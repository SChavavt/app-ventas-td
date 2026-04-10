
import streamlit as st
import streamlit.components.v1 as components
from streamlit.errors import StreamlitAPIException
import os
from pathlib import Path
from datetime import datetime, timedelta, date
import json
import base64
import uuid
import pandas as pd
import pdfplumber
from openpyxl import load_workbook
import unicodedata
from io import BytesIO
import time
import socket
import re
import gspread
import html
from typing import Dict, List, Optional
from difflib import SequenceMatcher
from urllib.parse import quote, urlsplit, urlunsplit, urlparse, unquote
from urllib.request import Request, urlopen
from urllib.error import URLError, HTTPError
from http.client import InvalidURL
from oauth2client.service_account import ServiceAccountCredentials
from pytz import timezone
from gspread.utils import rowcol_to_a1
from gspread.exceptions import APIError


# NEW: Import boto3 for AWS S3
import boto3

# --- STREAMLIT CONFIGURATION ---
st.set_page_config(page_title="App Vendedores TD", layout="wide")

REFRESH_COOLDOWN = 60
PENDING_SUBMISSION_RETRY_SECONDS = 8
PENDING_SUBMISSION_MAX_RETRY_SECONDS = 60
PENDING_SUBMISSIONS_DIR = Path(".pedido_retry_cache")
S3_UPLOAD_MAX_RETRIES = 4
S3_UPLOAD_BASE_DELAY_SECONDS = 1.2
CONNECTION_STATUS_TTL_SECONDS = 20
PEDIDO_STATUS_MAX_AGE_SECONDS = 180


TAB1_PRESERVED_STATE_KEYS: set[str] = {
    "last_selected_vendedor",
    "current_tab_index",
    "tipo_envio_selector_global",
}


TAB1_FORM_STATE_KEYS_TO_CLEAR: set[str] = {
    "registrar_nota_venta_checkbox",
    "registro_cliente",
    "numero_cliente_rfc",
    "tipo_envio_original",
    "estatus_factura_origen",
    "subtipo_local_selector",
    "folio_factura_error_input",
    "nota_venta_input",
    "motivo_nota_venta_input",
    "folio_factura_input",
    "fecha_entrega_input",
    "comentario_detallado",
    "direccion_guia_retorno_foraneo",
    "resultado_esperado",
    "material_devuelto",
    "monto_devuelto",
    "area_responsable",
    "nombre_responsable",
    "motivo_detallado",
    "g_resultado_esperado",
    "g_descripcion_falla",
    "g_piezas_afectadas",
    "g_monto_estimado",
    "g_area_responsable",
    "g_nombre_responsable",
    "g_numero_serie",
    "g_fecha_compra",
    "direccion_guia_retorno",
    "direccion_envio_destino",
    "pedido_adjuntos",
    "comprobante_cliente",
    "estado_pago",
    "comprobante_uploader_final",
    "fecha_pago_input",
    "forma_pago_input",
    "monto_pago_input",
    "terminal_input",
    "banco_destino_input",
    "referencia_pago_input",
    "local_route_recibe",
    "local_route_calle_no",
    "local_route_tipo_inmueble",
    "local_route_acceso_privada",
    "local_route_municipio",
    "local_route_telefonos",
    "local_route_interior",
    "local_route_colonia",
    "local_route_cp",
    "local_route_forma_pago",
    "local_route_total_factura",
    "local_route_adeudo_anterior",
    "local_route_referencias",
    "local_route_hora_entrega_manual",
    "local_route_hora_entrega_selector",
    "local_route_hora_entrega_custom",
    "local_route_confirmed_payload",
    "local_route_confirmed_at",
    "local_route_generated_file",
    "local_route_generated_filename",
    "local_route_generated_at",
    "local_route_post_confirm_notice",
    "local_route_client_search",
    "local_route_selected_history_label",
    "local_route_selected_history_row",
}

TAB1_WARNING_FORM_BACKUP_KEY = "tab1_warning_form_backup"
TAB1_VENDOR_EMPTY_OPTION = ""

TAB1_RESTORE_EXCLUDED_KEYS: set[str] = {
    "pedido_adjuntos",
    "comprobante_cliente",
    "comprobante_uploader_final",
}

TAB1_SCROLL_RESTORE_FLAG_KEY = "tab1_restore_scroll_after_submit"
TAB1_FEEDBACK_ANCHOR_ID = "tab1-pedido-feedback-anchor"
TAB1_FORM_NONCE_KEY = "tab1_form_nonce"
TAB2_LOADING_MESSAGE_KEY = "tab2_modification_loading_message"
LOCAL_ROUTE_CONFIRMED_PAYLOAD_KEY = "local_route_confirmed_payload"
LOCAL_ROUTE_CONFIRMED_AT_KEY = "local_route_confirmed_at"
LOCAL_ROUTE_GENERATED_FILE_KEY = "local_route_generated_file"
LOCAL_ROUTE_GENERATED_FILENAME_KEY = "local_route_generated_filename"
LOCAL_ROUTE_GENERATED_AT_KEY = "local_route_generated_at"
LOCAL_ROUTE_POST_CONFIRM_NOTICE_KEY = "local_route_post_confirm_notice"
LOCAL_ROUTE_HOUR_AUTOMATIC_OPTION = "🧠 Automático por turno"



USUARIOS_VALIDOS = [
    "DIANASOFIA47",
    "ALEJANDRO38",
    "ANA45",
    "CURSOS92",
    "CASSANDRA93",
    "CECILIA94",
    "DANIELA73",
    "CARITO82",
    "GLORIA53",
    "JUAN24",
    "JOSE31",
    "KAREN58",
    "PAULINA57",
    "RUBEN67",
    "ROBERTO51",
    "FRANKO95",
]

VENDEDOR_NOMBRE_POR_ID = {
    "DIANASOFIA47": "DIANA SOFIA",
    "ALEJANDRO38": "ALEJANDRO RODRIGUEZ",
    "ANA45": "ANA KAREN ORTEGA MAHUAD",
    "CURSOS92": "CURSOS Y EVENTOS",
    "CASSANDRA93": "CASSANDRA MIROSLAVA",
    "CECILIA94": "CECILIA SEPULVEDA",
    "DANIELA73": "DANIELA LOPEZ RAMIREZ",
    "CARITO82": "GRISELDA CAROLINA SANCHEZ GARCIA",
    "GLORIA53": "GLORIA MICHELLE GARCIA TORRES",
    "JUAN24": "JUAN CASTILLEJO",
    "JOSE31": "JOSE CORTES",
    "KAREN58": "KAREN JAQUELINE",
    "PAULINA57": "PAULINA TREJO",
    "RUBEN67": "RUBEN",
    "ROBERTO51": "DISTRIBUCION Y UNIVERSIDADES",
    "FRANKO95": "FRANKO",
}

TAB1_LOCAL_CDMX_DISABLE_ROUTE_IDS = {
    "JUAN24",
    "RUBEN67",
    "FRANKO95",
}



def normalize_case_text(value, placeholder: str = "N/A") -> str:
    """Return a clean string for optional case fields."""
    if value is None:
        return placeholder
    if isinstance(value, str):
        cleaned = value.strip()
        return cleaned if cleaned else placeholder
    return str(value)


def normalize_case_amount(value, placeholder: str = "N/A") -> str:
    """Format optional numeric fields, falling back to ``placeholder`` if empty."""
    try:
        amount = float(value)
    except (TypeError, ValueError):
        return placeholder
    return f"{amount:.2f}" if amount > 0 else placeholder


def format_estado_entrega(value) -> str:
    """Return delivery status text for local orders."""
    if value is None:
        return "Sin info de entrega"
    if isinstance(value, str):
        cleaned = value.strip()
        return cleaned if cleaned else "Sin info de entrega"
    if pd.isna(value):
        return "Sin info de entrega"
    cleaned = str(value).strip()
    return cleaned if cleaned else "Sin info de entrega"


def format_currency_for_route_sheet(value) -> str:
    """Return currency text like ``$7,400.00`` for the local route sheet."""
    try:
        amount = float(value)
    except (TypeError, ValueError):
        amount = 0.0
    return f"${amount:,.2f}"


def get_local_delivery_slot(turno_local: str) -> str:
    """Map local shift names to route sheet delivery time windows."""
    turno_normalizado = str(turno_local or "").strip()
    if turno_normalizado == "☀️ Local Mañana":
        return "9:00 AM a 2:00 PM"
    if turno_normalizado == "🌙 Local Tarde":
        return "3:00 PM a 7:00 PM"
    return turno_normalizado or "POR DEFINIR"


def resolve_local_delivery_slot(turno_local: str, hora_entrega_manual: str = "") -> str:
    """Resolve delivery slot using manual text first, then fallback to automatic mapping."""
    hora_manual_limpia = str(hora_entrega_manual or "").strip()
    if hora_manual_limpia:
        return hora_manual_limpia
    if str(turno_local or "").strip() == "🏙️ Local Mty":
        return "POR DEFINIR"
    return get_local_delivery_slot(turno_local)


def get_subtipo_local_excel_value(subtipo_local: str) -> str:
    """Normalize local shift label for Excel persistence."""
    turno_normalizado = str(subtipo_local or "").strip()
    if turno_normalizado == "🏙️ Local Mty":
        return "🌤️ Local Día"
    return turno_normalizado


LOCAL_TURNO_CDMX_IDS = {"RUBEN67", "JUAN24", "FRANKO95"}
TAB1_DUAL_VIEW_IDS = {"ALEJANDRO38", "CECILIA94"}


def get_local_shift_options(id_vendedor: str | None = None, force_cdmx_view: bool = False) -> list[str]:
    """Return local shift options, enabling CDMX only for approved users."""
    id_vendedor_normalizado = normalize_vendedor_id(id_vendedor or "")
    if force_cdmx_view or id_vendedor_normalizado in LOCAL_TURNO_CDMX_IDS:
        return ["🏙️ Local Mty", "🌆 Local CDMX", "🎓 Recoge en Aula"]

    opciones = ["☀️ Local Mañana", "🌙 Local Tarde", "🌵 Saltillo", "📦 Pasa a Bodega"]
    return opciones


def get_weekday_name_es(delivery_date: date) -> str:
    """Return the weekday in uppercase Spanish for the route sheet."""
    dias = [
        "LUNES",
        "MARTES",
        "MIERCOLES",
        "JUEVES",
        "VIERNES",
        "SABADO",
        "DOMINGO",
    ]
    if not isinstance(delivery_date, date):
        return ""
    return dias[delivery_date.weekday()]


def build_local_route_sheet(template_path: Path, payload: Dict[str, object]) -> BytesIO:
    """Fill the local delivery Excel template and return it in memory."""
    workbook = load_workbook(template_path)
    worksheet = workbook[workbook.sheetnames[0]]

    worksheet["B2"] = payload.get("fecha", "")
    worksheet["F2"] = payload.get("dia_entrega", "")
    worksheet["B3"] = payload.get("cliente", "")
    worksheet["F3"] = payload.get("hora_entrega", "")
    worksheet["B4"] = payload.get("recibe", "")
    worksheet["E5"] = payload.get("referencias", "")
    worksheet["B5"] = payload.get("calle_no", "")
    worksheet["B6"] = payload.get("tipo_inmueble", "")
    worksheet["D6"] = payload.get("interior", "")
    worksheet["B7"] = payload.get("acceso_privada", "")
    worksheet["D7"] = payload.get("colonia", "")
    worksheet["B8"] = payload.get("municipio", "")
    worksheet["D8"] = payload.get("cp", "")
    worksheet["B9"] = payload.get("telefonos", "")
    worksheet["D10"] = payload.get("estado_pago", "")
    worksheet["D11"] = payload.get("forma_pago", "")
    worksheet["D12"] = payload.get("vendedor", "")
    worksheet["G10"] = payload.get("total_factura", "")
    worksheet["G11"] = payload.get("adeudo_anterior", "")
    worksheet["G12"] = payload.get("gran_total", "")

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def slugify_local_route_client_name(cliente: str, fallback: str = "CLIENTE") -> str:
    """Return an uppercase ASCII filename-safe slug based on the client name."""
    normalized = unicodedata.normalize("NFKD", str(cliente or "").strip())
    ascii_name = normalized.encode("ascii", "ignore").decode("ascii")
    cleaned = re.sub(r"[^A-Za-z0-9]+", "_", ascii_name).strip("_")
    return (cleaned or fallback).upper()


def build_local_route_payload(
    fecha_entrega,
    registro_cliente: str,
    subtipo_local: str,
    recibe: str,
    referencias_hoja_ruta: str,
    calle_no: str,
    tipo_inmueble: str,
    interior: str,
    acceso_privada: str,
    colonia: str,
    municipio: str,
    cp: str,
    telefonos: str,
    estado_pago: str,
    forma_pago: str,
    vendedor: str,
    total_factura,
    adeudo_anterior,
    folio: str,
    hora_entrega_manual: str = "",
) -> Dict[str, str]:
    """Build the serialized payload used by the local route Excel and summary UI."""
    route_total_amount = float(total_factura or 0.0) + float(adeudo_anterior or 0.0)
    route_references = referencias_hoja_ruta.strip()

    return {
        "fecha": fecha_entrega.strftime('%Y-%m-%d') if isinstance(fecha_entrega, date) else "",
        "dia_entrega": get_weekday_name_es(fecha_entrega),
        "cliente": registro_cliente.strip(),
        "subtipo_local": subtipo_local.strip(),
        "hora_entrega": resolve_local_delivery_slot(subtipo_local, hora_entrega_manual),
        "recibe": recibe.strip(),
        "referencias": route_references,
        "calle_no": calle_no.strip(),
        "tipo_inmueble": tipo_inmueble.strip(),
        "interior": interior.strip(),
        "acceso_privada": acceso_privada.strip(),
        "colonia": colonia.strip(),
        "municipio": municipio.strip(),
        "cp": cp.strip(),
        "telefonos": telefonos.strip(),
        "estado_pago": (estado_pago or "NO PAGADO").strip(),
        "forma_pago": forma_pago.strip() or "TRANSFERENCIA",
        "vendedor": vendedor.strip(),
        "total_factura": format_currency_for_route_sheet(total_factura),
        "adeudo_anterior": format_currency_for_route_sheet(adeudo_anterior),
        "gran_total": format_currency_for_route_sheet(route_total_amount),
        "folio": folio.strip(),
    }


def get_local_route_missing_fields(payload: Dict[str, str]) -> List[str]:
    """Return required route fields that are still missing."""
    missing_fields = []
    local_delivery_mode = str(payload.get("subtipo_local", "") or "").strip()
    requires_address = local_delivery_mode != "📦 Pasa a Bodega"

    if not payload.get("cliente"):
        missing_fields.append("Cliente")
    if not payload.get("folio"):
        missing_fields.append("Folio de Factura")
    if requires_address and not payload.get("calle_no"):
        missing_fields.append("Calle y No.")
    return missing_fields


def build_local_route_file_from_payload(
    template_path: Path,
    payload: Dict[str, object],
) -> tuple[Optional[dict[str, str]], str]:
    """Return the generated local route file payload and filename."""
    if not template_path.exists():
        return None, ""

    generated_route_file = build_local_route_sheet(template_path, payload)
    generated_route_bytes = generated_route_file.getvalue()
    route_client_slug = slugify_local_route_client_name(payload.get("cliente", ""))
    route_filename = f"{route_client_slug}.xlsx"
    route_file_payload = {
        "name": route_filename,
        "content_b64": base64.b64encode(generated_route_bytes).decode("utf-8"),
    }
    return route_file_payload, route_filename


def parse_sheet_row_number(value) -> Optional[int]:
    """Return a normalized Google Sheet row number or ``None`` if missing."""
    if value is None:
        return None
    if isinstance(value, (int, float)) and not pd.isna(value):
        candidate = int(value)
        return candidate if candidate > 0 else None
    if isinstance(value, str):
        cleaned = value.strip()
        if not cleaned:
            return None
        try:
            candidate = int(float(cleaned))
        except ValueError:
            return None
        return candidate if candidate > 0 else None
    try:
        if pd.isna(value):
            return None
    except TypeError:
        return None
    try:
        candidate = int(float(value))
    except (TypeError, ValueError):
        return None
    return candidate if candidate > 0 else None


def normalize_vendedor_id(value: object) -> str:
    """Normalize vendor IDs to compare them safely across sheets/sessions."""
    return str(value or "").strip().upper()


def get_session_vendedor_name() -> str:
    """Return the logged-in vendor display name mapped from the session vendor ID."""
    return VENDEDOR_NOMBRE_POR_ID.get(
        normalize_vendedor_id(st.session_state.get("id_vendedor", "")),
        "",
    )


def ensure_selectbox_vendor_default(key: str, options: list[str], fallback: str = "Todos") -> int:
    """Preselect the current session vendor when it exists in the available options."""
    if not options:
        return 0

    session_vendor = get_session_vendedor_name()
    default_value = session_vendor if session_vendor in options else fallback
    if default_value not in options:
        default_value = options[0]

    current_value = st.session_state.get(key)
    if current_value not in options:
        st.session_state[key] = default_value
        current_value = default_value

    try:
        return options.index(current_value)
    except ValueError:
        return 0


def is_empty_folio(value: object) -> bool:
    """True cuando el folio está vacío o no tiene valor utilizable."""
    cleaned = str(value or "").strip()
    return cleaned == "" or cleaned.lower() in {"nan", "none"}


def clean_folio_for_ui(value: object) -> str:
    """Muestra el folio limpio en UI quitando prefijo de captura tardía (*)."""
    folio = str(value or "").strip()
    return folio[1:].strip() if folio.startswith("*") else folio


def is_devolucion_case_row(row: pd.Series) -> bool:
    """Detecta devoluciones usando el campo real disponible en la hoja."""
    for key in ("Tipo_Caso", "Tipo_Envio"):
        raw = str(row.get(key, "") or "").strip().lower()
        if "devoluci" in raw:
            return True
    return False


class CachedUploadedFile(BytesIO):
    """Archivo en memoria con atributo ``name`` para reutilizar carga a S3."""

    def __init__(self, filename: str, content: bytes):
        super().__init__(content)
        self.name = filename


def get_pending_submission_key() -> str:
    """Genera una llave por vendedor para cachear un pedido pendiente."""
    return (
        normalize_vendedor_id(st.session_state.get("id_vendedor", ""))
        or normalize_vendedor_id(st.session_state.get("last_selected_vendedor", ""))
        or "GLOBAL"
    )


def scroll_to_tab1_feedback_section() -> None:
    """Lleva la vista a la sección de mensajes del registro de pedidos."""
    components.html(
        f"""
        <script>
        (function() {{
            const parentWindow = window.parent;
            const anchorId = {json.dumps(TAB1_FEEDBACK_ANCHOR_ID)};

            function scrollToFeedback() {{
                const anchor = parentWindow.document.getElementById(anchorId);
                if (!anchor) return false;
                anchor.scrollIntoView({{ behavior: 'auto', block: 'start' }});
                return true;
            }}

            if (!scrollToFeedback()) {{
                setTimeout(scrollToFeedback, 150);
            }}
        }})();
        </script>
        """,
        height=0,
    )


def build_submission_identity() -> tuple[str, str, str]:
    """Construye un ID estable de pedido para reintentos y su prefijo de adjuntos."""
    zona_mexico = timezone("America/Mexico_City")
    now = datetime.now(zona_mexico)
    pedido_id = f"PED-{now.strftime('%Y%m%d%H%M%S')}-{uuid.uuid4().hex[:8].upper()}"
    hora_registro = now.strftime('%Y-%m-%d %H:%M:%S')
    s3_prefix = f"adjuntos_pedidos/{pedido_id}/"
    return pedido_id, hora_registro, s3_prefix


def _pending_submission_paths(cache_key: str) -> tuple[Path, Path]:
    cache_dir = PENDING_SUBMISSIONS_DIR / cache_key
    return cache_dir, cache_dir / "payload.json"


def _serialize_uploaded_files(files) -> list[dict]:
    serialized: list[dict] = []
    for file_obj in files or []:
        file_obj.seek(0)
        content = file_obj.read()
        file_obj.seek(0)
        serialized.append(
            {
                "name": file_obj.name,
                "content_b64": base64.b64encode(content).decode("utf-8"),
            }
        )
    return serialized


def _deserialize_uploaded_files(files_data: list[dict] | None):
    restored = []
    for item in files_data or []:
        try:
            content = base64.b64decode(item.get("content_b64", ""))
        except Exception:
            continue
        restored.append(CachedUploadedFile(item.get("name", "archivo.bin"), content))
    return restored


def save_pending_submission(cache_key: str, payload: dict, attempts: int = 0, next_retry_at: float = 0.0) -> None:
    """Guarda el payload de envío para reintentos automáticos."""
    cache_dir, payload_path = _pending_submission_paths(cache_key)
    cache_dir.mkdir(parents=True, exist_ok=True)
    record = {
        "cache_key": cache_key,
        "saved_at": time.time(),
        "attempts": attempts,
        "next_retry_at": next_retry_at,
        "payload": payload,
    }
    payload_path.write_text(json.dumps(record, ensure_ascii=False), encoding="utf-8")


def load_pending_submission(cache_key: str) -> Optional[dict]:
    """Carga un pedido pendiente desde disco."""
    _, payload_path = _pending_submission_paths(cache_key)
    if not payload_path.exists():
        return None
    try:
        return json.loads(payload_path.read_text(encoding="utf-8"))
    except Exception:
        return None


def clear_pending_submission(cache_key: str) -> None:
    """Elimina pedido pendiente guardado para el vendedor actual."""
    cache_dir, payload_path = _pending_submission_paths(cache_key)
    if payload_path.exists():
        payload_path.unlink(missing_ok=True)
    if cache_dir.exists():
        try:
            cache_dir.rmdir()
        except OSError:
            pass


def schedule_pending_submission_retry(cache_key: str, delay_seconds: int = PENDING_SUBMISSION_RETRY_SECONDS) -> None:
    """Agenda reintento automático con backoff para evitar saturar APIs externas."""
    pending = load_pending_submission(cache_key)
    if not pending:
        return
    attempts = int(pending.get("attempts", 0) or 0) + 1
    retry_seconds = min(
        PENDING_SUBMISSION_MAX_RETRY_SECONDS,
        max(1, delay_seconds) * (2 ** max(0, attempts - 1)),
    )
    save_pending_submission(
        cache_key,
        pending.get("payload", {}),
        attempts=attempts,
        next_retry_at=time.time() + retry_seconds,
    )


@st.cache_data(ttl=60)
def obtener_resumen_guias_vendedor(id_vendedor_norm: str, refresh_token: float | None = None) -> dict:
    """Obtiene resumen de guías cargadas para mostrar aviso rápido en encabezado."""
    _ = refresh_token
    if not id_vendedor_norm:
        return {"total": 0, "clientes": [], "keys": []}

    try:
        ws_ped = get_worksheet_operativa(refresh_token)
        df_ped = pd.DataFrame(ws_ped.get_all_records())
    except Exception:
        df_ped = pd.DataFrame()

    try:
        ws_casos = get_worksheet_casos_especiales()
        df_casos = pd.DataFrame(ws_casos.get_all_records())
    except Exception:
        df_casos = pd.DataFrame()

    for col in ["id_vendedor", "Adjuntos_Guia", "Cliente", "ID_Pedido", "Folio_Factura", "Completados_Limpiado"]:
        if col not in df_ped.columns:
            df_ped[col] = ""

    for col in ["id_vendedor", "Hoja_Ruta_Mensajero", "Adjuntos_Guia", "Cliente", "ID_Pedido", "Folio_Factura", "Completados_Limpiado"]:
        if col not in df_casos.columns:
            df_casos[col] = ""

    df_ped = df_ped[
        (df_ped["id_vendedor"].apply(normalize_vendedor_id) == id_vendedor_norm)
        & (df_ped["Adjuntos_Guia"].astype(str).str.strip() != "")
        & (df_ped["Completados_Limpiado"].fillna("").astype(str).str.strip() == "")
    ].copy()

    df_casos = df_casos[
        (df_casos["id_vendedor"].apply(normalize_vendedor_id) == id_vendedor_norm)
        & (df_casos["Hoja_Ruta_Mensajero"].astype(str).str.strip() != "")
        & (df_casos["Completados_Limpiado"].fillna("").astype(str).str.strip() == "")
    ].copy()

    clientes = []
    keys = []

    for _, row in df_ped.iterrows():
        cliente = str(row.get("Cliente", "")).strip()
        if cliente:
            clientes.append(cliente)
        pedido_ref = str(row.get("ID_Pedido", "")).strip() or str(row.get("Folio_Factura", "")).strip()
        guia_ref = str(row.get("Adjuntos_Guia", "")).strip()
        if pedido_ref and guia_ref:
            keys.append(f"{SHEET_PEDIDOS_OPERATIVOS}::{pedido_ref}::{guia_ref}")

    for _, row in df_casos.iterrows():
        cliente = str(row.get("Cliente", "")).strip()
        if cliente:
            clientes.append(cliente)
        pedido_ref = str(row.get("ID_Pedido", "")).strip() or str(row.get("Folio_Factura", "")).strip()
        guia_ref = str(row.get("Hoja_Ruta_Mensajero", "")).strip()
        if pedido_ref and guia_ref:
            keys.append(f"casos_especiales::{pedido_ref}::{guia_ref}")

    return {
        "total": int(len(df_ped) + len(df_casos)),
        "clientes": list(dict.fromkeys(clientes)),
        "keys": sorted(set(keys)),
    }


def load_sheet_records_with_row_numbers(worksheet):
    """Return DataFrame rows with their real Google Sheet indices preserved."""

    try:
        all_values = worksheet.get_all_values()
    except Exception:
        return pd.DataFrame(), []

    if not all_values:
        return pd.DataFrame(), []

    headers_raw = all_values[0]
    headers = [str(h).strip() for h in headers_raw]

    max_columns = len(headers)
    records: List[Dict[str, str]] = []
    row_numbers: List[int] = []

    for row_index, row_values in enumerate(all_values[1:], start=2):
        normalized_row = list(row_values[:max_columns])
        if len(normalized_row) < max_columns:
            normalized_row.extend([""] * (max_columns - len(normalized_row)))

        if not any(str(cell).strip() for cell in normalized_row):
            continue

        record = {
            headers[col_idx]: normalized_row[col_idx] if col_idx < len(normalized_row) else ""
            for col_idx in range(max_columns)
        }
        records.append(record)
        row_numbers.append(row_index)

    df_records = pd.DataFrame(records)
    if df_records.empty:
        return df_records, headers

    df_records.insert(0, "Sheet_Row_Number", row_numbers)
    return df_records, headers


@st.cache_data(ttl=90)
def get_tab3_pending_comprobante_dataset(
    refresh_token: float | None = None,
) -> tuple[pd.DataFrame, dict[str, list[str]]]:
    """Carga y cachea los datos necesarios para la pestaña de comprobantes."""
    _ = refresh_token

    dataframes_comprobante: list[pd.DataFrame] = []
    headers_by_source: dict[str, list[str]] = {}
    source_getters = [
        (SHEET_PEDIDOS_HISTORICOS, get_worksheet_historico),
        (SHEET_PEDIDOS_OPERATIVOS, get_worksheet_operativa),
    ]

    for source_name, getter in source_getters:
        worksheet_source = getter()
        if worksheet_source is None:
            continue

        ws_df, ws_headers = load_sheet_records_with_row_numbers(worksheet_source)
        if not ws_headers:
            continue

        headers_by_source[source_name] = ws_headers

        if ws_df.empty:
            continue

        ws_df["Fuente"] = source_name
        dataframes_comprobante.append(ws_df)

    if not dataframes_comprobante:
        return pd.DataFrame(), headers_by_source

    df_pedidos_comprobante = pd.concat(dataframes_comprobante, ignore_index=True)

    required_columns = [
        "Adjuntos_Guia",
        "Adjuntos",
        "Estado_Pago",
        "Vendedor_Registro",
        "ID_Pedido",
        "Cliente",
        "Folio_Factura",
    ]
    for col_name in required_columns:
        if col_name not in df_pedidos_comprobante.columns:
            df_pedidos_comprobante[col_name] = ""

    df_pedidos_comprobante["Folio_Factura"] = (
        df_pedidos_comprobante["Folio_Factura"].astype(str).replace("nan", "").str.strip()
    )

    vendedores_limpios = df_pedidos_comprobante["Vendedor_Registro"].astype(str).str.strip()
    vendedores_limpios = vendedores_limpios.replace({"nan": "", "None": ""})
    df_pedidos_comprobante["Vendedor_Registro"] = vendedores_limpios
    df_pedidos_comprobante.loc[
        df_pedidos_comprobante["Vendedor_Registro"] == "", "Vendedor_Registro"
    ] = "N/A"
    df_pedidos_comprobante.loc[
        ~df_pedidos_comprobante["Vendedor_Registro"].isin(VENDEDORES_LIST + ["N/A"]),
        "Vendedor_Registro",
    ] = "Otro/Desconocido"

    return df_pedidos_comprobante, headers_by_source


def extract_id_vendedor(data, placeholder: str = "N/A") -> str:
    """Return a readable vendor ID from heterogeneous row/dict structures."""

    if data is None:
        return placeholder

    candidate_keys = (
        "id_vendedor",
        "ID_Vendedor",
        "Id_Vendedor",
        "IDVendedor",
        "IdVendedor",
        "ID Vendedor",
        "Id Vendedor",
        "ID_VENDEDOR",
        "IDVENDEDOR",
    )

    for key in candidate_keys:
        value = None
        if hasattr(data, "get"):
            try:
                value = data.get(key)
            except Exception:
                value = None

        if value is None and isinstance(data, pd.Series) and key in data.index:
            value = data[key]

        if value is None:
            continue

        if isinstance(value, str):
            cleaned = value.strip()
            if cleaned and cleaned.lower() not in {"nan", "none"}:
                return cleaned
            continue

        try:
            if pd.isna(value):
                continue
        except Exception:
            pass

        cleaned = str(value).strip()
        if cleaned and cleaned.lower() not in {"nan", "none"}:
            return cleaned

    return placeholder


def extract_id_vendedor_mod(data, placeholder: str = "") -> str:
    """Return normalized modifier vendor IDs, handling multiple entries."""

    if data is None:
        return placeholder

    candidate_keys = (
        "id_vendedor_Mod",
        "ID_Vendedor_Mod",
        "Id_Vendedor_Mod",
        "IDVendedor_Mod",
        "ID_VENDEDOR_MOD",
    )

    raw_value = None
    for key in candidate_keys:
        value = None
        if hasattr(data, "get"):
            try:
                value = data.get(key)
            except Exception:
                value = None

        if value is None and isinstance(data, pd.Series) and key in data.index:
            value = data[key]

        if value is None:
            continue

        raw_value = value
        break

    if raw_value is None:
        return placeholder

    if isinstance(raw_value, str):
        tokens = [
            entry.strip()
            for entry in re.split(r"[;,\n]", raw_value)
            if entry and entry.strip()
        ]
    elif isinstance(raw_value, (list, tuple, set)):
        tokens = [str(entry).strip() for entry in raw_value if str(entry).strip()]
    else:
        normalized = str(raw_value).strip()
        tokens = [normalized] if normalized else []

    if not tokens:
        return placeholder

    unique_tokens: list[str] = []
    seen = set()
    for token in tokens:
        upper_token = token.upper()
        if upper_token not in seen:
            seen.add(upper_token)
            unique_tokens.append(upper_token)

    return ", ".join(unique_tokens) if unique_tokens else placeholder


def format_id_vendedor_with_mod(data, placeholder: str = "N/A") -> str:
    """Compose the display string for vendor IDs including modifiers."""

    id_principal = extract_id_vendedor(data, placeholder)
    id_modificador = extract_id_vendedor_mod(data, "")

    base_segment = f"🆔 **ID Vendedor:** `{id_principal}`"
    if id_modificador:
        base_segment += f"  |  🛠️ **ID Vendedor Mod:** `{id_modificador}`"

    return base_segment


def allow_refresh(key: str, container=st, cooldown: int = REFRESH_COOLDOWN) -> bool:
    """Rate-limit manual reloads to avoid hitting services too often."""
    now = time.time()
    last = st.session_state.get(key)
    if last and now - last < cooldown:
        container.warning("⚠️ Se recargó recientemente. Espera unos segundos.")
        return False
    st.session_state[key] = now
    return True


def clear_app_caches() -> None:
    """Reinicia las conexiones y datos cacheados para forzar una recarga completa."""
    st.cache_data.clear()

    # Limpiar solo funciones cacheadas que expongan `.clear()`.
    for cached_fn in (
        cargar_pedidos,
        get_google_sheets_client,
        get_worksheet_operativa,
        get_worksheet_historico,
        get_worksheet_clientes_locales,
        get_worksheet_zonas_remotas,
        get_s3_client,
    ):
        clear_fn = getattr(cached_fn, "clear", None)
        if callable(clear_fn):
            clear_fn()


def ensure_user_logged_in() -> str:
    """Muestra una pantalla de inicio de sesión simple y detiene la app hasta autenticar."""
    st.session_state.setdefault("id_vendedor", "")
    current_user = st.session_state.get("id_vendedor", "")

    if not current_user:
        usuario_param = st.query_params.get("usuario")
        if isinstance(usuario_param, (list, tuple)):
            usuario_param = usuario_param[0] if usuario_param else ""
        if usuario_param:
            candidate = str(usuario_param).strip().upper()
            if candidate and candidate in USUARIOS_VALIDOS:
                st.session_state["id_vendedor"] = candidate
                return candidate

    if current_user:
        return current_user

    st.markdown("## 🔐 Inicio de sesión")
    username_input = st.text_input("Usuario", key="login_usuario")

    if st.button("Ingresar", key="login_ingresar_btn"):
        candidate = username_input.strip()
        if candidate and candidate.upper() in USUARIOS_VALIDOS:
            normalized_candidate = candidate.upper()
            st.session_state["id_vendedor"] = normalized_candidate
            st.query_params["usuario"] = normalized_candidate
            st.rerun()
        else:
            st.error("❌ Usuario no válido. Verifica tu nombre y número.")

    st.stop()


def render_date_filter_controls(
    label: str,
    key_prefix: str,
    *,
    default_range_days: int = 7,
    recent_days_option: int | None = None,
    recent_days_label: str | None = None,
) -> tuple[date, date, bool, bool]:
    """Renderiza un control de fecha con opción de rango y devuelve la selección.

    Returns a tuple ``(fecha_inicio, fecha_fin, rango_activo, rango_valido)``.
    """

    use_range = st.checkbox(
        "🔁 Activar búsqueda por rango de fechas",
        key=f"{key_prefix}_usar_rango",
    )

    use_recent_days = False
    if recent_days_option is not None and recent_days_option > 0:
        recent_label = recent_days_label or f"Mostrar últimos {recent_days_option} días"
        recent_key = f"{key_prefix}_ultimos_{recent_days_option}_dias"

        if use_range and st.session_state.get(recent_key):
            st.session_state[recent_key] = False

        use_recent_days = st.checkbox(
            recent_label,
            key=recent_key,
            disabled=use_range,
        )

    if use_range:
        end_default = st.session_state.get(
            f"{key_prefix}_fecha_fin",
            datetime.now().date(),
        )
        start_default = st.session_state.get(
            f"{key_prefix}_fecha_inicio",
            end_default - timedelta(days=default_range_days),
        )

        if start_default > end_default:
            start_default = end_default

        start_date = st.date_input(
            "📅 Fecha inicial:",
            value=start_default,
            key=f"{key_prefix}_fecha_inicio",
        )
        end_date = st.date_input(
            "📅 Fecha final:",
            value=end_default if end_default >= start_date else start_date,
            key=f"{key_prefix}_fecha_fin",
        )

        is_valid = end_date >= start_date
        if not is_valid:
            st.error("La fecha final no puede ser anterior a la fecha inicial.")

        return start_date, end_date, True, is_valid

    if use_recent_days:
        end_date = datetime.now().date()
        start_date = end_date - timedelta(days=recent_days_option - 1)
        return start_date, end_date, False, True

    selected_date = st.date_input(
        label,
        value=st.session_state.get(
            f"{key_prefix}_fecha",
            datetime.now().date(),
        ),
        key=f"{key_prefix}_fecha",
    )

    return selected_date, selected_date, False, True


def build_vendor_filter_options(
    detected_values: list[object],
    *,
    include_all_option: bool = True,
) -> list[str]:
    """Construye opciones de vendedor usando catálogo oficial + valores extra detectados."""
    vendedores_detectados = sorted(
        {
            str(value).strip()
            for value in detected_values
            if str(value).strip() and str(value).strip().lower() not in {"none", "nan"}
        }
    )
    vendedores_extra = [v for v in vendedores_detectados if v not in VENDEDORES_LIST]
    opciones = VENDEDORES_LIST + vendedores_extra
    return (["Todos"] + opciones) if include_all_option else opciones


def render_lazy_tab_placeholder(tab_index: int, key_prefix: str, message: str) -> None:
    """Muestra aviso de pestaña diferida y botón para cargarla en la app."""
    st.caption(message)
    if st.button("🔄 Cargar esta pestaña ahora", key=f"{key_prefix}_load_now"):
        st.query_params.update({"tab": str(tab_index)})
        clear_app_caches()
        get_cached_connection_statuses.clear()
        st.rerun()


LAZY_TAB_MESSAGE = "ℹ️ Esta pestaña está inactiva. Presiona el botón '🔄 Cargar esta pestaña ahora' de abajo para activarla."


def reset_tab1_form_state(additional_preserved: dict[str, object] | None = None) -> None:
    """Elimina los valores capturados en el formulario principal, conservando envío y vendedor."""

    preserved_values = {
        key: st.session_state.get(key)
        for key in TAB1_PRESERVED_STATE_KEYS
    }

    if additional_preserved:
        preserved_values.update(additional_preserved)

    for key in TAB1_FORM_STATE_KEYS_TO_CLEAR:
        st.session_state.pop(key, None)

    for key, value in preserved_values.items():
        if value is None:
            continue

        # Evita sobrescribir el valor si ya existe y coincide
        if key in st.session_state and st.session_state[key] == value:
            continue

        st.session_state[key] = value

    st.session_state.pop(TAB1_WARNING_FORM_BACKUP_KEY, None)
    st.session_state[TAB1_FORM_NONCE_KEY] = int(st.session_state.get(TAB1_FORM_NONCE_KEY, 0) or 0) + 1


def backup_tab1_form_state_for_retry() -> None:
    """Respalda los valores del formulario de tab1 para restaurarlos tras una validación fallida."""

    st.session_state[TAB1_WARNING_FORM_BACKUP_KEY] = {
        key: st.session_state.get(key)
        for key in TAB1_FORM_STATE_KEYS_TO_CLEAR
        if key in st.session_state and key not in TAB1_RESTORE_EXCLUDED_KEYS
    }


def restore_tab1_form_state_for_retry() -> None:
    """Restaura valores del formulario respaldados cuando el envío no se pudo completar."""

    backup_values = st.session_state.pop(TAB1_WARNING_FORM_BACKUP_KEY, None)
    if not backup_values:
        return

    for key, value in backup_values.items():
        if key in TAB1_RESTORE_EXCLUDED_KEYS:
            continue
        try:
            st.session_state[key] = value
        except StreamlitAPIException:
            continue


def safe_batch_update(worksheet, data, retries: int = 5, base_delay: float = 1.0) -> None:
    """Realiza ``batch_update`` con reintentos ante errores de cuota."""
    for attempt in range(retries):
        try:
            worksheet.batch_update(data)
            return
        except APIError as e:
            status = getattr(getattr(e, "response", None), "status_code", None)
            if status == 429 and attempt < retries - 1:
                time.sleep(base_delay * (2 ** attempt))
            else:
                raise

# --- GOOGLE SHEETS CONFIGURATION ---
# Eliminamos la línea SERVICE_ACCOUNT_FILE ya que leeremos de secrets
GOOGLE_SHEET_ID = '1aWkSelodaz0nWfQx7FZAysGnIYGQFJxAN7RO3YgCiZY'
SHEET_PEDIDOS_OPERATIVOS = "data_pedidos"
SHEET_PEDIDOS_HISTORICOS = "datos_pedidos"
SHEET_CLIENTES_LOCALES = "Clientes_Locales"
SHEET_ZONAS_REMOTAS = "Zonas_Remotas"
CLIENTES_LOCALES_HEADERS = [
    "Cliente",
    "Recibe",
    "CalleyNumero",
    "Tipo_Inmueble",
    "Acceso_Privada",
    "Municipio",
    "Tels",
    "Interior",
    "Col",
    "C_P.",
    "Referencias",
]

def build_gspread_client():
    credentials_json_str = st.secrets["google_credentials"]
    creds_dict = json.loads(credentials_json_str)
    if "private_key" in creds_dict:
        creds_dict["private_key"] = creds_dict["private_key"].replace("\\n", "\n").strip()
    scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
    creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, scope)
    return gspread.authorize(creds)


def format_gspread_api_error(error: APIError) -> str:
    """Resume errores frecuentes de Google Sheets sin exponer secretos."""
    status = getattr(getattr(error, "response", None), "status_code", None)
    detail = str(error)
    if status == 403:
        return (
            "Permiso denegado al abrir Google Sheets. "
            "Verifica que la service account tenga acceso al archivo y que las APIs de Google Sheets/Drive estén habilitadas."
        )
    if status == 404:
        return (
            "No se encontró el Google Sheet configurado. "
            "Revisa que GOOGLE_SHEET_ID sea correcto y que el archivo siga existiendo."
        )
    if status == 429 or "RESOURCE_EXHAUSTED" in detail:
        return "La cuota de Google Sheets se agotó temporalmente. Intenta nuevamente en unos minutos."
    return f"Google Sheets devolvió un error inesperado ({status or 'sin código'}): {detail}"


@st.cache_resource
def get_google_sheets_client(refresh_token: float | None = None):
    _ = refresh_token
    def try_get_client():
        credentials_json_str = st.secrets["google_credentials"]
        creds_dict = json.loads(credentials_json_str)
        if "private_key" in creds_dict:
            creds_dict["private_key"] = creds_dict["private_key"].replace("\\n", "\n").strip()
        scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
        creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, scope)
        return gspread.authorize(creds)

    max_attempts = 5
    for attempt in range(max_attempts):
        try:
            client = try_get_client()
            _ = client.open_by_key(GOOGLE_SHEET_ID)
            st.session_state.pop("gsheet_error", None)
            return client
        except gspread.exceptions.APIError as e:
            status = getattr(getattr(e, "response", None), "status_code", None)
            if status == 429 or "RESOURCE_EXHAUSTED" in str(e):
                time.sleep(2 ** attempt)
                continue
            st.session_state["gsheet_error"] = (
                f"❌ Error al conectar con Google Sheets: {format_gspread_api_error(e)}"
            )
            return None
        except Exception as e:
            st.session_state["gsheet_error"] = f"❌ Error al conectar con Google Sheets: {e}"
            return None

    st.session_state["gsheet_error"] = st.session_state.get(
        "gsheet_error", "❌ No se pudo conectar con Google Sheets."
    )
    return None


def open_google_sheet(client):
    """Abre el spreadsheet principal y captura errores transitorios/permisos."""
    try:
        spreadsheet = client.open_by_key(GOOGLE_SHEET_ID)
        st.session_state.pop("gsheet_error", None)
        return spreadsheet
    except gspread.exceptions.APIError as e:
        st.session_state["gsheet_error"] = f"❌ Error al abrir Google Sheets: {format_gspread_api_error(e)}"
        return None
    except Exception as e:
        st.session_state["gsheet_error"] = f"❌ Error al abrir Google Sheets: {e}"
        return None

@st.cache_resource
def get_worksheet_operativa(refresh_token: float | None = None):
    client = get_google_sheets_client(refresh_token)
    if client is None:
        return None
    spreadsheet = open_google_sheet(client)
    if spreadsheet is None:
        return None
    return spreadsheet.worksheet(SHEET_PEDIDOS_OPERATIVOS)


@st.cache_resource
def get_worksheet_historico(refresh_token: float | None = None):
    client = get_google_sheets_client(refresh_token)
    if client is None:
        return None
    spreadsheet = open_google_sheet(client)
    if spreadsheet is None:
        return None
    return spreadsheet.worksheet(SHEET_PEDIDOS_HISTORICOS)


def get_worksheet(refresh_token: float | None = None):
    """Compatibilidad para tabs legadas que aún leen histórico."""
    return get_worksheet_historico(refresh_token)

@st.cache_resource
def get_worksheet_clientes_locales(refresh_token: float | None = None):
    client = get_google_sheets_client(refresh_token)
    if client is None:
        return None
    spreadsheet = open_google_sheet(client)
    if spreadsheet is None:
        return None
    return spreadsheet.worksheet(SHEET_CLIENTES_LOCALES)


@st.cache_resource
def get_worksheet_zonas_remotas(refresh_token: float | None = None):
    client = get_google_sheets_client(refresh_token)
    if client is None:
        return None
    spreadsheet = open_google_sheet(client)
    if spreadsheet is None:
        return None
    return spreadsheet.worksheet(SHEET_ZONAS_REMOTAS)

def get_worksheet_casos_especiales():
    client = build_gspread_client()
    spreadsheet = client.open_by_key(GOOGLE_SHEET_ID)
    return spreadsheet.worksheet("casos_especiales")


@st.cache_data(ttl=300, show_spinner=False)
def get_remote_postal_codes() -> set[str]:
    """Obtiene los códigos postales de la hoja Zonas_Remotas como strings normalizados."""
    worksheet = get_worksheet_zonas_remotas(st.session_state.get("remote_zones_refresh_token"))
    if worksheet is None:
        return set()

    try:
        valores = worksheet.col_values(1)[1:]  # omite encabezado
    except Exception:
        return set()

    codigos: set[str] = set()
    for value in valores:
        digits = re.sub(r"\D", "", str(value or "").strip())
        if digits:
            codigos.add(digits.zfill(5) if len(digits) <= 5 else digits)
    return codigos


def normalize_client_history_text(value: object) -> str:
    """Normaliza texto para búsquedas flexibles tolerando acentos y captura irregular."""
    normalized = unicodedata.normalize("NFKD", str(value or "").strip().lower())
    ascii_text = normalized.encode("ascii", "ignore").decode("ascii")
    ascii_text = re.sub(r"[^a-z0-9 ]+", " ", ascii_text)
    return re.sub(r"\s+", " ", ascii_text).strip()


def normalize_phone_for_sheets(value: object) -> str:
    """Normaliza teléfonos para que Google Sheets los trate siempre como texto plano."""
    phone = str(value or "").strip()
    if not phone:
        return ""
    if phone.startswith("'"):
        return phone
    if phone.startswith("+"):
        # El prefijo "'" fuerza texto en Sheets y evita que "+" se interprete como fórmula.
        return f"'{phone}"
    return phone


def display_phone_from_sheets(value: object) -> str:
    """Convierte el valor guardado en Sheets a una representación amigable en UI."""
    phone = str(value or "").strip()
    if phone.startswith("'") and len(phone) > 1 and phone[1] == "+":
        return phone[1:]
    return phone


def build_clientes_locales_record_from_form() -> dict[str, str]:
    """Construye un registro de Clientes_Locales usando el estado actual del formulario."""
    return {
        "Cliente": str(st.session_state.get("registro_cliente", "") or "").strip(),
        "Recibe": str(st.session_state.get("local_route_recibe", "") or "").strip(),
        "CalleyNumero": str(st.session_state.get("local_route_calle_no", "") or "").strip(),
        "Tipo_Inmueble": str(st.session_state.get("local_route_tipo_inmueble", "") or "").strip(),
        "Acceso_Privada": str(st.session_state.get("local_route_acceso_privada", "") or "").strip(),
        "Municipio": str(st.session_state.get("local_route_municipio", "") or "").strip(),
        "Tels": normalize_phone_for_sheets(st.session_state.get("local_route_telefonos", "")),
        "Interior": str(st.session_state.get("local_route_interior", "") or "").strip(),
        "Col": str(st.session_state.get("local_route_colonia", "") or "").strip(),
        "C_P.": str(st.session_state.get("local_route_cp", "") or "").strip(),
        "Referencias": str(st.session_state.get("local_route_referencias", "") or "").strip(),
    }


def ensure_clientes_locales_headers(worksheet) -> list[str]:
    """Garantiza que Clientes_Locales tenga los encabezados exactos requeridos."""
    if worksheet is None:
        raise Exception("No se pudo abrir la hoja Clientes_Locales.")
    last_header_cell = rowcol_to_a1(1, len(CLIENTES_LOCALES_HEADERS))
    headers_range = f"A1:{last_header_cell}"
    current_headers = worksheet.row_values(1)
    if current_headers != CLIENTES_LOCALES_HEADERS:
        worksheet.update(headers_range, [CLIENTES_LOCALES_HEADERS], value_input_option="RAW")
        get_sheet_headers.clear()
    try:
        # Columna G = Tels. Formato texto plano para prevenir evaluaciones como fórmula.
        worksheet.format("G:G", {"numberFormat": {"type": "TEXT"}})
    except Exception:
        # Mejor esfuerzo: si el formato falla, la normalización + RAW sigue protegiendo los teléfonos.
        pass
    return CLIENTES_LOCALES_HEADERS


@st.cache_data(ttl=120)
def load_clientes_locales_dataset(refresh_token: float | None = None) -> pd.DataFrame:
    """Carga Clientes_Locales con metadatos normalizados para coincidencias flexibles."""
    worksheet = get_worksheet_clientes_locales(refresh_token)
    if worksheet is None:
        return pd.DataFrame()

    last_header_cell = rowcol_to_a1(1, len(CLIENTES_LOCALES_HEADERS))
    headers_range = f"A1:{last_header_cell}"
    rows = worksheet.get_all_values()
    if not rows:
        worksheet.update(headers_range, [CLIENTES_LOCALES_HEADERS], value_input_option="RAW")
        return pd.DataFrame(columns=CLIENTES_LOCALES_HEADERS + ["Sheet_Row_Number", "normalized_cliente"])

    headers = [str(value).strip() for value in rows[0]]
    if headers != CLIENTES_LOCALES_HEADERS:
        worksheet.update(headers_range, [CLIENTES_LOCALES_HEADERS], value_input_option="RAW")
        headers = CLIENTES_LOCALES_HEADERS

    records = []
    for row_number, row_values in enumerate(rows[1:], start=2):
        padded = list(row_values[: len(CLIENTES_LOCALES_HEADERS)])
        if len(padded) < len(CLIENTES_LOCALES_HEADERS):
            padded.extend([""] * (len(CLIENTES_LOCALES_HEADERS) - len(padded)))
        if not any(str(cell).strip() for cell in padded):
            continue
        record = dict(zip(CLIENTES_LOCALES_HEADERS, padded))
        record["Sheet_Row_Number"] = row_number
        record["normalized_cliente"] = normalize_client_history_text(record.get("Cliente", ""))
        records.append(record)

    return pd.DataFrame(records)


def _client_name_prefix_tokens_match(query_tokens: list[str], name_tokens: list[str]) -> bool:
    """Valida coincidencias progresivas token por token sin mezclar nombres distintos."""
    if not query_tokens or not name_tokens or len(query_tokens) > len(name_tokens):
        return False
    return all(name_token.startswith(query_token) for query_token, name_token in zip(query_tokens, name_tokens))



def find_clientes_locales_matches(search_text: str, dataset: pd.DataFrame, limit: int = 8) -> list[dict]:
    """Busca coincidencias estrictas para evitar confundir clientes con apellidos similares."""
    normalized_query = normalize_client_history_text(search_text)
    if not normalized_query or dataset.empty:
        return []

    query_tokens = normalized_query.split()
    matches: list[dict] = []
    for _, row in dataset.iterrows():
        normalized_name = str(row.get("normalized_cliente", "") or "")
        if not normalized_name:
            continue

        name_tokens = normalized_name.split()
        ratio = SequenceMatcher(None, normalized_query, normalized_name).ratio()
        is_exact = normalized_query == normalized_name
        is_prefix = normalized_name.startswith(normalized_query)
        is_token_prefix = len(normalized_query) >= 4 and _client_name_prefix_tokens_match(query_tokens, name_tokens)
        is_near_exact = (
            len(query_tokens) == len(name_tokens)
            and query_tokens[:2] == name_tokens[:2]
            and ratio >= 0.96
        )

        if is_exact:
            score = 10.0
        elif is_prefix:
            score = 8.0 + ratio
        elif is_token_prefix:
            score = 6.0 + ratio
        elif is_near_exact:
            score = 4.0 + ratio
        else:
            continue

        row_dict = row.to_dict()
        row_dict["_match_score"] = score
        matches.append(row_dict)

    matches.sort(
        key=lambda item: (
            -float(item.get("_match_score", 0)),
            len(str(item.get("Cliente", "") or "")),
        )
    )
    return matches[:limit]


def apply_cliente_local_record_to_session(
    record: dict,
    *,
    route_prefix: str = "local_route",
    pending_cliente_key: str = "local_route_pending_registro_cliente",
) -> None:
    """Rellena campos de hoja de ruta usando un prefijo de session_state."""
    tipo_inmueble_options = {
        "Consultorio",
        "Clínica",
        "Hospital",
        "Casa",
        "Departamento",
        "Oficina",
        "Local comercial",
        "Otro",
    }
    acceso_privada_options = {
        "No aplica",
        "Aplica",
        "Acceso controlado",
        "Requiere autorización previa",
    }

    st.session_state[pending_cliente_key] = str(record.get("Cliente", "") or "").strip()
    st.session_state[f"{route_prefix}_recibe"] = str(record.get("Recibe", "") or "").strip()
    st.session_state[f"{route_prefix}_calle_no"] = str(record.get("CalleyNumero", "") or "").strip()
    tipo_inmueble = str(record.get("Tipo_Inmueble", "") or "").strip()
    if tipo_inmueble in tipo_inmueble_options:
        st.session_state[f"{route_prefix}_tipo_inmueble"] = tipo_inmueble
    acceso_privada = str(record.get("Acceso_Privada", "") or "").strip()
    if acceso_privada in acceso_privada_options:
        st.session_state[f"{route_prefix}_acceso_privada"] = acceso_privada
    st.session_state[f"{route_prefix}_municipio"] = str(record.get("Municipio", "") or "").strip()
    st.session_state[f"{route_prefix}_telefonos"] = display_phone_from_sheets(record.get("Tels", ""))
    st.session_state[f"{route_prefix}_interior"] = str(record.get("Interior", "") or "").strip()
    st.session_state[f"{route_prefix}_colonia"] = str(record.get("Col", "") or "").strip()
    st.session_state[f"{route_prefix}_cp"] = str(record.get("C_P.", "") or "").strip()
    st.session_state[f"{route_prefix}_referencias"] = str(record.get("Referencias", "") or "").strip()



def apply_cliente_local_to_session(record: dict) -> None:
    """Rellena la hoja de ruta local con la información guardada del cliente."""
    apply_cliente_local_record_to_session(record)


def upsert_cliente_local_if_missing(record: dict[str, str]) -> tuple[bool, str]:
    """Guarda un cliente nuevo solo si no existe ya en Clientes_Locales."""
    client_name = record.get("Cliente", "").strip()
    if not client_name:
        return False, "Nombre de cliente vacío."

    worksheet = get_worksheet_clientes_locales()
    headers = ensure_clientes_locales_headers(worksheet)
    dataset = load_clientes_locales_dataset()
    normalized_name = normalize_client_history_text(client_name)
    if not dataset.empty and dataset["normalized_cliente"].astype(str).eq(normalized_name).any():
        return False, "El cliente ya existe en el historial."

    values = [record.get(header, "") for header in headers]
    worksheet.append_row(values, value_input_option="RAW")
    load_clientes_locales_dataset.clear()
    return True, "Cliente agregado al historial."


def update_existing_cliente_local(row_number: int, record: dict[str, str]) -> None:
    """Actualiza un cliente existente en Clientes_Locales usando el renglón real."""
    worksheet = get_worksheet_clientes_locales()
    headers = ensure_clientes_locales_headers(worksheet)
    start_cell = rowcol_to_a1(row_number, 1)
    end_cell = rowcol_to_a1(row_number, len(headers))
    values = [record.get(header, "") for header in headers]
    worksheet.update(f"{start_cell}:{end_cell}", [values], value_input_option="RAW")
    load_clientes_locales_dataset.clear()

@st.cache_data(ttl=300)
def get_sheet_headers(sheet_name: str):
    """Obtiene y cachea los encabezados de la hoja especificada."""
    if sheet_name == "casos_especiales":
        ws = get_worksheet_casos_especiales()
    elif sheet_name == SHEET_PEDIDOS_HISTORICOS:
        ws = get_worksheet_historico()
    else:
        ws = get_worksheet_operativa()
    return ws.row_values(1) if ws else []


@st.cache_data(ttl=180)
def obtener_devoluciones_autorizadas_sin_folio(id_vendedor_normalizado: str) -> int:
    """Cuenta devoluciones autorizadas sin Folio Nuevo para el vendedor actual."""
    if not id_vendedor_normalizado:
        return 0

    try:
        ws_casos = get_worksheet_casos_especiales()
        df_casos, _ = load_sheet_records_with_row_numbers(ws_casos)
    except Exception:
        return 0

    if df_casos.empty:
        return 0

    if "id_vendedor" not in df_casos.columns:
        df_casos["id_vendedor"] = ""
    if "Seguimiento" not in df_casos.columns:
        df_casos["Seguimiento"] = ""
    if "Folio_Factura" not in df_casos.columns:
        df_casos["Folio_Factura"] = ""

    seguimiento_autorizacion = "Autorización de devolución"
    df_alertas = df_casos[df_casos.apply(is_devolucion_case_row, axis=1)].copy()
    df_alertas = df_alertas[
        df_alertas["id_vendedor"].apply(normalize_vendedor_id) == id_vendedor_normalizado
    ]
    mask = (
        df_alertas["Seguimiento"].astype(str).str.strip().eq(seguimiento_autorizacion)
        & df_alertas["Folio_Factura"].apply(is_empty_folio)
    )
    return int(mask.sum())


# --- AWS S3 CONFIGURATION (NEW) ---
# Load AWS credentials from Streamlit secrets
try:
    AWS_ACCESS_KEY_ID = st.secrets["aws_access_key_id"]
    AWS_SECRET_ACCESS_KEY = st.secrets["aws_secret_access_key"]
    AWS_REGION = st.secrets["aws_region"]
    S3_BUCKET_NAME = st.secrets["s3_bucket_name"]
except KeyError as e:
    st.error(f"❌ Error: AWS S3 credentials not found in Streamlit secrets. Make sure your .streamlit/secrets.toml file is correctly configured. Missing key: {e}")
    st.stop()


  # --- AUTHENTICATION AND CLIENT FUNCTIONS ---

  # Removed the old load_credentials_from_file and get_gspread_client functions
  # as they are replaced by get_google_sheets_client()

  # NEW: Build S3 client
@st.cache_resource
def get_s3_client():
    """Initializes and returns an S3 client."""
    try:
        s3 = boto3.client(
            's3',
            aws_access_key_id=AWS_ACCESS_KEY_ID,
            aws_secret_access_key=AWS_SECRET_ACCESS_KEY,
            region_name=AWS_REGION
        )
        st.session_state.pop("s3_error", None)
        return s3
    except Exception as e:
        st.session_state["s3_error"] = f"❌ Error al inicializar el cliente S3: {e}"
        return None


@st.cache_data(ttl=60)
def check_basic_internet_connectivity(timeout: float = 5.0) -> tuple[bool, str]:
    """Comprueba si hay conexión básica a Internet realizando una solicitud simple."""
    # Usamos el endpoint generate_204, recomendado por Google para comprobar
    # conectividad sin desencadenar respuestas 403 destinadas a los navegadores.
    test_url = "https://clients3.google.com/generate_204"
    try:
        request = Request(
            test_url,
            headers={
                # Algunos endpoints devuelven 403 si falta un User-Agent.
                "User-Agent": "Mozilla/5.0 (compatible; StreamlitApp/1.0)"
            },
        )
        with urlopen(request, timeout=timeout):
            pass
        return True, "Conexión a Internet estable."
    except HTTPError as exc:
        return False, f"Error HTTP al verificar Internet ({exc.code})."
    except (URLError, InvalidURL, TimeoutError, socket.timeout) as exc:
        return False, f"No hay conexión estable a Internet: {exc}"
    except Exception as exc:  # pragma: no cover - captura errores imprevistos
        return False, f"Error inesperado de Internet: {exc}"


def build_connection_statuses(g_client, s3_client) -> list[dict[str, object]]:
    """Genera el estado de conexión para los servicios críticos de la app."""

    statuses: list[dict[str, object]] = []

    internet_ok, internet_message = check_basic_internet_connectivity()
    statuses.append(
        {
            "name": "Internet",
            "ok": internet_ok,
            "message": internet_message,
            "critical": True,
        }
    )

    if g_client is not None:
        statuses.append(
            {
                "name": "Google Sheets",
                "ok": True,
                "message": "Conexión con Google Sheets activa.",
                "critical": True,
            }
        )
    else:
        statuses.append(
            {
                "name": "Google Sheets",
                "ok": False,
                "message": st.session_state.get(
                    "gsheet_error",
                    "❌ Error desconocido al conectar con Google Sheets.",
                ),
                "critical": True,
            }
        )

    if s3_client is not None:
        try:
            s3_client.head_bucket(Bucket=S3_BUCKET_NAME)
            statuses.append(
                {
                    "name": "AWS S3",
                    "ok": True,
                    "message": "Conexión con AWS S3 verificada.",
                    "critical": True,
                }
            )
        except Exception as exc:
            statuses.append(
                {
                    "name": "AWS S3",
                    "ok": False,
                    "message": f"❌ Error al verificar AWS S3: {exc}",
                    "critical": True,
                }
            )
    else:
        statuses.append(
            {
                "name": "AWS S3",
                "ok": False,
                "message": st.session_state.get(
                    "s3_error",
                    "❌ Error desconocido al inicializar AWS S3.",
                ),
                "critical": True,
            }
        )

    return statuses


def display_connection_status_badge(statuses: list[dict[str, object]]) -> None:
    """Muestra un indicador fijo en pantalla con el estado de las conexiones."""

    overall_ok = all(status.get("ok", False) for status in statuses)
    icon = "🟢" if overall_ok else "🔴"
    label = "Conexión segura" if overall_ok else "Problemas de conexión"

    detail_lines = [
        f"{status['name']}: {'✅' if status.get('ok') else '❌'} {status.get('message', '')}"
        for status in statuses
    ]
    escaped_lines = [html.escape(line, quote=True) for line in detail_lines]
    tooltip_text = "&#10;".join(escaped_lines)

    status_class = "ok" if overall_ok else "error"

    badge_html = f"""
    <style>
    .connection-status-container {{
        position: sticky;
        top: 4.5rem;
        z-index: 1000;
        display: flex;
        justify-content: flex-end;
        width: 100%;
        max-width: 1200px;
        margin: 0 auto 0.5rem;
        padding: 0 1.5rem;
    }}
    .connection-status-badge {{
        display: inline-flex;
        align-items: center;
        padding: 0.45rem 0.9rem;
        border-radius: 999px;
        box-shadow: 0 4px 12px rgba(0, 0, 0, 0.12);
        font-weight: 600;
        font-size: 0.95rem;
        transition: transform 0.2s ease, box-shadow 0.2s ease;
        cursor: default;
        gap: 0.5rem;
    }}
    .connection-status-badge--ok {{
        background: linear-gradient(135deg, #1f8f4d, #16693a);
        color: #ffffff;
    }}
    .connection-status-badge--error {{
        background: linear-gradient(135deg, #e74c3c, #c0392b);
        color: #ffffff;
    }}
    .connection-status-badge:hover {{
        transform: translateY(-1px);
        box-shadow: 0 6px 16px rgba(0, 0, 0, 0.18);
    }}
    .connection-status-icon {{
        font-size: 1.1rem;
    }}
    @media (max-width: 768px) {{
        .connection-status-container {{
            top: 3.5rem;
            padding: 0 1rem;
        }}
        .connection-status-badge {{
            width: 100%;
            justify-content: center;
        }}
    }}
    </style>
    <div class="connection-status-container">
        <div class="connection-status-badge connection-status-badge--{status_class}" title="{tooltip_text}">
            <span class="connection-status-icon">{icon}</span>
            <span>{label}</span>
        </div>
    </div>
    """

    st.markdown(badge_html, unsafe_allow_html=True)


# ✅ Clientes listos para usar en cualquier parte
g_spread_client = get_google_sheets_client()
s3_client = get_s3_client()

def upload_file_to_s3(s3_client, bucket_name, file_obj, s3_key):
    """
    Sube un archivo a un bucket de S3.

    Args:
        s3_client: El cliente S3 inicializado.
        bucket_name: El nombre del bucket S3.
        file_obj: El objeto de archivo cargado por st.file_uploader.
        s3_key: La ruta completa y nombre del archivo en S3 (ej. 'pedido_id/filename.pdf').

    Returns:
        tuple: (True, URL del archivo, None) si tiene éxito.
               (False, None, str(error)) cuando ocurre un problema.
    """
    retryable_markers = (
        "timeout",
        "timed out",
        "connection",
        "temporarily unavailable",
        "throttl",
        "slowdown",
        "internalerror",
        "service unavailable",
        "requesttimeout",
        "expiredtoken",
    )

    def _is_retryable_upload_error(error: Exception) -> bool:
        error_text = str(error).lower()
        if any(marker in error_text for marker in retryable_markers):
            return True
        response = getattr(error, "response", None)
        if isinstance(response, dict):
            error_code = str(response.get("Error", {}).get("Code", "")).lower()
            status_code = int(response.get("ResponseMetadata", {}).get("HTTPStatusCode", 0) or 0)
            if error_code in {
                "requesttimeout",
                "throttling",
                "throttlingexception",
                "slowdown",
                "internalerror",
                "serviceunavailable",
                "expiredtoken",
            }:
                return True
            if status_code in {408, 429, 500, 502, 503, 504}:
                return True
        return False

    last_error: Optional[Exception] = None
    for attempt in range(1, S3_UPLOAD_MAX_RETRIES + 1):
        try:
            # Asegúrate de que el puntero del archivo esté al principio
            file_obj.seek(0)
            s3_client.upload_fileobj(file_obj, bucket_name, s3_key)
            file_url = f"https://{bucket_name}.s3.{AWS_REGION}.amazonaws.com/{s3_key}"
            return True, file_url, None
        except Exception as e:
            last_error = e
            if attempt >= S3_UPLOAD_MAX_RETRIES or not _is_retryable_upload_error(e):
                break

            sleep_seconds = S3_UPLOAD_BASE_DELAY_SECONDS * (2 ** (attempt - 1))
            time.sleep(sleep_seconds)

    error_detail = str(last_error) if last_error else "Error desconocido"
    return False, None, (
        f"{error_detail} (reintentos agotados: {S3_UPLOAD_MAX_RETRIES})"
        if last_error
        else error_detail
    )


def upload_files_or_fail(files, s3_client, bucket, prefix):
    uploaded_urls = []
    for file_obj in files or []:
        file_obj.seek(0)
        safe_name = file_obj.name.replace(" ", "_")
        s3_key = f"{prefix}{safe_name}"
        ok, url, error = upload_file_to_s3(s3_client, bucket, file_obj, s3_key)
        if not ok:
            raise Exception(f"Error subiendo {file_obj.name}: {error}")
        uploaded_urls.append(url)
    return uploaded_urls


def append_row_with_confirmation(
    worksheet,
    values,
    pedido_id,
    id_col_index,
    retries=5,
    base_delay=1.0,
):
    def ensure_worksheet_capacity(target_row: int):
        """Expande la hoja si el siguiente renglón excede el límite actual."""

        try:
            current_rows = worksheet.row_count
            if target_row <= current_rows:
                return
            rows_to_add = (target_row - current_rows) + 50  # agrega margen para futuras inserciones
            worksheet.add_rows(rows_to_add)
        except Exception as capacity_error:
            raise Exception(f"No se pudo expandir la hoja: {capacity_error}")

    last_error = None
    for attempt in range(retries):
        try:
            existing_values = worksheet.get_all_values()
            if any(
                len(row) > id_col_index and row[id_col_index] == pedido_id
                for row in existing_values
            ):
                return True
            existing_rows = len(existing_values) + 1
            ensure_worksheet_capacity(existing_rows)

            start_cell = rowcol_to_a1(existing_rows, 1)
            end_cell = rowcol_to_a1(existing_rows, len(values))
            worksheet.update(
                f"{start_cell}:{end_cell}",
                [values],
                value_input_option="USER_ENTERED",
            )
            time.sleep(1 + attempt * 0.5)

            appended_row = worksheet.row_values(existing_rows)
            if len(appended_row) > id_col_index and appended_row[id_col_index] == pedido_id:
                return True
            recent_values = worksheet.get_all_values()[-20:]
            if any(
                len(row) > id_col_index and row[id_col_index] == pedido_id
                for row in recent_values
            ):
                return True
            raise Exception("La escritura no se confirmó")
        except Exception as e:
            last_error = e
            time.sleep(base_delay * (attempt + 1))
    raise Exception(f"No se pudo confirmar la escritura en Google Sheets: {last_error}")
    
# --- Función para actualizar una celda de Google Sheets de forma segura ---
def update_gsheet_cell(worksheet, headers, row_index, col_name, value):
    try:
        if col_name not in headers:
            st.error(f"❌ Error: La columna '{col_name}' no se encontró en Google Sheets para la actualización.")
            return False
        col_index = headers.index(col_name) + 1
        worksheet.update_cell(row_index, col_index, value)
        return True
    except Exception as e:
        st.error(f"❌ Error al actualizar la celda ({row_index}, {col_name}) en Google Sheets: {e}")
        return False


def set_pedido_submission_status(
    status: str,
    message: str,
    detail: str | None = None,
    attachments: list[str] | None = None,
    missing_attachments_warning: bool = False,
    client_name: str = "",
) -> None:
    """Guarda el resultado del registro de un pedido para mostrarlo en la UI."""
    st.session_state["pedido_submission_status"] = {
        "event_id": uuid.uuid4().hex,
        "created_at": time.time(),
        "status": status,
        "message": message,
        "detail": detail or "",
        "attachments": attachments or [],
        "missing_attachments_warning": missing_attachments_warning,
        "client_name": client_name,
    }


def rerun_with_pedido_loading(message: str = "⏳ Actualizando el estado del pedido...") -> None:
    """Muestra aviso de carga para tab1 y relanza la app sin forzar cambio de pestaña."""
    st.session_state["pedido_submission_loading_message"] = message
    st.rerun()


def rerun_with_tab2_loading(message: str = "⏳ Actualizando la modificación del pedido...") -> None:
    """Muestra aviso de carga al modificar antes de relanzar la app."""
    st.session_state[TAB2_LOADING_MESSAGE_KEY] = message
    st.rerun()


def clear_order_related_caches() -> None:
    """Limpia cachés de lectura para reflejar pedidos recién registrados sin recargar la app."""
    for fn_name in (
        "cargar_pedidos",
        "cargar_pedidos_ventas_reportes",
        "cargar_pedidos_combinados",
        "cargar_pedidos_busqueda",
        "obtener_resumen_guias_vendedor",
        "get_tab3_pending_comprobante_dataset",
        "get_tab4_casos_especiales_dataset",
    ):
        clear_fn = getattr(globals().get(fn_name), "clear", None)
        if not callable(clear_fn):
            continue
        try:
            clear_fn()
        except Exception:
            continue


@st.cache_data(ttl=CONNECTION_STATUS_TTL_SECONDS, show_spinner=False)
def get_cached_connection_statuses() -> list[dict[str, object]]:
    """Cachea la verificación de conectividad para reducir recargas visuales frecuentes."""
    return build_connection_statuses(g_spread_client, s3_client)



@st.cache_data(ttl=300)
def cargar_pedidos():
    sheet = g_spread_client.open_by_key("1aWkSelodaz0nWfQx7FZAysGnIYGQFJxAN7RO3YgCiZY").worksheet(SHEET_PEDIDOS_OPERATIVOS)
    data = sheet.get_all_records()
    return pd.DataFrame(data)


@st.cache_data(ttl=300)
def cargar_pedidos_ventas_reportes():
    """Carga pedidos de data_pedidos + datos_pedidos para la vista de reportes."""
    spreadsheet = g_spread_client.open_by_key("1aWkSelodaz0nWfQx7FZAysGnIYGQFJxAN7RO3YgCiZY")
    frames: list[pd.DataFrame] = []
    for nombre_hoja in (SHEET_PEDIDOS_OPERATIVOS, SHEET_PEDIDOS_HISTORICOS):
        try:
            sheet = spreadsheet.worksheet(nombre_hoja)
            frame = pd.DataFrame(sheet.get_all_records())
            if not frame.empty:
                frame["Fuente"] = nombre_hoja
            frames.append(frame)
        except Exception:
            continue

    if not frames:
        return pd.DataFrame()

    return pd.concat(frames, ignore_index=True, sort=False)


usuario_activo = ensure_user_logged_in()

connection_statuses = get_cached_connection_statuses()
display_connection_status_badge(connection_statuses)

status_by_name = {status["name"]: status for status in connection_statuses}

internet_status = status_by_name.get("Internet")
if internet_status and not internet_status.get("ok", False):
    st.warning(internet_status.get("message", "Problema al verificar la conexión a Internet."))

gsheet_status = status_by_name.get("Google Sheets")
if gsheet_status and not gsheet_status.get("ok", False):
    st.error(gsheet_status.get("message", "No se pudo conectar con Google Sheets."))
    if st.button("Reintentar conexión con Google Sheets", key="retry_gsheets_badge"):
        get_google_sheets_client.clear()
        get_cached_connection_statuses.clear()
        st.session_state.pop("gsheet_error", None)
        st.rerun()
    st.stop()

s3_status = status_by_name.get("AWS S3")
if s3_status and not s3_status.get("ok", False):
    st.error(s3_status.get("message", "No se pudo conectar con AWS S3."))
    if st.button("Reintentar conexión con AWS S3", key="retry_s3_badge"):
        get_s3_client.clear()
        get_cached_connection_statuses.clear()
        st.session_state.pop("s3_error", None)
        st.rerun()
    st.stop()

st.markdown(f"### 👋 Bienvenido, {usuario_activo}")

st.markdown(
    """
    <style>
    .remote-zone-status {
        border-radius: 10px;
        padding: 10px 12px;
        font-weight: 800;
        color: var(--text-color);
    }
    .remote-zone-status--remote {
        background: rgba(239, 68, 68, 0.18);
        border: 1px solid rgba(239, 68, 68, 0.48);
    }
    .remote-zone-status--ok {
        background: rgba(34, 197, 94, 0.2);
        border: 1px solid rgba(34, 197, 94, 0.5);
    }
    div[data-testid="stTextInput"] input {
        color: var(--text-color);
        caret-color: var(--text-color);
    }
    div[data-testid="stTextInput"] div[data-baseweb="input"] {
        border-color: rgba(100, 116, 139, 0.45);
    }
    div[data-testid="stTextInput"] div[data-baseweb="input"]:focus-within {
        border-color: #22c55e;
        box-shadow: 0 0 0 1px rgba(34, 197, 94, 0.7);
    }
    </style>
    """,
    unsafe_allow_html=True,
)

remote_postal_codes = get_remote_postal_codes()
_home_col_left, home_col_validator, _home_col_right = st.columns([1, 1.2, 1])
with home_col_validator:
    st.markdown("##### ⚡ Verificador de Zonas Remotas")
    cp_input = st.text_input(
        "Código postal",
        key="remote_zone_cp_input",
        placeholder="Escribe o pega CP",
        max_chars=10,
        label_visibility="collapsed",
        help="Se valida en cuanto escribes o pegas el código postal.",
    )

    cp_digits = re.sub(r"\D", "", str(cp_input or "").strip())
    cp_normalized = cp_digits.zfill(5) if cp_digits and len(cp_digits) <= 5 else cp_digits

    if cp_normalized:
        if cp_normalized in remote_postal_codes:
            st.markdown(
                (
                    "<div class='remote-zone-status remote-zone-status--remote'>"
                    f"🔴 {cp_normalized} · Zona remota"
                    "</div>"
                ),
                unsafe_allow_html=True,
            )
        else:
            st.markdown(
                (
                    "<div class='remote-zone-status remote-zone-status--ok'>"
                    f"🟢 {cp_normalized} · No es zona remota"
                    "</div>"
                ),
                unsafe_allow_html=True,
            )
    else:
        st.caption("Revisa rápido si un CP es zona remota.")

if st.button("🔄 Recargar Página y Conexión", help="Haz clic aquí si algo no carga o da error de Google Sheets."):
    if allow_refresh("main_last_refresh"):
        clear_app_caches()
        get_cached_connection_statuses.clear()
        st.rerun()

st.title("🛒 App de Vendedores TD")
st.write("¡Bienvenido! Aquí puedes registrar y gestionar tus pedidos.")

id_vendedor_sesion_global = normalize_vendedor_id(st.session_state.get("id_vendedor", ""))
if id_vendedor_sesion_global:
    pendientes_devoluciones_home = obtener_devoluciones_autorizadas_sin_folio(id_vendedor_sesion_global)
    if pendientes_devoluciones_home > 0:
        st.warning(
            f"⚠️ Aviso: tienes devoluciones autorizadas sin Folio Nuevo, Pestaña 📁 Casos Especiales. ({pendientes_devoluciones_home})"
        )

    resumen_guias = obtener_resumen_guias_vendedor(
        id_vendedor_sesion_global,
        st.session_state.get("guias_refresh_token"),
    )
    total_guias = int(resumen_guias.get("total", 0) or 0)
    clientes_guias = resumen_guias.get("clientes", []) or []
    current_home_keys = set(resumen_guias.get("keys", []) or [])
    prev_home_keys_raw = st.session_state.get("home_guias_keys", [])
    prev_home_keys = set(prev_home_keys_raw if isinstance(prev_home_keys_raw, list) else [])
    nuevas_home_keys = sorted(current_home_keys - prev_home_keys)

    if prev_home_keys and nuevas_home_keys:
        st.warning(f"🔔 Aviso rápido: tienes {len(nuevas_home_keys)} guía(s) nueva(s).")

    if total_guias > 0:
        clientes_preview = ", ".join(clientes_guias[:2])
        if len(clientes_guias) > 2:
            clientes_preview = f"{clientes_preview} y {len(clientes_guias) - 2} más"
        clientes_msg = f" Clientes: {clientes_preview}." if clientes_preview else ""
        st.info(f"📦 Aviso: tienes {total_guias} pedido(s) con guía cargada.{clientes_msg}")

    st.session_state["home_guias_keys"] = sorted(current_home_keys)

def normalizar(texto):
    return unicodedata.normalize('NFKD', texto).encode('ASCII', 'ignore').decode('utf-8').lower()


def normalizar_busqueda_libre(texto: object) -> str:
    """Normaliza texto para búsquedas flexibles ignorando espacios."""
    return normalizar(str(texto or "")).replace(" ", "")

@st.cache_data(ttl=300)
def obtener_prefijo_s3(pedido_id):
    posibles_prefijos = [
        f"{pedido_id}/", f"adjuntos_pedidos/{pedido_id}/",
        f"adjuntos_pedidos/{pedido_id}", f"{pedido_id}"
    ]
    for prefix in posibles_prefijos:
        try:
            respuesta = s3_client.list_objects_v2(Bucket=S3_BUCKET_NAME, Prefix=prefix, MaxKeys=1)
            if "Contents" in respuesta:
                return prefix if prefix.endswith("/") else prefix + "/"
        except Exception:
            continue
    return None

@st.cache_data(ttl=300)
def obtener_archivos_pdf_validos(prefix):
    try:
        respuesta = s3_client.list_objects_v2(Bucket=S3_BUCKET_NAME, Prefix=prefix)
        archivos = respuesta.get("Contents", [])
        return [f for f in archivos if f["Key"].lower().endswith(".pdf") and any(x in f["Key"].lower() for x in ["guia", "guía", "descarga"])]
    except Exception as e:
        st.error(f"❌ Error al listar archivos en S3 para prefijo {prefix}: {e}")
        return []

@st.cache_data(ttl=300)
def obtener_todos_los_archivos(prefix):
    try:
        respuesta = s3_client.list_objects_v2(Bucket=S3_BUCKET_NAME, Prefix=prefix)
        return respuesta.get("Contents", [])
    except Exception:
        return []

@st.cache_data(ttl=600)
def extraer_texto_pdf(s3_key):
    try:
        response = s3_client.get_object(Bucket=S3_BUCKET_NAME, Key=s3_key)
        with pdfplumber.open(BytesIO(response["Body"].read())) as pdf:
            return "\n".join(page.extract_text() or "" for page in pdf.pages)
    except Exception as e:
        return f"[ERROR AL LEER PDF]: {e}"

def generar_url_s3(s3_key):
    return s3_client.generate_presigned_url(
        'get_object',
        Params={'Bucket': S3_BUCKET_NAME, 'Key': s3_key},
        ExpiresIn=3600
    )

# --- Utilidades y renderizado de casos especiales ---
def partir_urls(value):
    """Normaliza campos de adjuntos que pueden venir como JSON o texto separado."""
    if value is None:
        return []
    s = str(value).strip()
    if not s or s.lower() in ("nan", "none", "n/a"):
        return []
    urls = []
    try:
        obj = json.loads(s)
        if isinstance(obj, list):
            for it in obj:
                if isinstance(it, str) and it.strip():
                    urls.append(it.strip())
                elif isinstance(it, dict):
                    for k in ("url", "URL", "href", "link"):
                        if k in it and str(it[k]).strip():
                            urls.append(str(it[k]).strip())
        elif isinstance(obj, dict):
            for k in ("url", "URL", "href", "link"):
                if k in obj and str(obj[k]).strip():
                    urls.append(str(obj[k]).strip())
    except Exception:
        for p in re.split(r"[,\n;]+", s):
            p = p.strip()
            if p:
                urls.append(p)
    out, seen = [], set()
    for u in urls:
        if u not in seen:
            seen.add(u); out.append(u)
    return out

def __s(v):
    return "" if v is None else str(v).strip()

def __has(v):
    s = __s(v)
    return bool(s) and s.lower() not in ("nan", "none", "n/a")

def __is_url(v):
    s = __s(v).lower()
    return s.startswith("http://") or s.startswith("https://")

def __link(url, label=None):
    u = __s(url)
    if __is_url(u):
        import os
        return f"[{label or (os.path.basename(u) or 'Abrir')}]({u})"
    return u

IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp"}
PDF_EXTENSIONS = {".pdf"}

MAX_INLINE_PDF_BYTES = 10 * 1024 * 1024  # 10 MB límite para incrustar PDFs en base64


def _normalize_url(value: str) -> str:
    """Return a URL with unsafe characters percent-encoded."""
    raw = __s(value)
    if not raw:
        return ""

    try:
        parts = urlsplit(raw)
        if not parts.scheme or not parts.netloc:
            return raw

        safe_path = quote(parts.path, safe="/%:@")
        safe_query = quote(parts.query, safe="=&%:@")
        safe_fragment = quote(parts.fragment, safe="=&%:@")
        return urlunsplit((parts.scheme, parts.netloc, safe_path, safe_query, safe_fragment))
    except Exception:
        return raw


def _render_pdf_iframe_from_base64(data: bytes) -> None:
    """Render a PDF preview from raw bytes encoded in base64."""
    if not data:
        st.info("El archivo no contiene datos para mostrar.")
        return
    b64_pdf = base64.b64encode(data).decode("utf-8")
    iframe = (
        "<iframe src=\"data:application/pdf;base64,{data}\" width=\"100%\" height=\"600\" style=\"border:none;\"></iframe>"
    ).format(data=b64_pdf)
    components.html(iframe, height=620, scrolling=True)


def _render_pdf_iframe_via_google(url: str) -> None:
    """Fallback to Google Docs viewer for remote PDF URLs."""
    viewer_url = f"https://docs.google.com/gview?url={quote(_normalize_url(url), safe='')}\u0026embedded=true"
    iframe = (
        "<iframe src=\"{src}\" width=\"100%\" height=\"600\" style=\"border:none;\" allow=\"fullscreen\"></iframe>"
    ).format(src=html.escape(viewer_url, quote=True))
    components.html(iframe, height=620, scrolling=True)


def _clean_url_path(value: str) -> str:
    """Remove query/hash parameters from a URL or filename."""
    cleaned = __s(value)
    if not cleaned:
        return ""
    return cleaned.split("?")[0].split("#")[0]


def _infer_extension(value: str) -> str:
    """Infer lowercase file extension from a path or URL."""
    cleaned = _clean_url_path(value)
    return os.path.splitext(cleaned)[1].lower()


def _infer_display_name(value: str) -> str:
    """Return a friendly filename to display for a URL or path."""
    cleaned = _clean_url_path(value)
    name = os.path.basename(cleaned)
    return name or cleaned or "Archivo"


def render_remote_file_preview(url: str, display_label: str) -> None:
    """Render an inline preview for a remote file when possible."""
    if not __is_url(url):
        st.info("El archivo no es una URL válida para previsualizar.")
        return

    normalized_url = _normalize_url(url)
    if not normalized_url:
        st.info("El archivo no es una URL válida para previsualizar.")
        return

    ext = _infer_extension(normalized_url)
    if ext in IMAGE_EXTENSIONS:
        st.image(normalized_url, caption=display_label, use_container_width=True)
    elif ext in PDF_EXTENSIONS:
        rendered = False
        try:
            request = Request(normalized_url, headers={"User-Agent": "Mozilla/5.0"})
            with urlopen(request, timeout=10) as response:
                data = response.read(MAX_INLINE_PDF_BYTES + 1)
                if len(data) > MAX_INLINE_PDF_BYTES:
                    raise ValueError("PDF supera el límite para vista previa embebida")
                _render_pdf_iframe_from_base64(data)
                rendered = True
        except (HTTPError, URLError, ValueError, TimeoutError, OSError, InvalidURL):
            st.caption("No se pudo generar la vista previa directa del PDF. Se usa visor alternativo.")

        if not rendered:
            _render_pdf_iframe_via_google(normalized_url)
    else:
        st.info("Vista previa no disponible para este tipo de archivo.")


def add_url_preview_expander(url: str, display_label: str) -> None:
    """Attach an expander with a preview for a given URL."""
    if not __is_url(url):
        return
    with st.expander(f"👁️ Vista previa • {display_label}", expanded=False):
        render_remote_file_preview(url, display_label)


def render_attachment_link(url: str, label: str | None = None, icon: str | None = None, bullet: bool = True) -> None:
    """Render a file link and automatically include an expandable preview."""
    if not __has(url):
        return

    display_label = label or _infer_display_name(url)
    prefix = f"{icon} " if icon else ""

    if bullet:
        if __is_url(url):
            sanitized = _normalize_url(url)
            st.markdown(f"- {prefix}[{display_label}]({sanitized or url})")
        else:
            st.markdown(f"- {prefix}{__s(url)}")
    else:
        if __is_url(url):
            sanitized = _normalize_url(url)
            st.markdown(f"{prefix}[{display_label}]({sanitized or url})")
        else:
            st.markdown(f"{prefix}{__s(url)}")

    if __is_url(url):
        add_url_preview_expander(url, display_label)


def render_uploaded_file_preview(file_obj) -> None:
    """Show a preview expander for an uploaded Streamlit file."""
    file_name = getattr(file_obj, "name", "Archivo")
    display_label = file_name or "Archivo"
    ext = _infer_extension(file_name)

    with st.expander(f"👁️ Vista previa • {display_label}", expanded=False):
        if ext in IMAGE_EXTENSIONS:
            file_obj.seek(0)
            st.image(file_obj.read(), caption=display_label, use_container_width=True)
            file_obj.seek(0)
        elif ext in PDF_EXTENSIONS:
            file_obj.seek(0)
            data = file_obj.read()
            file_obj.seek(0)
            if data:
                b64_pdf = base64.b64encode(data).decode("utf-8")
                iframe = (
                    "<iframe src=\"data:application/pdf;base64,{data}\" width=\"100%\" height=\"600\" style=\"border:none;\"></iframe>"
                ).format(data=b64_pdf)
                components.html(iframe, height=620, scrolling=True)
            else:
                st.info("El archivo no contiene datos para mostrar.")
        else:
            st.info("Vista previa no disponible para este tipo de archivo.")


def render_uploaded_files_preview(title: str, files) -> None:
    """Render previews for uploaded files unless Tab 1 is active."""
    if st.session_state.get("current_tab_index", 0) == 0:
        return

    if not files:
        return

    st.markdown(f"##### 👁️ {title}")
    for file_obj in files:
        render_uploaded_file_preview(file_obj)

def render_caso_especial(row):
    tipo = __s(row.get("Tipo_Envio", ""))
    is_dev = (tipo == "🔁 Devolución")
    title = "🧾 Caso Especial – 🔁 Devolución" if is_dev else "🧾 Caso Especial – 🛠 Garantía"
    st.markdown(f"### {title}")

    vendedor = row.get("Vendedor_Registro", "") or row.get("Vendedor", "")
    id_vendedor_segment = format_id_vendedor_with_mod(row)
    hora = row.get("Hora_Registro", "")

    if is_dev:
        folio_nuevo = clean_folio_for_ui(row.get("Folio_Factura", ""))
        folio_error = row.get("Folio_Factura_Error", "")
        st.markdown(
            f"📄 **Folio Nuevo:** `{folio_nuevo or 'N/A'}`  |  "
            f"📄 **Folio Error:** `{folio_error or 'N/A'}`  |  "
            f"🧑‍💼 **Vendedor:** `{vendedor or 'N/A'}`  |  "
            f"{id_vendedor_segment}  |  "
            f"🕒 **Hora:** `{hora or 'N/A'}`"
        )
    else:
        st.markdown(
            f"📄 **Folio:** `{clean_folio_for_ui(row.get('Folio_Factura','')) or 'N/A'}`  |  "
            f"🧑‍💼 **Vendedor:** `{vendedor or 'N/A'}`  |  "
            f"{id_vendedor_segment}  |  "
            f"🕒 **Hora:** `{hora or 'N/A'}`"
        )

        num_serie = __s(row.get("Numero_Serie", ""))
        fec_compra = __s(row.get("Fecha_Compra", "")) or __s(row.get("FechaCompra", ""))
        if __has(num_serie) or __has(fec_compra):
            st.markdown("**🧷 Datos de compra y serie:**")
            st.markdown(f"- **Número de serie / lote:** `{num_serie or 'N/A'}`")
            st.markdown(f"- **Fecha de compra:** `{fec_compra or 'N/A'}`")

    st.markdown(
        f"**👤 Cliente:** {row.get('Cliente','N/A')}  |  **RFC:** {row.get('Numero_Cliente_RFC','') or 'N/A'}"
    )
    st.markdown(
        f"**Estado:** {row.get('Estado','') or 'N/A'}  |  "
        f"**Estado del Caso:** {row.get('Estado_Caso','') or 'N/A'}  |  "
        f"**Turno:** {row.get('Turno','') or 'N/A'}  |  "
        f"**Tipo Envío Original:** {row.get('Tipo_Envio_Original','') or 'N/A'}"
    )
    st.markdown(f"**📌 Seguimiento:** {row.get('Seguimiento', 'N/A')}")

    rt = __s(row.get("Refacturacion_Tipo",""))
    rs = __s(row.get("Refacturacion_Subtipo",""))
    rf = __s(row.get("Folio_Factura_Refacturada",""))
    if __has(rt) or __has(rs) or __has(rf):
        st.markdown("**♻️ Refacturación:**")
        if __has(rt): st.markdown(f"- **Tipo:** {rt}")
        if __has(rs): st.markdown(f"- **Subtipo:** {rs}")
        if __has(rf): st.markdown(f"- **Folio refacturado:** {rf}")

    if __has(row.get("Resultado_Esperado","")):
        st.markdown(f"**🎯 Resultado Esperado:** {row.get('Resultado_Esperado')}")
    if __has(row.get("Motivo_Detallado","")):
        st.markdown("**📝 Motivo / Descripción:**")
        st.info(__s(row.get("Motivo_Detallado","")))
    if __has(row.get("Material_Devuelto","")):
        st.markdown("**📦 Piezas / Material:**")
        st.info(__s(row.get("Material_Devuelto","")))
    if __has(row.get("Monto_Devuelto","")):
        st.markdown(f"**💵 Monto (dev./estimado):** {row.get('Monto_Devuelto')}")

    if __has(row.get("Area_Responsable","")) or __has(row.get("Nombre_Responsable","")):
        st.markdown(
            f"**🏢 Área Responsable:** {row.get('Area_Responsable','') or 'N/A'}  |  "
            f"**👥 Responsable del Error:** {row.get('Nombre_Responsable','') or 'N/A'}"
        )

    dir_guia = row.get("Direccion_Guia_Retorno", "")
    dir_envio = row.get("Direccion_Envio", "")
    if __has(dir_guia) or __has(dir_envio):
        st.markdown("#### 🏠 Direcciones")
        if __has(dir_guia):
            st.markdown(f"- **Guía de retorno:** {__s(dir_guia)}")
        if __has(dir_envio):
            st.markdown(f"- **Envío destino:** {__s(dir_envio)}")

    if __has(row.get("Fecha_Entrega","")) or __has(row.get("Fecha_Recepcion_Devolucion","")) or __has(row.get("Estado_Recepcion","")):
        st.markdown(
            f"**📅 Fecha Entrega/Cierre:** {row.get('Fecha_Entrega','') or 'N/A'}  |  "
            f"**📅 Recepción:** {row.get('Fecha_Recepcion_Devolucion','') or 'N/A'}  |  "
            f"**📦 Recepción:** {row.get('Estado_Recepcion','') or 'N/A'}"
        )

    nota = __s(row.get("Nota_Credito_URL",""))
    docad = __s(row.get("Documento_Adicional_URL",""))
    if __has(nota):
        if __is_url(nota):
            st.markdown(f"**🧾 Nota de Crédito:** {__link(nota, 'Nota de Crédito')}")
            add_url_preview_expander(nota, "Nota de Crédito")
        else:
            st.markdown(f"**🧾 Nota de Crédito:** {nota}")
    if __has(docad):
        if __is_url(docad):
            st.markdown(f"**📂 Documento Adicional:** {__link(docad, 'Documento Adicional')}")
            add_url_preview_expander(docad, "Documento Adicional")
        else:
            st.markdown(f"**📂 Documento Adicional:** {docad}")
    if __has(row.get("Comentarios_Admin_Devolucion","")):
        st.markdown("**🗒️ Comentario Administrativo:**")
        st.info(__s(row.get("Comentarios_Admin_Devolucion","")))

    mod_txt = __s(row.get("Modificacion_Surtido",""))
    adj_mod_raw = row.get("Adjuntos_Surtido","")
    adj_mod = partir_urls(adj_mod_raw)
    if __has(mod_txt) or adj_mod:
        st.markdown("#### 🛠 Modificación de surtido")
        if __has(mod_txt):
            st.info(mod_txt)
        if adj_mod:
            st.markdown("**Archivos de modificación:**")
            for u in adj_mod:
                render_attachment_link(u)

    with st.expander("📎 Archivos (Adjuntos y Guía)", expanded=False):
        adj_raw = row.get("Adjuntos","")
        adj = partir_urls(adj_raw)
        guia = __s(row.get("Hoja_Ruta_Mensajero","")) or __s(row.get("Adjuntos_Guia",""))
        has_any = False
        if adj:
            has_any = True
            st.markdown("**Adjuntos:**")
            for u in adj:
                render_attachment_link(u)
        if __has(guia) and __is_url(guia):
            has_any = True
            st.markdown("**Guía:**")
            render_attachment_link(guia, "Abrir guía")
        if not has_any:
            st.info("Sin archivos registrados en la hoja.")
    st.markdown("---")

# --- Initialize Gspread Client and S3 Client ---
s3_client = get_s3_client()  # Initialize S3 client

# Removed the old try-except block for client initialization

# --- Tab Definition ---
id_vendedor_tabs = normalize_vendedor_id(st.session_state.get("id_vendedor", ""))
tab1_view_mode_tabs = str(st.session_state.get("tab1_shipping_view_mode", "mty")).strip().lower()
show_tab_ventas_reportes = (
    id_vendedor_tabs in LOCAL_TURNO_CDMX_IDS
    or (id_vendedor_tabs in TAB1_DUAL_VIEW_IDS and tab1_view_mode_tabs == "cdmx")
)
tabs_labels = ["🛒 Registrar Nuevo Pedido"]
if show_tab_ventas_reportes:
    tabs_labels.append("📊 Ventas y Reportes")
tabs_labels.extend([
    "✏️ Modificar Pedido Existente",
    "📦 Guías Cargadas",
    "🧾 Pedidos Pendientes de Comprobante",
    "📁 Casos Especiales",
    "⏳ Pedidos No Entregados",
    "⬇️ Descargar Datos",
    "🔍 Buscar Pedido",
])

# Leer índice de pestaña desde los parámetros de la URL.
# Si falta o viene inválido, usar la pestaña actual de sesión para evitar rebotes en reruns.
raw_tab_param = st.query_params.get("tab")
default_tab: int | None = None

if raw_tab_param is not None:
    try:
        default_tab = int(raw_tab_param[0]) if isinstance(raw_tab_param, list) else int(raw_tab_param)
    except (TypeError, ValueError):
        default_tab = None

if default_tab is None:
    session_tab = st.session_state.get("current_tab_index")
    if isinstance(session_tab, int):
        default_tab = session_tab
    else:
        default_tab = 0

if tabs_labels:
    default_tab = max(0, min(len(tabs_labels) - 1, default_tab))
else:
    default_tab = 0

st.session_state.setdefault("current_tab_index", default_tab)

# Crear pestañas y mantener referencia
tabs = st.tabs(tabs_labels)

components.html(
    f"""
    <script>
    (function() {{
        const desiredIndex = {default_tab};
        const parentWindow = window.parent;
        const parentDocument = parentWindow.document;

        function updateQueryParam(index) {{
            try {{
                const url = new URL(parentWindow.location.href);
                url.searchParams.set('tab', index);
                parentWindow.history.replaceState(null, '', url.toString());
            }} catch (error) {{
                console.error('Error syncing tab param', error);
            }}
        }}

        function attachListeners() {{
            const buttons = parentDocument.querySelectorAll('div[data-baseweb="tab-list"] button');
            if (!buttons || !buttons.length) {{
                setTimeout(attachListeners, 100);
                return;
            }}

            buttons.forEach((button, index) => {{
                if (button.getAttribute('data-tab-listener') === 'true') {{
                    return;
                }}
                button.setAttribute('data-tab-listener', 'true');
                button.addEventListener('click', () => updateQueryParam(index));
            }});

            if (desiredIndex >= 0 && desiredIndex < buttons.length) {{
                const targetButton = buttons[desiredIndex];
                if (targetButton && targetButton.getAttribute('aria-selected') !== 'true') {{
                    targetButton.click();
                }}
            }}
        }}

        if (document.readyState === 'complete') {{
            attachListeners();
        }} else {{
            window.addEventListener('load', attachListeners);
            document.addEventListener('DOMContentLoaded', attachListeners);
            setTimeout(attachListeners, 500);
        }}
    }})();
    </script>
    """,
    height=0,
)
tab1 = tabs[0]
tab_ventas_reportes = tabs[1] if show_tab_ventas_reportes else None
tab_offset = 1 if show_tab_ventas_reportes else 0
tab2 = tabs[1 + tab_offset]
tab5 = tabs[2 + tab_offset]
tab3 = tabs[3 + tab_offset]
tab4 = tabs[4 + tab_offset]
tab6 = tabs[5 + tab_offset]
tab7 = tabs[6 + tab_offset]
tab8 = tabs[7 + tab_offset]
TAB_INDEX_TAB1 = 0
TAB_INDEX_REPORTES = 1 if show_tab_ventas_reportes else None
TAB_INDEX_TAB2 = 1 + tab_offset
TAB_INDEX_TAB5 = 2 + tab_offset
TAB_INDEX_TAB3 = 3 + tab_offset
TAB_INDEX_TAB4 = 4 + tab_offset
TAB_INDEX_TAB6 = 5 + tab_offset
TAB_INDEX_TAB7 = 6 + tab_offset
TAB_INDEX_TAB8 = 7 + tab_offset

# --- List of Vendors (reusable and explicitly alphabetically sorted) ---
VENDEDORES_LIST = sorted([
    "DIANA SOFIA",
    "ALEJANDRO RODRIGUEZ",
    "ANA KAREN ORTEGA MAHUAD",
    "CECILIA SEPULVEDA",
    "CURSOS Y EVENTOS",
    "CASSANDRA MIROSLAVA",
    "DANIELA LOPEZ RAMIREZ",
    "DISTRIBUCION Y UNIVERSIDADES",
    "GLORIA MICHELLE GARCIA TORRES",
    "GRISELDA CAROLINA SANCHEZ GARCIA",
    "JOSE CORTES",
    "JUAN CASTILLEJO",
    "KAREN JAQUELINE",
    "PAULINA TREJO",
    "RUBEN",
    "ROBERTO LEGRA",
    "FRANKO"
])

# Initialize session state for vendor (default linked to logged-in vendor ID)
id_vendedor_logeado = normalize_vendedor_id(st.session_state.get("id_vendedor", ""))
vendedor_predeterminado = VENDEDOR_NOMBRE_POR_ID.get(
    id_vendedor_logeado,
    TAB1_VENDOR_EMPTY_OPTION,
)
if (
    not st.session_state.get("last_selected_vendedor")
    or st.session_state.get("last_selected_vendedor") not in VENDEDORES_LIST
):
    st.session_state.last_selected_vendedor = vendedor_predeterminado

# --- TAB 1: REGISTER NEW ORDER ---
with tab1:
    restore_tab1_form_state_for_retry()
    pending_cache_key = get_pending_submission_key()
    pending_submission_record = load_pending_submission(pending_cache_key)
    has_pending_submission = bool(pending_submission_record and pending_submission_record.get("payload"))
    submission_payload_override = None
    existing_tab1_status = st.session_state.get("pedido_submission_status") or {}
    if (
        existing_tab1_status.get("status") in {"success", "warning", "error"}
        and st.session_state.get("pedido_submit_disabled")
    ):
        # Evita que un estado previo deje bloqueado el botón de registro
        # en un nuevo intento real de captura.
        st.session_state["pedido_submit_disabled"] = False
        st.session_state.pop("pedido_submit_disabled_at", None)

    tab1_is_active = default_tab == TAB_INDEX_TAB1
    if tab1_is_active:
        st.session_state["current_tab_index"] = TAB_INDEX_TAB1
    st.header("📝 Nuevo Pedido")
    id_vendedor_tab1 = normalize_vendedor_id(st.session_state.get("id_vendedor", ""))
    tab1_allow_pedidos_cdmx_option = id_vendedor_tab1 not in LOCAL_TURNO_CDMX_IDS
    tab1_is_dual_view_user = id_vendedor_tab1 in TAB1_DUAL_VIEW_IDS
    tab1_view_mode_key = "tab1_shipping_view_mode"
    if tab1_is_dual_view_user:
        current_view_mode = st.session_state.get(tab1_view_mode_key, "mty")
        if current_view_mode not in {"mty", "cdmx"}:
            current_view_mode = "mty"
            st.session_state[tab1_view_mode_key] = current_view_mode
        st.caption("Vista de captura para vendedores:")
        col_view_mty, col_view_cdmx = st.columns(2)
        with col_view_mty:
            if st.button("Vista vendedores MTY", use_container_width=True):
                st.session_state[tab1_view_mode_key] = "mty"
                current_view_mode = "mty"
        with col_view_cdmx:
            if st.button("Vista vendedores CDMX", use_container_width=True):
                st.session_state[tab1_view_mode_key] = "cdmx"
                current_view_mode = "cdmx"
    else:
        st.session_state.pop(tab1_view_mode_key, None)
        current_view_mode = "mty"

    tab1_special_shipping = current_view_mode == "cdmx"
    if tab1_special_shipping:
        tipo_envio_options = [
            "🚚 Foráneo CDMX",
            "📍 Local CDMX",
            "🔁 Devolución",
            "🛠 Garantía",
            "📋 Solicitudes de Guía",
            "🎓 Cursos y Eventos",
        ]
    else:
        tipo_envio_options = [
            "🚚 Foráneo",
            "📍 Local",
            "🔁 Devolución",
            "🛠 Garantía",
            "📋 Solicitudes de Guía",
            "🎓 Cursos y Eventos",
        ]
    if tab1_allow_pedidos_cdmx_option:
        tipo_envio_options.insert(1, "🏙️ Pedidos CDMX")

    current_tipo_envio = st.session_state.get("tipo_envio_selector_global", tipo_envio_options[0])
    if tab1_special_shipping:
        if current_tipo_envio in {"🚚 Pedido Foráneo", "🚚 Foráneo"}:
            current_tipo_envio = "🚚 Foráneo CDMX"
        elif current_tipo_envio in {"📍 Pedido Local", "📍 Local"}:
            current_tipo_envio = "📍 Local CDMX"
        elif current_tipo_envio in {"🏙️ Pedido CDMX", "🏙️ Pedidos CDMX"} and tab1_allow_pedidos_cdmx_option:
            current_tipo_envio = "🏙️ Pedidos CDMX"
    else:
        if current_tipo_envio in {"🚚 Pedido Foráneo", "🚚 Foráneo CDMX"}:
            current_tipo_envio = "🚚 Foráneo"
        elif current_tipo_envio in {"📍 Pedido Local", "📍 Local CDMX"}:
            current_tipo_envio = "📍 Local"
        elif current_tipo_envio in {"🏙️ Pedido CDMX", "🏙️ Pedidos CDMX"} and tab1_allow_pedidos_cdmx_option:
            current_tipo_envio = "🏙️ Pedidos CDMX"
    if current_tipo_envio not in tipo_envio_options:
        current_tipo_envio = tipo_envio_options[0]
        st.session_state["tipo_envio_selector_global"] = current_tipo_envio

    tipo_envio_ui = st.selectbox(
        "📦 Tipo de Envío",
        tipo_envio_options,
        index=tipo_envio_options.index(current_tipo_envio),
        key="tipo_envio_selector_global",
    )
    tipo_envio = tipo_envio_ui
    tipo_envio_excel = tipo_envio_ui
    if tipo_envio_ui in {"🚚 Foráneo CDMX", "🚚 Foráneo", "🏙️ Pedidos CDMX"}:
        tipo_envio = "🚚 Pedido Foráneo"
        tipo_envio_excel = "🏙️ Pedidos CDMX" if tipo_envio_ui == "🏙️ Pedidos CDMX" else "🚚 Pedido Foráneo"
    elif tipo_envio_ui in {"📍 Local CDMX", "📍 Local"}:
        tipo_envio = "📍 Pedido Local"
        tipo_envio_excel = "📍 Pedido Local"

    tipo_envio_original = ""
    if tipo_envio == "🔁 Devolución":
        tipo_envio_original = st.selectbox(
            "📦 Tipo de Envío Original",
            ["🚚 Foráneo","📍 Local"],
            index=0,
            key="tipo_envio_original",
            help="Selecciona el tipo de envío del pedido que se va a devolver.",
        )
    else:
        st.session_state.pop("tipo_envio_original", None)

    subtipo_local = ""
    is_local_pasa_bodega = False
    is_local_recoge_aula = False
    is_devolucion_local = tipo_envio == "🔁 Devolución" and tipo_envio_original == "📍 Local"
    usa_logica_local = tipo_envio == "📍 Pedido Local" or is_devolucion_local
    expand_payment_details_default = (
        id_vendedor_tab1 in TAB1_LOCAL_CDMX_DISABLE_ROUTE_IDS
        or (tab1_is_dual_view_user and tab1_special_shipping)
    )
    usa_hoja_ruta_local = usa_logica_local
    if usa_logica_local:
        if tipo_envio == "📍 Pedido Local":
            st.markdown("---")
            st.subheader("⏰ Detalle de Pedido Local")
            local_shift_options = get_local_shift_options(
                st.session_state.get("id_vendedor", "") if tab1_special_shipping else None,
                force_cdmx_view=tab1_special_shipping,
            )
            current_subtipo_local = st.session_state.get("subtipo_local_selector", local_shift_options[0])
            if current_subtipo_local not in local_shift_options:
                current_subtipo_local = local_shift_options[0]
                st.session_state["subtipo_local_selector"] = current_subtipo_local
            subtipo_local = st.selectbox(
                "Turno/Locales",
                local_shift_options,
                index=local_shift_options.index(current_subtipo_local),
                key="subtipo_local_selector",
                help="Selecciona el turno o tipo de entrega para pedidos locales."
            )
            is_local_pasa_bodega = subtipo_local == "📦 Pasa a Bodega"
            is_local_recoge_aula = subtipo_local == "🎓 Recoge en Aula"
        else:
            # Para devolución local no se muestra selector de turno/locales.
            subtipo_local = "☀️ Local Mañana"
            st.session_state["subtipo_local_selector"] = subtipo_local

        if is_local_recoge_aula:
            usa_hoja_ruta_local = False

        if not usa_hoja_ruta_local:
            st.session_state["local_route_selected_history_label"] = None
            st.session_state["local_route_selected_history_row"] = None
            if is_local_recoge_aula:
                st.session_state["local_route_hora_entrega_manual"] = ""
                st.session_state.pop("local_route_hora_entrega_selector", None)
                st.session_state.pop("local_route_hora_entrega_custom", None)
                st.caption(
                    "ℹ️ Para **🎓 Recoge en Aula** no se usa hoja de ruta ni actualización de `Clientes_Locales`."
                )
        elif is_local_pasa_bodega:
            st.session_state["local_route_selected_history_label"] = None
            st.session_state["local_route_selected_history_row"] = None
            st.caption(
                "ℹ️ Para **📦 Pasa a Bodega** no se usa búsqueda automática de cliente ni actualización de `Clientes_Locales`."
            )
        else:
            pending_registro_cliente = st.session_state.pop("local_route_pending_registro_cliente", None)
            if pending_registro_cliente is not None:
                st.session_state["registro_cliente"] = pending_registro_cliente

            st.markdown("---")
            st.subheader("🤝 Cliente con búsqueda automática")
            st.caption(
                "Escribe el nombre del cliente y dale ENTER. La app buscará coincidencias en el historial local."
            )
            registro_cliente = st.text_input(
                "🤝 Cliente",
                key="registro_cliente",
                placeholder="Escribe o pega el nombre del cliente",
                help="Busca coincidencias más estrictas en Clientes_Locales, priorizando nombres exactos o capturas progresivas del mismo nombre.",
            )

            clientes_locales_df = load_clientes_locales_dataset()
            client_history_matches = find_clientes_locales_matches(registro_cliente, clientes_locales_df)
            client_history_options: dict[str, dict] = {}
            normalized_registro_cliente = normalize_client_history_text(registro_cliente)
            exact_match_label = None
            for match in client_history_matches:
                display_label = f"{str(match.get('Cliente', '')).strip()} | C.P. {str(match.get('C_P.', '')).strip() or 'N/A'}"
                suffix = 2
                base_label = display_label
                while display_label in client_history_options:
                    display_label = f"{base_label} ({suffix})"
                    suffix += 1
                client_history_options[display_label] = match
                if normalized_registro_cliente and normalize_client_history_text(match.get("Cliente", "")) == normalized_registro_cliente:
                    exact_match_label = display_label

            selected_history_row = st.session_state.get("local_route_selected_history_row")
            previous_history_label = st.session_state.get("local_route_selected_history_label")

            if exact_match_label and exact_match_label in client_history_options:
                selected_history_label = exact_match_label
                selected_history_record = client_history_options[exact_match_label]
                st.caption(f"✅ Cliente encontrado en historial: {selected_history_label}")
                selected_row_number = parse_sheet_row_number(selected_history_record.get("Sheet_Row_Number"))
                if (
                    selected_history_row != selected_row_number
                    or str(st.session_state.get("registro_cliente", "") or "").strip() != str(selected_history_record.get("Cliente", "") or "").strip()
                ):
                    st.session_state["local_route_selected_history_label"] = selected_history_label
                    st.session_state["local_route_selected_history_row"] = selected_row_number
                    apply_cliente_local_to_session(selected_history_record)
                    st.rerun()
            elif len(client_history_options) == 1:
                selected_history_label, selected_history_record = next(iter(client_history_options.items()))
                st.caption(f"✅ Coincidencia encontrada: {selected_history_label}")
                selected_row_number = parse_sheet_row_number(selected_history_record.get("Sheet_Row_Number"))
                if (
                    selected_history_row != selected_row_number
                    or str(st.session_state.get("registro_cliente", "") or "").strip() != str(selected_history_record.get("Cliente", "") or "").strip()
                ):
                    st.session_state["local_route_selected_history_label"] = selected_history_label
                    st.session_state["local_route_selected_history_row"] = selected_row_number
                    apply_cliente_local_to_session(selected_history_record)
                    st.rerun()
            elif client_history_options:
                option_labels = list(client_history_options.keys())
                selected_history_index = None
                if previous_history_label in client_history_options:
                    selected_history_index = option_labels.index(previous_history_label)

                selected_history_label = st.radio(
                    "Coincidencias encontradas",
                    options=option_labels,
                    index=selected_history_index,
                    key="local_route_selected_history_label",
                    help="Selecciona una coincidencia para cargar la información guardada del cliente.",
                )
                selected_history_record = client_history_options.get(selected_history_label)
                if selected_history_record:
                    selected_row_number = parse_sheet_row_number(selected_history_record.get("Sheet_Row_Number"))
                    selected_history_name = str(selected_history_record.get("Cliente", "") or "").strip()
                    if (
                        selected_history_row != selected_row_number
                        or str(st.session_state.get("registro_cliente", "") or "").strip() != selected_history_name
                    ):
                        st.session_state["local_route_selected_history_row"] = selected_row_number
                        apply_cliente_local_to_session(selected_history_record)
                        st.rerun()
            elif registro_cliente.strip():
                st.caption("🆕 Cliente nuevo sin historial. Puedes continuar y al registrar el pedido se agregará al historial local.")
                st.session_state["local_route_selected_history_label"] = None
                st.session_state["local_route_selected_history_row"] = None

    registrar_nota_venta = st.checkbox(
        "🧾 Registrar nota de venta",
        key="registrar_nota_venta_checkbox",
        help="Activa para capturar los datos de una nota de venta.",
    )

    # -------------------------------
    # Inicialización de variables
    # -------------------------------
    vendedor = ""
    registro_cliente = ""
    numero_cliente_rfc = ""
    nota_venta = ""
    motivo_nota_venta = ""
    folio_factura_input_value = ""
    folio_factura = ""
    folio_factura_error = ""  # 🆕 NUEVO para devoluciones
    fecha_entrega = datetime.now().date()
    comentario = ""
    uploaded_files = []

    # Variables Devolución
    tipo_envio_original = ""
    estatus_origen_factura = ""
    resultado_esperado = ""
    material_devuelto = ""
    motivo_detallado = ""
    area_responsable = ""
    nombre_responsable = ""
    monto_devuelto = 0.0
    comprobante_cliente = None
    aplica_pago = "No"

    # Variables Garantía
    g_resultado_esperado = ""
    g_descripcion_falla = ""
    g_piezas_afectadas = ""
    g_monto_estimado = 0.0
    g_area_responsable = ""
    g_nombre_responsable = ""
    g_numero_serie = ""
    g_fecha_compra = None
    direccion_guia_retorno = ""
    direccion_envio_destino = ""

    # Variables Estado de Pago
    comprobante_pago_files = []
    fecha_pago = None
    forma_pago = ""
    terminal = ""
    banco_destino = ""
    monto_pago = 0.0
    referencia_pago = ""
    estado_pago = "🔴 No Pagado"

    # Variables Hoja de Ruta Local
    local_route_recibe = ""
    local_route_calle_no = ""
    local_route_tipo_inmueble = ""
    local_route_acceso_privada = ""
    local_route_municipio = "MONTERREY"
    local_route_telefonos = ""
    local_route_interior = ""
    local_route_colonia = ""
    local_route_cp = ""
    local_route_forma_pago = "TRANSFERENCIA"
    local_route_total_factura = 0.0
    local_route_adeudo_anterior = 0.0
    local_route_referencias = ""
    local_route_hora_entrega = str(st.session_state.get("local_route_hora_entrega_manual", "") or "").strip()

    # -------------------------------
    # --- FORMULARIO PRINCIPAL ---
    # -------------------------------
    form_nonce = int(st.session_state.get(TAB1_FORM_NONCE_KEY, 0) or 0)
    route_post_confirm_notice = (
        st.session_state.get(LOCAL_ROUTE_POST_CONFIRM_NOTICE_KEY)
        if usa_hoja_ruta_local
        else None
    )
    with st.form(key=f"new_pedido_form_{form_nonce}", clear_on_submit=False):
        st.markdown("---")
        st.subheader("Información Básica del Cliente y Pedido")

        id_vendedor_form = normalize_vendedor_id(st.session_state.get("id_vendedor", ""))
        default_vendedor = VENDEDOR_NOMBRE_POR_ID.get(id_vendedor_form, TAB1_VENDOR_EMPTY_OPTION)
        vendor_options_tab1 = [TAB1_VENDOR_EMPTY_OPTION] + VENDEDORES_LIST

        # Si el usuario logeado tiene mapeo, siempre forzamos ese vendedor como base del formulario.
        if default_vendedor in VENDEDORES_LIST:
            st.session_state.last_selected_vendedor = default_vendedor

        selected_vendedor_state = st.session_state.get("last_selected_vendedor", TAB1_VENDOR_EMPTY_OPTION)
        if selected_vendedor_state not in vendor_options_tab1:
            selected_vendedor_state = default_vendedor if default_vendedor in vendor_options_tab1 else TAB1_VENDOR_EMPTY_OPTION
            st.session_state.last_selected_vendedor = selected_vendedor_state

        try:
            initial_vendedor_index = vendor_options_tab1.index(selected_vendedor_state)
        except ValueError:
            initial_vendedor_index = 0

        vendedor = st.selectbox("👤 Vendedor", vendor_options_tab1, index=initial_vendedor_index)
        if vendedor != st.session_state.get("last_selected_vendedor", None):
            st.session_state.last_selected_vendedor = vendedor

        if not usa_logica_local or is_local_pasa_bodega or not usa_hoja_ruta_local:
            registro_cliente = st.text_input("🤝 Cliente", key="registro_cliente")
        else:
            registro_cliente = str(st.session_state.get("registro_cliente", "") or "").strip()

        # Número de cliente / RFC para Casos Especiales (Devolución y Garantía)
        if tipo_envio in ["🔁 Devolución", "🛠 Garantía"]:
            numero_cliente_rfc = st.text_input("🆔 Número de Cliente o RFC (opcional)", key="numero_cliente_rfc")

        # Datos adicionales de Devolución (tipo_envio_original se captura fuera del formulario)
        if tipo_envio == "🔁 Devolución":
            estatus_origen_factura = st.selectbox(
                "📊 Estatus de Factura Origen",
                ["Pagado", "Crédito", "Otro"],
                index=0,
                key="estatus_factura_origen",
                help="Indica el estatus de la factura original asociada al pedido devuelto."
            )

            aplica_pago = st.radio(
                "✅ Aplica pago",
                options=["Sí", "No"],
                index=1,
                horizontal=True,
                key="aplica_pago_selector",
                help="Se llena este campo cuando el cliente va pagar alguna diferencia.",
            )

            # 🆕 NUEVO: Folio Error arriba del folio normal
            folio_factura_error = st.text_input(
                "📄 Folio Error (factura equivocada, si aplica)",
                key="folio_factura_error_input"
            )

        if registrar_nota_venta:
            nota_venta = st.text_input(
                "🧾 Nota de Venta",
                key="nota_venta_input",
                help="Ingresa el número de nota de venta si aplica. Se guardará en la misma columna que el folio.",
            )
            motivo_nota_venta = st.text_area(
                "✏️ Motivo de nota de venta",
                key="motivo_nota_venta_input",
                help="Describe el motivo de la nota de venta, si se registró una.",
            )
            st.session_state.pop("folio_factura_input", None)
        else:
            # Folio normal (renombrado a 'Folio Nuevo' en devoluciones)
            folio_label = "📄 Folio Nuevo" if tipo_envio == "🔁 Devolución" else "📄 Folio de Factura"
            folio_factura_input_value = st.text_input(folio_label, key="folio_factura_input")

        # Campos de pedido normal (no Casos Especiales)
        if tipo_envio not in ["🔁 Devolución", "🛠 Garantía"]:
            fecha_entrega = st.date_input(
                "🗓 Fecha de Entrega Requerida",
                value=datetime.now().date(),
                key="fecha_entrega_input",
            )
            if usa_logica_local and not is_local_pasa_bodega and not is_local_recoge_aula:
                local_route_hour_options = [
                    "9:00 AM a 2:00 PM",
                    "3:00 PM a 7:00 PM",
                    "10:00 AM a 7:00 PM",
                ]
                if not tab1_special_shipping:
                    local_route_hour_options = [LOCAL_ROUTE_HOUR_AUTOMATIC_OPTION, *local_route_hour_options]

                hora_entrega_actual = str(st.session_state.get("local_route_hora_entrega_manual", "") or "").strip()
                if hora_entrega_actual in local_route_hour_options:
                    default_hora_selector = hora_entrega_actual
                elif tab1_special_shipping:
                    default_hora_selector = local_route_hour_options[0]
                else:
                    default_hora_selector = LOCAL_ROUTE_HOUR_AUTOMATIC_OPTION

                if st.session_state.get("local_route_hora_entrega_selector") != default_hora_selector:
                    st.session_state["local_route_hora_entrega_selector"] = default_hora_selector

                hora_entrega_selector = st.selectbox(
                    "🕒 HORA DE ENTREGA",
                    local_route_hour_options,
                    key="local_route_hora_entrega_selector",
                    help="Selecciona el horario de entrega para el pedido local.",
                )

                if hora_entrega_selector == LOCAL_ROUTE_HOUR_AUTOMATIC_OPTION:
                    st.session_state["local_route_hora_entrega_manual"] = ""
                    local_route_hora_entrega = ""
                else:
                    local_route_hora_entrega = hora_entrega_selector
                    st.session_state["local_route_hora_entrega_manual"] = local_route_hora_entrega
                st.session_state.pop("local_route_hora_entrega_custom", None)

        comentario = st.text_area(
            "💬 Comentario / Descripción Detallada",
            key="comentario_detallado",
        )

        if usa_logica_local:
            if usa_hoja_ruta_local and not is_local_pasa_bodega:
                st.markdown("### 🗺️ Hoja de Ruta Local")
                col_local_1, col_local_2 = st.columns(2)
                with col_local_1:
                    local_route_recibe = st.text_input("🙋 Recibe", key="local_route_recibe")
                    local_route_calle_no = st.text_input("📍 CALLE Y NO.", key="local_route_calle_no")
                    local_route_tipo_inmueble = st.selectbox(
                        "🏢 TIPO INMUEBLE",
                        [
                            "Consultorio",
                            "Clínica",
                            "Hospital",
                            "Casa",
                            "Departamento",
                            "Oficina",
                            "Local comercial",
                            "Otro",
                        ],
                        key="local_route_tipo_inmueble",
                    )
                    local_route_acceso_privada = st.selectbox(
                        "🚧 ACCESO PRIVADA",
                        [
                            "No aplica",
                            "Aplica",
                            "Acceso controlado",
                            "Requiere autorización previa",
                        ],
                        key="local_route_acceso_privada",
                    )
                    local_route_municipio = st.text_input("🗺️ MUNICIPIO", key="local_route_municipio")
                    local_route_telefonos = st.text_input("☎️ TELS", key="local_route_telefonos")
                with col_local_2:
                    local_route_interior = st.text_input("🚪 INTERIOR", key="local_route_interior")
                    local_route_colonia = st.text_input("🏘️ COL.", key="local_route_colonia")
                    local_route_cp = st.text_input("📮 C.P.", key="local_route_cp")
                    local_route_forma_pago = st.selectbox(
                        "💳 FORMA DE PAGO",
                        [
                            "Transferencia",
                            "Depósito en Efectivo",
                            "Tarjeta de Débito",
                            "Tarjeta de Crédito",
                            "Credito TD",
                            "Cheque",
                        ],
                        key="local_route_forma_pago",
                    )
                    local_route_total_factura = st.number_input(
                        "💵 TOTAL FACTURA",
                        min_value=0.0,
                        format="%.2f",
                        key="local_route_total_factura",
                    )
                    local_route_adeudo_anterior = st.number_input(
                        "💸 ADEUDO ANT.",
                        min_value=0.0,
                        format="%.2f",
                        key="local_route_adeudo_anterior",
                    )
                local_route_referencias = st.text_area(
                    "📝 REFERENCIAS Y/O COMENTARIOS (solo hoja de ruta)",
                    key="local_route_referencias",
                    help="Este campo se usa únicamente en la hoja de ruta y en el historial de Clientes_Locales.",
                )

            st.markdown("---")
            st.subheader("💰 Estado de Pago")
            opciones_estado_pago = (
                ["🎟️ No Aplica", "🔴 No Pagado", "✅ Pagado"]
                if registrar_nota_venta
                else ["🔴 No Pagado", "✅ Pagado", "💳 CREDITO"]
            )
            if st.session_state.get("estado_pago") not in opciones_estado_pago:
                st.session_state["estado_pago"] = opciones_estado_pago[0]

            estado_pago = st.selectbox(
                "Estado de Pago",
                opciones_estado_pago,
                index=0,
                key="estado_pago",
            )

            selected_history_row = parse_sheet_row_number(st.session_state.get("local_route_selected_history_row"))
            show_update_client_button = (
                usa_hoja_ruta_local
                and (not is_local_pasa_bodega)
                and (selected_history_row is not None)
            )
            if show_update_client_button:
                update_client_history_button = st.form_submit_button(
                    "📝 Actualizar info del cliente",
                    help="Actualiza el registro histórico del cliente seleccionado con los datos actuales del formulario.",
                )
                st.caption(
                    "Este botón actualiza la base histórica del cliente. "
                    "Usar solo si se requiere almacenar nueva info del cliente del formulario actual"
                )
            else:
                update_client_history_button = False

            requiere_captura_pago = estado_pago == "✅ Pagado"

            comprobante_pago_files = st.file_uploader(
                "💲 Comprobante(s) de Pago",
                type=["pdf", "jpg", "jpeg", "png"],
                accept_multiple_files=True,
                key="comprobante_uploader_final"
            )
            render_uploaded_files_preview("Comprobantes de pago seleccionados", comprobante_pago_files)

            if requiere_captura_pago:
                st.warning("⚠️ Estado en PAGADO: debes adjuntar al menos un comprobante antes de registrar el pedido.")
            else:
                st.caption("ℹ️ Los Comprobantes son obligatorios cuando el estado sea '✅ Pagado'.")

            with st.expander(
                "🧾 Detalles del Pago (opcional)",
                expanded=expand_payment_details_default,
            ):
                col1, col2, col3 = st.columns(3)
                with col1:
                    fecha_pago = st.date_input("📅 Fecha del Pago", value=datetime.today().date(), key="fecha_pago_input")
                with col2:
                    forma_pago = st.selectbox("💳 Forma de Pago", [
                        "Transferencia", "Depósito en Efectivo", "Tarjeta de Débito", "Tarjeta de Crédito", "Cheque"
                    ], key="forma_pago_input")
                with col3:
                    monto_pago = st.number_input("💲 Monto del Pago", min_value=0.0, format="%.2f", key="monto_pago_input")

                col4, col5 = st.columns(2)
                with col4:
                    if forma_pago in ["Tarjeta de Débito", "Tarjeta de Crédito"]:
                        terminal = st.selectbox(
                            "🏧 Terminal",
                            [
                                "BANORTE",
                                "AFIRME",
                                "VELPAY",
                                "CLIP",
                                "PAYPAL",
                                "BBVA",
                                "CONEKTA",
                                "MERCADO PAGO",
                            ],
                            key="terminal_input",
                        )
                        banco_destino = ""
                    else:
                        banco_destino = st.selectbox("🏦 Banco Destino", ["BANORTE", "BANAMEX", "AFIRME", "BANCOMER OP", "BANCOMER CURSOS"], key="banco_destino_input")
                        terminal = ""
                with col5:
                    referencia_pago = st.text_input("🔢 Referencia (opcional)", key="referencia_pago_input")

        else:
            confirm_route_button = False
            update_client_history_button = False

        if tipo_envio == "🚚 Pedido Foráneo":
            direccion_guia_retorno = st.text_area(
                "📬 Dirección para Envió (Obligatorio al Solicitar Guia)",
                key="direccion_guia_retorno_foraneo",
            )

        # --- Campos adicionales para Devolución ---
        if tipo_envio == "🔁 Devolución":
            st.markdown("### 🔁 Información de Devolución")

            resultado_esperado = st.selectbox(
                "🎯 Resultado Esperado",
                ["Cambio de Producto", "Devolución de Dinero", "Saldo a Favor", "Material Faltante"],
                key="resultado_esperado"
            )

            material_devuelto = st.text_area(
                "📦 Material a Devolver (códigos, descripciones, cantidades y monto individual con IVA)",
                key="material_devuelto"
            )

            monto_devuelto = st.number_input(
                "💲 Total de Materiales a Devolver (con IVA)",
                min_value=0.0,
                format="%.2f",
                key="monto_devuelto"
            )

            area_responsable = st.selectbox(
                "🏷 Área Responsable del Error",
                ["Vendedor", "Almacén", "Cliente", "Proveedor", "Otro"],
                key="area_responsable"
            )

            if area_responsable in ["Vendedor", "Almacén"]:
                nombre_responsable = st.text_input("👤 Nombre del Empleado Responsable", key="nombre_responsable")
            else:
                nombre_responsable = "No aplica"

            motivo_detallado = st.text_area("📝 Explicación Detallada del Caso", key="motivo_detallado")

        # --- Campos adicionales para Garantía ---
        if tipo_envio == "🛠 Garantía":
            st.markdown("### 🛠 Información de Garantía")
            aplica_pago = "No"

            g_resultado_esperado = st.selectbox(
                "🎯 Resultado Esperado",
                ["Reparación", "Cambio por Garantía", "Nota de Crédito"],
                key="g_resultado_esperado"
            )

            g_descripcion_falla = st.text_area(
                "🧩 Descripción de la Falla (detallada)",
                key="g_descripcion_falla"
            )

            g_piezas_afectadas = st.text_area(
                "🧰 Piezas/Material afectado (códigos, descripciones, cantidades y monto individual con IVA si aplica)",
                key="g_piezas_afectadas"
            )

            g_monto_estimado = st.number_input(
                "💲 Monto estimado de atención (con IVA, opcional)",
                min_value=0.0,
                format="%.2f",
                key="g_monto_estimado"
            )

            g_area_responsable = st.selectbox(
                "🏷 Área posiblemente responsable",
                ["Vendedor", "Almacén", "Cliente", "Proveedor", "Otro"],
                key="g_area_responsable"
            )

            if g_area_responsable in ["Vendedor", "Almacén"]:
                g_nombre_responsable = st.text_input("👤 Nombre del Empleado Responsable", key="g_nombre_responsable")
            else:
                g_nombre_responsable = "No aplica"

            col_g1, col_g2 = st.columns(2)
            with col_g1:
                g_numero_serie = st.text_input("🔢 Número de serie / lote (opcional)", key="g_numero_serie")
            with col_g2:
                g_fecha_compra = st.date_input("🗓 Fecha de compra (opcional)", value=None, key="g_fecha_compra")

        if tipo_envio in ["🔁 Devolución", "🛠 Garantía"]:
            st.markdown("### 🏠 Direcciones")
            direccion_guia_retorno = st.text_area(
                "📬 Dirección para guía de retorno",
                key="direccion_guia_retorno",
            )
            direccion_envio_destino = st.text_area(
                "📦 Dirección de envío destino",
                key="direccion_envio_destino",
            )
        else:
            aplica_pago = "No"

        st.markdown("---")
        st.subheader("📎 Adjuntos del Pedido")
        uploaded_files = st.file_uploader(
            "Sube archivos del pedido",
            type=["pdf", "jpg", "jpeg", "png", "xlsx", "docx"],
            accept_multiple_files=True,
            key="pedido_adjuntos",
        )
        render_uploaded_files_preview("Archivos del pedido seleccionados", uploaded_files)

        auto_route_filename = ""
        if usa_hoja_ruta_local and not is_local_pasa_bodega:
            auto_route_filename = st.session_state.get(LOCAL_ROUTE_GENERATED_FILENAME_KEY, "")
            if route_post_confirm_notice and route_post_confirm_notice.get("filename"):
                auto_route_filename = route_post_confirm_notice.get("filename", "")

        if auto_route_filename:
            st.info(f"📎 Hoja de ruta adjuntada automáticamente: {auto_route_filename}")

        if usa_hoja_ruta_local and not is_local_pasa_bodega:
            st.markdown("---")
            st.subheader("👀 Vista previa opcional del Excel local")
            st.caption(
                "Este botón es solo para revisar cómo quedará el Excel con datos automáticos como día, "
                "horario y gran total. Aunque no lo uses, el archivo se genera y se adjunta al registrar el pedido."
            )
            route_notice_placeholder = st.empty()
            confirm_route_button = st.form_submit_button(
                "👀 Ver / actualizar vista previa del Excel (opcional)",
                help="Genera una vista previa del Excel local con la información capturada hasta este momento.",
            )
            if route_post_confirm_notice:
                route_notice_filename = route_post_confirm_notice.get("filename", "")
                with route_notice_placeholder.container():
                    st.success("✅ Vista previa del Excel actualizada correctamente.")
                    if route_notice_filename:
                        st.caption(f"📎 Archivo generado para revisión: `{route_notice_filename}`")

        # --- Evidencias/Comprobantes PARA DEVOLUCIONES y GARANTÍAS ---
        if tipo_envio in ["🔁 Devolución", "🛠 Garantía"]:
            st.markdown("---")
            st.subheader("📎 Evidencias / Comprobantes del Caso")
            comprobante_cliente = st.file_uploader(
                "Sube evidencia(s) del caso (comprobantes, fotos, PDF, etc.)",
                type=["pdf", "jpg", "jpeg", "png"],
                accept_multiple_files=True,
                key="comprobante_cliente",
                help="Sube archivos relacionados con esta devolución o garantía"
            )
            render_uploaded_files_preview("Evidencias seleccionadas", comprobante_cliente)

        # Confirmación antes de registrar
        confirmation_detail = ""
        if tipo_envio not in ["🔁 Devolución", "🛠 Garantía"] and fecha_entrega:
            fecha_entrega_texto = fecha_entrega.strftime("%d/%m/%Y") if hasattr(fecha_entrega, "strftime") else str(fecha_entrega)
            confirmation_detail += f" | Fecha de Entrega Seleccionada: {fecha_entrega_texto}"

        if tipo_envio == "📍 Pedido Local":
            turno_local = subtipo_local if subtipo_local else "Sin turno"
            confirmation_detail += f" | Turno: {turno_local}"

        st.info(f"✅ Tipo de envío seleccionado: {tipo_envio}{confirmation_detail}")

        # -------------------------------
        # SECCIÓN DE ESTADO DE PAGO (dentro del form para evitar recargas al adjuntar archivos)
        # -------------------------------
        if tipo_envio in ["🚚 Pedido Foráneo", "🏙️ Pedido CDMX"]:
            st.markdown("---")
            st.subheader("💰 Estado de Pago")
            opciones_estado_pago = (
                ["🎟️ No Aplica", "🔴 No Pagado", "✅ Pagado"]
                if registrar_nota_venta
                else ["🔴 No Pagado", "✅ Pagado", "💳 CREDITO"]
            )
            if st.session_state.get("estado_pago") not in opciones_estado_pago:
                st.session_state["estado_pago"] = opciones_estado_pago[0]

            estado_pago = st.selectbox(
                "Estado de Pago",
                opciones_estado_pago,
                index=0,
                key="estado_pago",
            )

            requiere_captura_pago = estado_pago == "✅ Pagado"

            comprobante_pago_files = st.file_uploader(
                "💲 Comprobante(s) de Pago",
                type=["pdf", "jpg", "jpeg", "png"],
                accept_multiple_files=True,
                key="comprobante_uploader_final"
            )
            render_uploaded_files_preview("Comprobantes de pago seleccionados", comprobante_pago_files)

            if requiere_captura_pago:
                st.warning("⚠️ Estado en PAGADO: debes adjuntar al menos un comprobante antes de registrar el pedido.")
            else:
                st.caption("ℹ️ Los Comprobantes son obligatorios cuando el estado sea '✅ Pagado'.")

            with st.expander(
                "🧾 Detalles del Pago (opcional)",
                expanded=expand_payment_details_default,
            ):
                col1, col2, col3 = st.columns(3)
                with col1:
                    fecha_pago = st.date_input("📅 Fecha del Pago", value=datetime.today().date(), key="fecha_pago_input")
                with col2:
                    forma_pago = st.selectbox("💳 Forma de Pago", [
                        "Transferencia", "Depósito en Efectivo", "Tarjeta de Débito", "Tarjeta de Crédito", "Cheque"
                    ], key="forma_pago_input")
                with col3:
                    monto_pago = st.number_input("💲 Monto del Pago", min_value=0.0, format="%.2f", key="monto_pago_input")

                col4, col5 = st.columns(2)
                with col4:
                    if forma_pago in ["Tarjeta de Débito", "Tarjeta de Crédito"]:
                        terminal = st.selectbox(
                            "🏧 Terminal",
                            [
                                "BANORTE",
                                "AFIRME",
                                "VELPAY",
                                "CLIP",
                                "PAYPAL",
                                "BBVA",
                                "CONEKTA",
                                "MERCADO PAGO",
                            ],
                            key="terminal_input",
                        )
                        banco_destino = ""
                    else:
                        banco_destino = st.selectbox("🏦 Banco Destino", ["BANORTE", "BANAMEX", "AFIRME", "BANCOMER OP", "BANCOMER CURSOS"], key="banco_destino_input")
                        terminal = ""
                with col5:
                    referencia_pago = st.text_input("🔢 Referencia (opcional)", key="referencia_pago_input")

        # AL FINAL DEL FORMULARIO: botón submit
        submit_button = st.form_submit_button(
            "✅ Registrar Pedido",
            disabled=st.session_state.get("pedido_submit_disabled", False) or has_pending_submission,
            on_click=backup_tab1_form_state_for_retry,
        )

    should_process_submission = submit_button
    if route_post_confirm_notice:
        st.session_state.pop(LOCAL_ROUTE_POST_CONFIRM_NOTICE_KEY, None)

    if usa_hoja_ruta_local and not is_local_pasa_bodega:
        route_template_path = Path("plantillas") / "FORMATO DE ENTREGA LOCAL limpia.xlsx"
        selected_history_row = parse_sheet_row_number(st.session_state.get("local_route_selected_history_row"))
        current_folio_for_route = (
            nota_venta.strip()
            if registrar_nota_venta and isinstance(nota_venta, str)
            else folio_factura_input_value.strip()
        )
        current_route_payload = build_local_route_payload(
            fecha_entrega=fecha_entrega,
            registro_cliente=registro_cliente,
            subtipo_local=subtipo_local,
            hora_entrega_manual=local_route_hora_entrega,
            recibe=local_route_recibe,
            referencias_hoja_ruta=local_route_referencias,
            calle_no=local_route_calle_no,
            tipo_inmueble=local_route_tipo_inmueble,
            interior=local_route_interior,
            acceso_privada=local_route_acceso_privada,
            colonia=local_route_colonia,
            municipio=local_route_municipio,
            cp=local_route_cp,
            telefonos=local_route_telefonos,
            estado_pago=estado_pago,
            forma_pago=local_route_forma_pago,
            vendedor=vendedor,
            total_factura=local_route_total_factura,
            adeudo_anterior=local_route_adeudo_anterior,
            folio=current_folio_for_route,
        )
        current_cliente_local_record = build_clientes_locales_record_from_form()

        if update_client_history_button:
            if selected_history_row is None:
                st.warning("⚠️ Selecciona un cliente del historial antes de actualizar su información.")
            elif not current_cliente_local_record["Cliente"]:
                st.warning("⚠️ Captura el nombre del cliente antes de actualizar el historial.")
            else:
                try:
                    update_existing_cliente_local(selected_history_row, current_cliente_local_record)
                    st.success("✅ La información histórica del cliente fue actualizada correctamente.")
                except Exception as e:
                    st.error(f"❌ No se pudo actualizar Clientes_Locales: {e}")

        if confirm_route_button:
            st.session_state[LOCAL_ROUTE_CONFIRMED_PAYLOAD_KEY] = current_route_payload
            st.session_state[LOCAL_ROUTE_CONFIRMED_AT_KEY] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            route_generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            generated_route_file_data, route_filename = build_local_route_file_from_payload(
                route_template_path,
                current_route_payload,
            )
            if not generated_route_file_data:
                st.session_state.pop(LOCAL_ROUTE_GENERATED_FILE_KEY, None)
                st.session_state.pop(LOCAL_ROUTE_GENERATED_FILENAME_KEY, None)
                st.session_state.pop(LOCAL_ROUTE_GENERATED_AT_KEY, None)
            else:
                st.session_state[LOCAL_ROUTE_GENERATED_FILE_KEY] = generated_route_file_data
                st.session_state[LOCAL_ROUTE_GENERATED_FILENAME_KEY] = route_filename
                st.session_state[LOCAL_ROUTE_GENERATED_AT_KEY] = route_generated_at

            st.session_state[LOCAL_ROUTE_POST_CONFIRM_NOTICE_KEY] = {
                "filename": route_filename,
                "confirmed_at": route_generated_at,
            }
            if "route_notice_placeholder" in locals():
                with route_notice_placeholder.container():
                    st.success("✅ Hoja de ruta actualizada correctamente.")
                    if route_filename:
                        st.caption(f"📎 Hoja de ruta generada: `{route_filename}`")

        confirmed_route_payload = st.session_state.get(LOCAL_ROUTE_CONFIRMED_PAYLOAD_KEY)
        confirmed_route_timestamp = st.session_state.get(LOCAL_ROUTE_CONFIRMED_AT_KEY, "")
        route_missing_fields = get_local_route_missing_fields(current_route_payload)

        if (
            not tab1_special_shipping
            and subtipo_local not in ["☀️ Local Mañana", "🌙 Local Tarde"]
            and not str(local_route_hora_entrega or "").strip()
        ):
            st.warning(
                "⚠️ Selecciona `HORA DE ENTREGA` para personalizar este turno. "
                "Si lo dejas vacío, se aplica la lógica automática por turno (en 🏙️ Local Mty quedará como `POR DEFINIR`)."
            )

        st.markdown("---")
        st.subheader("📄 Vista previa de hoja de ruta local")

        if confirmed_route_payload:
            confirmed_missing_fields = get_local_route_missing_fields(confirmed_route_payload)
            generated_route_file_data = st.session_state.get(LOCAL_ROUTE_GENERATED_FILE_KEY)
            generated_route_filename = st.session_state.get(LOCAL_ROUTE_GENERATED_FILENAME_KEY, "")
            resumen_items = [
                f"Cliente: {confirmed_route_payload.get('cliente') or 'N/A'}",
                f"Folio: {confirmed_route_payload.get('folio') or 'N/A'}",
                f"Recibe: {confirmed_route_payload.get('recibe') or 'N/A'}",
                f"Dirección: {confirmed_route_payload.get('calle_no') or 'N/A'}",
                f"Municipio: {confirmed_route_payload.get('municipio') or 'N/A'}",
                f"Teléfonos: {confirmed_route_payload.get('telefonos') or 'N/A'}",
                f"Referencias: {confirmed_route_payload.get('referencias') or 'N/A'}",
                f"Forma de pago: {confirmed_route_payload.get('forma_pago') or 'N/A'}",
                f"Total factura: {confirmed_route_payload.get('total_factura')}",
                f"Adeudo anterior: {confirmed_route_payload.get('adeudo_anterior')}",
                f"Gran total: {confirmed_route_payload.get('gran_total')}",
            ]
            st.caption(
                f"Última vista previa generada: {confirmed_route_timestamp or 'Sin fecha'} | "
                f"Horario asignado: {confirmed_route_payload['hora_entrega']} | "
                f"Día de entrega: {confirmed_route_payload['dia_entrega']} | "
                f"Gran total a cobrar: {confirmed_route_payload['gran_total']}"
            )
            st.caption(" | ".join(resumen_items))

            if generated_route_file_data and generated_route_filename:
                try:
                    generated_route_bytes = base64.b64decode(
                        generated_route_file_data.get("content_b64", "")
                    )
                except Exception:
                    generated_route_bytes = b""

                if generated_route_bytes:
                    st.download_button(
                        label="📥 Descargar Excel generado",
                        data=generated_route_bytes,
                        file_name=generated_route_filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="local_route_download_generated_excel",
                        help="Descarga la vista previa del Excel local para revisarla antes de subir el pedido.",
                        on_click="ignore",
                    )
                else:
                    st.caption("No fue posible recuperar el Excel generado para la vista previa.")

            if confirmed_missing_fields:
                st.caption("Faltan datos por revisar en la vista previa: " + ", ".join(confirmed_missing_fields))

            if not route_template_path.exists():
                st.error(f"No se encontró la plantilla de hoja de ruta en: {route_template_path}")

        if confirmed_route_payload and current_route_payload != confirmed_route_payload:
            st.warning(
                "⚠️ Hiciste cambios después de la última vista previa. "
                "Si quieres revisar el Excel antes de enviar, vuelve a presionar "
                "'Ver / actualizar vista previa del Excel (opcional)'. Al registrar el pedido se adjuntará automáticamente "
                "la versión más reciente con los datos actuales."
            )

    else:
        st.session_state.pop(LOCAL_ROUTE_GENERATED_FILE_KEY, None)
        st.session_state.pop(LOCAL_ROUTE_GENERATED_FILENAME_KEY, None)
        st.session_state.pop(LOCAL_ROUTE_GENERATED_AT_KEY, None)

    if submit_button:
        # Si el usuario envía un nuevo pedido, limpia feedback anterior para
        # evitar confusión visual con mensajes de un envío pasado.
        st.session_state.pop("pedido_submission_status", None)
        st.session_state.pop("pedido_status_toast_event_id", None)
        st.session_state[TAB1_SCROLL_RESTORE_FLAG_KEY] = True
        st.session_state["current_tab_index"] = TAB_INDEX_TAB1
        st.query_params.update({"tab": "0"})
        st.session_state["pedido_submit_disabled"] = True
        st.session_state["pedido_submit_disabled_at"] = time.time()

    if not should_process_submission and pending_submission_record:
        retry_at = float(pending_submission_record.get("next_retry_at", 0) or 0)
        attempts_done = int(pending_submission_record.get("attempts", 0) or 0)
        remaining_seconds = max(0, int(retry_at - time.time()))

        st.warning(
            "⚠️ Hay un pedido pendiente de reintento. Para evitar duplicados se bloqueó temporalmente el botón de registrar."
        )
        action_col1, action_col2 = st.columns(2)
        with action_col1:
            retry_now = st.button("🔄 Reintentar ahora", key="retry_pending_now")
        with action_col2:
            cancel_pending = st.button("🗑️ Cancelar pedido pendiente", key="cancel_pending_retry")

        if cancel_pending:
            clear_pending_submission(pending_cache_key)
            set_pedido_submission_status(
                "warning",
                "⚠️ Se canceló el pedido pendiente. Ya puedes capturar y enviar uno nuevo.",
            )
            rerun_with_pedido_loading("⏳ Cancelando pedido pendiente...")

        if retry_now:
            submission_payload_override = pending_submission_record.get("payload", {}) or {}
            should_process_submission = True
            st.info(
                f"🔄 Reintentando ahora el pedido pendiente (intento #{attempts_done + 1})."
            )
        elif remaining_seconds <= 0:
            submission_payload_override = pending_submission_record.get("payload", {}) or {}
            should_process_submission = True
            st.info(
                f"🔄 Reintentando automáticamente el último pedido fallido (intento #{attempts_done + 1})."
            )
        else:
            st.info(
                f"⏱️ Reintento automático en {remaining_seconds}s (intentos previos: {attempts_done})."
            )
            time.sleep(min(1, remaining_seconds))
            st.rerun()

    if not registrar_nota_venta:
        nota_venta = ""
        motivo_nota_venta = ""

    folio_factura = (
        nota_venta.strip() if registrar_nota_venta and isinstance(nota_venta, str) else ""
    )
    if not folio_factura:
        folio_factura = (
            folio_factura_input_value.strip()
            if isinstance(folio_factura_input_value, str)
            else ""
        )
    motivo_nota_venta = (
        motivo_nota_venta.strip()
        if registrar_nota_venta and isinstance(motivo_nota_venta, str)
        else ""
    )
    has_new_capture_signal = any(
        [
            bool(str(registro_cliente or "").strip()),
            bool(str(folio_factura_input_value or "").strip()),
            bool(str(comentario or "").strip()),
            bool(uploaded_files),
            bool(comprobante_pago_files),
            bool(comprobante_cliente),
        ]
    )

    st.markdown(f"<div id='{TAB1_FEEDBACK_ANCHOR_ID}'></div>", unsafe_allow_html=True)

    message_container = st.container()

    with message_container:
        loading_message = st.session_state.pop("pedido_submission_loading_message", None)
        if loading_message:
            st.info(loading_message)

        status_data = st.session_state.get("pedido_submission_status")
        if status_data:
            if status_data.get("status") == "success" and has_new_capture_signal:
                st.session_state.pop("pedido_submission_status", None)
                st.session_state.pop("pedido_status_toast_event_id", None)
                status_data = None
            if status_data:
                status_age_seconds = time.time() - float(status_data.get("created_at", 0) or 0)
                if status_age_seconds > PEDIDO_STATUS_MAX_AGE_SECONDS:
                    st.session_state.pop("pedido_submission_status", None)
                    st.session_state.pop("pedido_status_toast_event_id", None)
                    status_data = None
        if st.session_state.get(TAB1_SCROLL_RESTORE_FLAG_KEY, False) and (loading_message or status_data):
            scroll_to_tab1_feedback_section()
            st.session_state[TAB1_SCROLL_RESTORE_FLAG_KEY] = False
        if status_data:
            event_id = status_data.get("event_id")
            status = status_data.get("status")
            detail = status_data.get("detail")
            attachments = status_data.get("attachments") or []

            # Permite registrar otro pedido sin depender del botón "Aceptar".
            # El estado de bloqueo se usa solo mientras el envío está en curso.
            if st.session_state.get("pedido_submit_disabled") and status in {"success", "warning", "error"}:
                st.session_state["pedido_submit_disabled"] = False
                st.session_state.pop("pedido_submit_disabled_at", None)

            should_toast = (
                bool(event_id)
                and st.session_state.get("pedido_status_toast_event_id") != event_id
            )

            if status == "success":
                st.success(status_data.get("message", "✅ Pedido registrado correctamente."))
                if should_toast:
                    cliente_toast = str(status_data.get("client_name", "")).strip()
                    mensaje_toast = (
                        f"✅ Pedido de {cliente_toast} registrado correctamente"
                        if cliente_toast
                        else "✅ Pedido registrado correctamente"
                    )
                    st.toast(mensaje_toast)
                    st.session_state["pedido_status_toast_event_id"] = event_id
                if attachments:
                    st.info("📎 Archivos subidos: " + ", ".join(os.path.basename(url) for url in attachments))
                if detail:
                    st.write(detail)
                if status_data.get("missing_attachments_warning"):
                    st.warning("⚠️ Pedido registrado sin archivos adjuntos.")
            elif status == "warning":
                st.warning(status_data.get("message", "⚠️ Revisa los campos obligatorios."))
                if should_toast:
                    st.toast("⚠️ Revisa los datos antes de reenviar")
                    st.session_state["pedido_status_toast_event_id"] = event_id
                if detail:
                    st.write(detail)
            else:
                error_message = status_data.get("message", "❌ Falla al subir el pedido.")
                if detail:
                    error_message = f"{error_message}\n\n🔍 Detalle: {detail}"
                st.error(error_message)
                if should_toast:
                    st.toast("❌ Error al registrar el pedido")
                    st.session_state["pedido_status_toast_event_id"] = event_id

            def clear_pedido_status_message() -> None:
                """Limpia el aviso y prepara el formulario para capturar un pedido nuevo."""
                st.session_state[TAB1_SCROLL_RESTORE_FLAG_KEY] = False
                reset_tab1_form_state()
                st.session_state["last_selected_vendedor"] = VENDEDOR_NOMBRE_POR_ID.get(
                    normalize_vendedor_id(st.session_state.get("id_vendedor", "")),
                    TAB1_VENDOR_EMPTY_OPTION,
                )
                st.session_state.pop("pedido_submission_status", None)
                st.session_state["pedido_submit_disabled"] = False
                st.session_state.pop("pedido_submit_disabled_at", None)

            if status == "success":
                st.button(
                    "Limpiar mensaje de éxito",
                    key="acknowledge_pedido_status",
                    on_click=clear_pedido_status_message,
                )


    # -------------------------------
    # Registro del Pedido
    # -------------------------------
    if should_process_submission:
        st.session_state[TAB1_SCROLL_RESTORE_FLAG_KEY] = True
        st.info("⏳ Registrando pedido, espera la confirmación final...")
        try:
            pedido_id = ""
            hora_registro = ""
            s3_prefix = ""
            if submission_payload_override:
                vendedor = submission_payload_override.get("vendedor", vendedor)
                registro_cliente = submission_payload_override.get("registro_cliente", registro_cliente)
                numero_cliente_rfc = submission_payload_override.get("numero_cliente_rfc", numero_cliente_rfc)
                folio_factura = submission_payload_override.get("folio_factura", folio_factura)
                folio_factura_error = submission_payload_override.get("folio_factura_error", folio_factura_error)
                motivo_nota_venta = submission_payload_override.get("motivo_nota_venta", motivo_nota_venta)
                tipo_envio = submission_payload_override.get("tipo_envio", tipo_envio)
                tipo_envio_excel = submission_payload_override.get("tipo_envio_excel", tipo_envio_excel)
                tipo_envio_original = submission_payload_override.get("tipo_envio_original", tipo_envio_original)
                estatus_origen_factura = submission_payload_override.get("estatus_origen_factura", estatus_origen_factura)
                aplica_pago = submission_payload_override.get("aplica_pago", aplica_pago)
                resultado_esperado = submission_payload_override.get("resultado_esperado", resultado_esperado)
                material_devuelto = submission_payload_override.get("material_devuelto", material_devuelto)
                motivo_detallado = submission_payload_override.get("motivo_detallado", motivo_detallado)
                area_responsable = submission_payload_override.get("area_responsable", area_responsable)
                nombre_responsable = submission_payload_override.get("nombre_responsable", nombre_responsable)
                monto_devuelto = float(submission_payload_override.get("monto_devuelto", monto_devuelto) or 0)
                g_resultado_esperado = submission_payload_override.get("g_resultado_esperado", g_resultado_esperado)
                g_descripcion_falla = submission_payload_override.get("g_descripcion_falla", g_descripcion_falla)
                g_piezas_afectadas = submission_payload_override.get("g_piezas_afectadas", g_piezas_afectadas)
                g_monto_estimado = float(submission_payload_override.get("g_monto_estimado", g_monto_estimado) or 0)
                g_area_responsable = submission_payload_override.get("g_area_responsable", g_area_responsable)
                g_nombre_responsable = submission_payload_override.get("g_nombre_responsable", g_nombre_responsable)
                g_numero_serie = submission_payload_override.get("g_numero_serie", g_numero_serie)
                g_fecha_compra_str = submission_payload_override.get("g_fecha_compra")
                if g_fecha_compra_str:
                    try:
                        g_fecha_compra = datetime.strptime(g_fecha_compra_str, "%Y-%m-%d").date()
                    except ValueError:
                        g_fecha_compra = None
                direccion_guia_retorno = submission_payload_override.get("direccion_guia_retorno", direccion_guia_retorno)
                direccion_envio_destino = submission_payload_override.get("direccion_envio_destino", direccion_envio_destino)
                estado_pago = submission_payload_override.get("estado_pago", estado_pago)
                fecha_pago = submission_payload_override.get("fecha_pago", fecha_pago)
                forma_pago = submission_payload_override.get("forma_pago", forma_pago)
                terminal = submission_payload_override.get("terminal", terminal)
                banco_destino = submission_payload_override.get("banco_destino", banco_destino)
                monto_pago = float(submission_payload_override.get("monto_pago", monto_pago) or 0)
                referencia_pago = submission_payload_override.get("referencia_pago", referencia_pago)
                comentario = submission_payload_override.get("comentario", comentario)
                subtipo_local = submission_payload_override.get("subtipo_local", subtipo_local)
                local_route_hora_entrega = submission_payload_override.get("local_route_hora_entrega", local_route_hora_entrega)
                local_route_recibe = submission_payload_override.get("local_route_recibe", local_route_recibe)
                local_route_calle_no = submission_payload_override.get("local_route_calle_no", local_route_calle_no)
                local_route_tipo_inmueble = submission_payload_override.get("local_route_tipo_inmueble", local_route_tipo_inmueble)
                local_route_acceso_privada = submission_payload_override.get("local_route_acceso_privada", local_route_acceso_privada)
                local_route_municipio = submission_payload_override.get("local_route_municipio", local_route_municipio)
                local_route_telefonos = submission_payload_override.get("local_route_telefonos", local_route_telefonos)
                local_route_interior = submission_payload_override.get("local_route_interior", local_route_interior)
                local_route_colonia = submission_payload_override.get("local_route_colonia", local_route_colonia)
                local_route_cp = submission_payload_override.get("local_route_cp", local_route_cp)
                local_route_forma_pago = submission_payload_override.get("local_route_forma_pago", local_route_forma_pago)
                local_route_total_factura = float(submission_payload_override.get("local_route_total_factura", local_route_total_factura) or 0)
                local_route_adeudo_anterior = float(submission_payload_override.get("local_route_adeudo_anterior", local_route_adeudo_anterior) or 0)
                local_route_referencias = submission_payload_override.get("local_route_referencias", local_route_referencias)
                fecha_entrega_str = submission_payload_override.get("fecha_entrega")
                if fecha_entrega_str:
                    try:
                        fecha_entrega = datetime.strptime(fecha_entrega_str, "%Y-%m-%d").date()
                    except ValueError:
                        fecha_entrega = datetime.now().date()
                pedido_id = str(submission_payload_override.get("pedido_id", "") or "").strip()
                hora_registro = str(submission_payload_override.get("hora_registro", "") or "").strip()
                s3_prefix = str(submission_payload_override.get("s3_prefix", "") or "").strip()
                uploaded_files = _deserialize_uploaded_files(submission_payload_override.get("uploaded_files"))
                comprobante_pago_files = _deserialize_uploaded_files(submission_payload_override.get("comprobante_pago_files"))
                comprobante_cliente = _deserialize_uploaded_files(submission_payload_override.get("comprobante_cliente"))
                auto_route_files = _deserialize_uploaded_files(submission_payload_override.get("auto_route_files"))
            else:
                auto_route_files = _deserialize_uploaded_files(
                    [st.session_state.get(LOCAL_ROUTE_GENERATED_FILE_KEY)]
                    if usa_hoja_ruta_local and st.session_state.get(LOCAL_ROUTE_GENERATED_FILE_KEY)
                    else []
                )

            if usa_hoja_ruta_local:
                route_template_path = Path("plantillas") / "FORMATO DE ENTREGA LOCAL limpia.xlsx"
                current_route_payload_for_submission = build_local_route_payload(
                    fecha_entrega=fecha_entrega,
                    registro_cliente=registro_cliente,
                    subtipo_local=subtipo_local,
                    hora_entrega_manual=local_route_hora_entrega,
                    recibe=local_route_recibe,
                    referencias_hoja_ruta=local_route_referencias,
                    calle_no=local_route_calle_no,
                    tipo_inmueble=local_route_tipo_inmueble,
                    interior=local_route_interior,
                    acceso_privada=local_route_acceso_privada,
                    colonia=local_route_colonia,
                    municipio=local_route_municipio,
                    cp=local_route_cp,
                    telefonos=local_route_telefonos,
                    estado_pago=estado_pago,
                    forma_pago=local_route_forma_pago,
                    vendedor=vendedor,
                    total_factura=local_route_total_factura,
                    adeudo_anterior=local_route_adeudo_anterior,
                    folio=folio_factura,
                )
                route_missing_fields_for_submission = get_local_route_missing_fields(
                    current_route_payload_for_submission
                )
                if route_missing_fields_for_submission:
                    clear_pending_submission(pending_cache_key)
                    set_pedido_submission_status(
                        "warning",
                        "⚠️ El pedido local no se subió. Completa los datos obligatorios de la hoja de ruta.",
                        "Faltan: " + ", ".join(route_missing_fields_for_submission),
                    )
                    st.session_state["pedido_submit_disabled"] = False
                    st.session_state.pop("pedido_submit_disabled_at", None)
                    rerun_with_pedido_loading()

                generated_route_file_data, generated_route_filename = build_local_route_file_from_payload(
                    route_template_path,
                    current_route_payload_for_submission,
                )
                if not generated_route_file_data:
                    clear_pending_submission(pending_cache_key)
                    set_pedido_submission_status(
                        "error",
                        "❌ El pedido local no se subió.",
                        f"No se encontró la plantilla de hoja de ruta en: {route_template_path}",
                    )
                    st.session_state["pedido_submit_disabled"] = False
                    st.session_state.pop("pedido_submit_disabled_at", None)
                    rerun_with_pedido_loading()

                st.session_state[LOCAL_ROUTE_CONFIRMED_PAYLOAD_KEY] = current_route_payload_for_submission
                st.session_state[LOCAL_ROUTE_CONFIRMED_AT_KEY] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                st.session_state[LOCAL_ROUTE_GENERATED_FILE_KEY] = generated_route_file_data
                st.session_state[LOCAL_ROUTE_GENERATED_FILENAME_KEY] = generated_route_filename
                st.session_state[LOCAL_ROUTE_GENERATED_AT_KEY] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                auto_route_files = _deserialize_uploaded_files([generated_route_file_data])

            if not vendedor or not registro_cliente:
                set_pedido_submission_status(
                    "warning",
                    "⚠️ El pedido no se subió. Completa los campos obligatorios e inténtalo de nuevo.",
                )
                st.session_state["pedido_submit_disabled"] = False
                st.session_state.pop("pedido_submit_disabled_at", None)
                rerun_with_pedido_loading("⏳ Recargando formulario...")

            if not submission_payload_override:
                pedido_id, hora_registro, s3_prefix = build_submission_identity()
                payload_to_retry = {
                    "pedido_id": pedido_id,
                    "hora_registro": hora_registro,
                    "s3_prefix": s3_prefix,
                    "tipo_envio": tipo_envio,
                    "tipo_envio_excel": tipo_envio_excel,
                    "vendedor": vendedor,
                    "registro_cliente": registro_cliente,
                    "numero_cliente_rfc": numero_cliente_rfc,
                    "folio_factura": folio_factura,
                    "folio_factura_error": folio_factura_error,
                    "motivo_nota_venta": motivo_nota_venta,
                    "tipo_envio_original": tipo_envio_original,
                    "estatus_origen_factura": estatus_origen_factura,
                    "aplica_pago": aplica_pago,
                    "resultado_esperado": resultado_esperado,
                    "material_devuelto": material_devuelto,
                    "motivo_detallado": motivo_detallado,
                    "area_responsable": area_responsable,
                    "nombre_responsable": nombre_responsable,
                    "monto_devuelto": monto_devuelto,
                    "g_resultado_esperado": g_resultado_esperado,
                    "g_descripcion_falla": g_descripcion_falla,
                    "g_piezas_afectadas": g_piezas_afectadas,
                    "g_monto_estimado": g_monto_estimado,
                    "g_area_responsable": g_area_responsable,
                    "g_nombre_responsable": g_nombre_responsable,
                    "g_numero_serie": g_numero_serie,
                    "g_fecha_compra": g_fecha_compra.strftime('%Y-%m-%d') if g_fecha_compra else "",
                    "direccion_guia_retorno": direccion_guia_retorno,
                    "direccion_envio_destino": direccion_envio_destino,
                    "estado_pago": estado_pago,
                    "fecha_pago": fecha_pago if isinstance(fecha_pago, str) else (fecha_pago.strftime('%Y-%m-%d') if fecha_pago else ""),
                    "forma_pago": forma_pago,
                    "terminal": terminal,
                    "banco_destino": banco_destino,
                    "monto_pago": monto_pago,
                    "referencia_pago": referencia_pago,
                    "comentario": comentario,
                    "subtipo_local": subtipo_local,
                    "local_route_hora_entrega": local_route_hora_entrega,
                    "local_route_recibe": local_route_recibe,
                    "local_route_calle_no": local_route_calle_no,
                    "local_route_tipo_inmueble": local_route_tipo_inmueble,
                    "local_route_acceso_privada": local_route_acceso_privada,
                    "local_route_municipio": local_route_municipio,
                    "local_route_telefonos": local_route_telefonos,
                    "local_route_interior": local_route_interior,
                    "local_route_colonia": local_route_colonia,
                    "local_route_cp": local_route_cp,
                    "local_route_forma_pago": local_route_forma_pago,
                    "local_route_total_factura": local_route_total_factura,
                    "local_route_adeudo_anterior": local_route_adeudo_anterior,
                    "local_route_referencias": local_route_referencias,
                    "fecha_entrega": fecha_entrega.strftime('%Y-%m-%d') if fecha_entrega else "",
                    "uploaded_files": _serialize_uploaded_files(uploaded_files),
                    "comprobante_pago_files": _serialize_uploaded_files(comprobante_pago_files),
                    "comprobante_cliente": _serialize_uploaded_files(comprobante_cliente),
                    "auto_route_files": _serialize_uploaded_files(auto_route_files),
                }
                save_pending_submission(pending_cache_key, payload_to_retry)
            else:
                if not all([pedido_id, hora_registro, s3_prefix]):
                    pedido_id, hora_registro, s3_prefix = build_submission_identity()

            pedido_sin_adjuntos = not (
                uploaded_files or comprobante_pago_files or comprobante_cliente or auto_route_files
            )
            aviso_estado_pago_auto = ""

            pedidos_con_estado_pago = [
                "🚚 Pedido Foráneo",
                "🏙️ Pedido CDMX",
                "📍 Pedido Local",
                "🎓 Cursos y Eventos",
            ]
            if tipo_envio == "🔁 Devolución" and tipo_envio_original == "📍 Local":
                pedidos_con_estado_pago.append("🔁 Devolución")

            if (
                tipo_envio in pedidos_con_estado_pago
                and comprobante_pago_files
                and estado_pago != "✅ Pagado"
            ):
                estado_pago = "✅ Pagado"
                aviso_estado_pago_auto = (
                    "ℹ️ Se detectó al menos un comprobante de pago y el pedido fue marcado "
                    "automáticamente como '✅ Pagado'."
                )
                st.info(aviso_estado_pago_auto)

            # Normalización de campos para Casos Especiales
            if tipo_envio == "🔁 Devolución":
                resultado_esperado = normalize_case_text(resultado_esperado)
                material_devuelto = normalize_case_text(material_devuelto)
                motivo_detallado = normalize_case_text(motivo_detallado)
                nombre_responsable = normalize_case_text(nombre_responsable)
            if tipo_envio == "🛠 Garantía":
                g_resultado_esperado = normalize_case_text(g_resultado_esperado)
                g_descripcion_falla = normalize_case_text(g_descripcion_falla)
                g_piezas_afectadas = normalize_case_text(g_piezas_afectadas)
                g_nombre_responsable = normalize_case_text(g_nombre_responsable)
                g_numero_serie = normalize_case_text(g_numero_serie)
            if tipo_envio in ["🔁 Devolución", "🛠 Garantía"]:
                direccion_guia_retorno = normalize_case_text(direccion_guia_retorno)
                direccion_envio_destino = normalize_case_text(direccion_envio_destino)

            # Validar comprobante de pago para tipos normales
            if (
                tipo_envio in pedidos_con_estado_pago
                and estado_pago == "✅ Pagado"
                and not comprobante_pago_files
            ):
                clear_pending_submission(pending_cache_key)
                set_pedido_submission_status(
                    "warning",
                    "⚠️ El pedido no se subió. Adjunta un comprobante si el pedido está marcado como pagado.",
                )
                st.session_state["pedido_submit_disabled"] = False
                st.session_state.pop("pedido_submit_disabled_at", None)
                rerun_with_pedido_loading()

            # Acceso a la hoja
            headers = []
            try:
                if tipo_envio in ["🔁 Devolución", "🛠 Garantía"]:
                    worksheet = get_worksheet_casos_especiales()
                    if worksheet is None:
                        set_pedido_submission_status(
                            "error",
                            "❌ Falla al subir el pedido.",
                            "No fue posible acceder a la hoja de casos especiales.",
                        )
                        rerun_with_pedido_loading()

                    headers = worksheet.row_values(1)
                    required_headers = ["Direccion_Guia_Retorno", "Direccion_Envio", "Estatus_OrigenF"]
                    missing_headers = [col for col in required_headers if col not in headers]
                    if missing_headers:
                        try:
                            new_headers = headers + missing_headers
                            worksheet.update('A1', [new_headers])
                            get_sheet_headers.clear()
                            headers = worksheet.row_values(1)
                        except Exception as header_error:
                            set_pedido_submission_status(
                                "error",
                                "❌ Falla al subir el pedido.",
                                f"No se pudieron preparar las columnas de direcciones: {header_error}",
                            )
                            rerun_with_pedido_loading()
                else:
                    worksheet = get_worksheet_operativa()
                    if worksheet is None:
                        set_pedido_submission_status(
                            "error",
                            "❌ Falla al subir el pedido.",
                            "No fue posible acceder a la hoja de pedidos.",
                        )
                        rerun_with_pedido_loading()
                    headers = worksheet.row_values(1)
                    required_headers = []
                    if tipo_envio == "🚚 Pedido Foráneo":
                        required_headers.append("Direccion_Guia_Retorno")
                    if required_headers:
                        missing_headers = [col for col in required_headers if col not in headers]
                        if missing_headers:
                            try:
                                new_headers = headers + missing_headers
                                worksheet.update('A1', [new_headers])
                                get_sheet_headers.clear()
                                headers = worksheet.row_values(1)
                            except Exception as header_error:
                                set_pedido_submission_status(
                                    "error",
                                    "❌ Falla al subir el pedido.",
                                    f"No se pudieron preparar las columnas de direcciones: {header_error}",
                                )
                                rerun_with_pedido_loading()

                if not headers:
                    set_pedido_submission_status(
                        "error",
                        "❌ Falla al subir el pedido.",
                        "La hoja de cálculo está vacía.",
                    )
                    rerun_with_pedido_loading()

                # Reutilizar identidad estable del envío actual para evitar duplicados en reintentos.
                if not all([pedido_id, hora_registro, s3_prefix]):
                    pedido_id, hora_registro, s3_prefix = build_submission_identity()

            except gspread.exceptions.APIError as e:
                if "RESOURCE_EXHAUSTED" in str(e):
                    set_pedido_submission_status(
                        "warning",
                        "⚠️ Cuota de Google Sheets alcanzada. Reintentando...",
                    )
                    st.cache_resource.clear()
                    time.sleep(6)
                    rerun_with_pedido_loading()
                else:
                    set_pedido_submission_status(
                        "error",
                        "❌ Falla al subir el pedido.",
                        f"Error al acceder a Google Sheets: {e}",
                    )
                    rerun_with_pedido_loading()

            adjuntos_urls = []
            try:
                adjuntos_urls.extend(
                    upload_files_or_fail(
                        uploaded_files,
                        s3_client,
                        S3_BUCKET_NAME,
                        s3_prefix,
                    )
                )
                adjuntos_urls.extend(
                    upload_files_or_fail(
                        comprobante_pago_files,
                        s3_client,
                        S3_BUCKET_NAME,
                        s3_prefix,
                    )
                )
                adjuntos_urls.extend(
                    upload_files_or_fail(
                        comprobante_cliente,
                        s3_client,
                        S3_BUCKET_NAME,
                        s3_prefix,
                    )
                )
                adjuntos_urls.extend(
                    upload_files_or_fail(
                        auto_route_files,
                        s3_client,
                        S3_BUCKET_NAME,
                        s3_prefix,
                    )
                )
            except Exception as e:
                schedule_pending_submission_retry(pending_cache_key)
                set_pedido_submission_status(
                    status="error",
                    message="❌ No se pudieron subir los archivos del pedido.",
                    detail=str(e),
                )
                rerun_with_pedido_loading()

            adjuntos_str = ", ".join(adjuntos_urls)

            # Mapeo de columnas a valores
            values = []
            for header in headers:
                if header == "ID_Pedido":
                    values.append(pedido_id)
                elif header == "Hora_Registro":
                    values.append(hora_registro)
                elif header.lower() == "id_vendedor":
                    values.append(st.session_state.get("id_vendedor", ""))
                elif header in ["Vendedor", "Vendedor_Registro"]:
                    values.append(vendedor)
                elif header in ["Cliente", "RegistroCliente"]:
                    values.append(registro_cliente)
                elif header == "Numero_Cliente_RFC":
                    if tipo_envio in ["🔁 Devolución", "🛠 Garantía"]:
                        values.append(numero_cliente_rfc)
                    else:
                        values.append("")
                elif header == "Folio_Factura":
                    values.append(folio_factura)  # en devoluciones es "Folio Nuevo" o Nota de Venta
                elif header == "Folio_Factura_Error":  # 🆕 mapeo adicional
                    values.append(folio_factura_error if tipo_envio == "🔁 Devolución" else "")
                elif header == "Motivo_NotaVenta":
                    values.append(motivo_nota_venta)
                elif header == "Tipo_Envio":
                    values.append(tipo_envio_excel)
                elif header == "Tipo_Envio_Original":
                    values.append(tipo_envio_original if tipo_envio == "🔁 Devolución" else "")
                elif header == "Estatus_OrigenF":
                    values.append(estatus_origen_factura if tipo_envio == "🔁 Devolución" else "")
                elif header == "Turno":
                    values.append(get_subtipo_local_excel_value(subtipo_local))
                elif header == "Fecha_Entrega":
                    if tipo_envio in ["🔁 Devolución", "🛠 Garantía"]:
                        values.append("")
                    else:
                        values.append(fecha_entrega.strftime('%Y-%m-%d'))
                elif header == "Comentario":
                    values.append(comentario)
                elif header == "Adjuntos":
                    values.append(adjuntos_str)
                elif header == "Adjuntos_Surtido":
                    values.append("")
                elif header == "Estado":
                    values.append("🟡 Pendiente")
                elif header == "Estado_Pago":
                    if tipo_envio in ["🚚 Pedido Foráneo", "🏙️ Pedido CDMX", "📍 Pedido Local"] or (
                        tipo_envio == "🔁 Devolución" and tipo_envio_original == "📍 Local"
                    ):
                        values.append(estado_pago)
                    else:
                        values.append("")
                elif header == "Aplica_Pago":
                    values.append("Sí" if aplica_pago == "Sí" else "No")
                elif header == "Fecha_Pago_Comprobante":
                    if tipo_envio in ["🚚 Pedido Foráneo", "🏙️ Pedido CDMX", "📍 Pedido Local"] or (
                        tipo_envio == "🔁 Devolución" and tipo_envio_original == "📍 Local"
                    ):
                        values.append(fecha_pago if isinstance(fecha_pago, str) else (fecha_pago.strftime('%Y-%m-%d') if fecha_pago else ""))
                    else:
                        values.append("")
                elif header == "Forma_Pago_Comprobante":
                    if tipo_envio in ["🚚 Pedido Foráneo", "🏙️ Pedido CDMX"]:
                        values.append(forma_pago)
                    elif tipo_envio == "📍 Pedido Local" or (tipo_envio == "🔁 Devolución" and tipo_envio_original == "📍 Local"):
                        values.append(local_route_forma_pago)
                    else:
                        values.append("")
                elif header == "Terminal":
                    if tipo_envio in ["🚚 Pedido Foráneo", "🏙️ Pedido CDMX", "📍 Pedido Local"] or (
                        tipo_envio == "🔁 Devolución" and tipo_envio_original == "📍 Local"
                    ):
                        values.append(terminal)
                    else:
                        values.append("")
                elif header == "Banco_Destino_Pago":
                    if tipo_envio in ["🚚 Pedido Foráneo", "🏙️ Pedido CDMX", "📍 Pedido Local"] or (
                        tipo_envio == "🔁 Devolución" and tipo_envio_original == "📍 Local"
                    ):
                        values.append(banco_destino)
                    else:
                        values.append("")
                elif header == "Monto_Comprobante":
                    if tipo_envio in ["🚚 Pedido Foráneo", "🏙️ Pedido CDMX"]:
                        values.append(f"{monto_pago:.2f}" if monto_pago > 0 else "")
                    elif tipo_envio == "📍 Pedido Local" or (tipo_envio == "🔁 Devolución" and tipo_envio_original == "📍 Local"):
                        monto_comprobante_local = float(local_route_total_factura or 0) + float(local_route_adeudo_anterior or 0)
                        values.append(f"{monto_comprobante_local:.2f}" if monto_comprobante_local > 0 else "")
                    else:
                        values.append("")
                elif header == "Referencia_Comprobante":
                    if tipo_envio in ["🚚 Pedido Foráneo", "🏙️ Pedido CDMX", "📍 Pedido Local"] or (
                        tipo_envio == "🔁 Devolución" and tipo_envio_original == "📍 Local"
                    ):
                        values.append(referencia_pago)
                    else:
                        values.append("")
                elif header in ["Fecha_Completado", "Hora_Proceso", "Modificacion_Surtido"]:
                    values.append("")

                # -------- Campos Casos Especiales (reutilizados) --------
                elif header == "Resultado_Esperado":
                    if tipo_envio == "🔁 Devolución":
                        values.append(resultado_esperado)
                    elif tipo_envio == "🛠 Garantía":
                        values.append(g_resultado_esperado)
                    else:
                        values.append("")
                elif header == "Material_Devuelto":
                    if tipo_envio == "🔁 Devolución":
                        values.append(material_devuelto)
                    elif tipo_envio == "🛠 Garantía":
                        values.append(g_piezas_afectadas)  # Reuso columna para piezas afectadas
                    else:
                        values.append("")
                elif header == "Monto_Devuelto":
                    if tipo_envio == "🔁 Devolución":
                        values.append(normalize_case_amount(monto_devuelto))
                    elif tipo_envio == "🛠 Garantía":
                        values.append(normalize_case_amount(g_monto_estimado))
                    else:
                        values.append("")
                elif header == "Motivo_Detallado":
                    if tipo_envio == "🔁 Devolución":
                        values.append(motivo_detallado)
                    elif tipo_envio == "🛠 Garantía":
                        values.append(g_descripcion_falla)
                    else:
                        values.append("")
                elif header == "Area_Responsable":
                    if tipo_envio == "🔁 Devolución":
                        values.append(area_responsable)
                    elif tipo_envio == "🛠 Garantía":
                        values.append(g_area_responsable)
                    else:
                        values.append("")
                elif header == "Nombre_Responsable":
                    if tipo_envio == "🔁 Devolución":
                        values.append(nombre_responsable)
                    elif tipo_envio == "🛠 Garantía":
                        values.append(g_nombre_responsable)
                    else:
                        values.append("")
                elif header == "Direccion_Guia_Retorno":
                    if tipo_envio in ["🔁 Devolución", "🛠 Garantía"]:
                        values.append(direccion_guia_retorno)
                    elif tipo_envio == "🚚 Pedido Foráneo" and direccion_guia_retorno.strip():
                        values.append(direccion_guia_retorno)
                    else:
                        values.append("")
                elif header == "Direccion_Envio":
                    if tipo_envio in ["🔁 Devolución", "🛠 Garantía"]:
                        values.append(direccion_envio_destino)
                    else:
                        values.append("")
                # -------- Opcionales si existen en la hoja --------
                elif header == "Numero_Serie":
                    values.append(g_numero_serie if tipo_envio == "🛠 Garantía" else "")
                elif header in ["Fecha_Compra", "FechaCompra"]:
                    if tipo_envio == "🛠 Garantía":
                        values.append(g_fecha_compra.strftime('%Y-%m-%d') if g_fecha_compra else "")
                    else:
                        values.append("")
                else:
                    values.append("")

            try:
                id_col_index = headers.index("ID_Pedido")
            except ValueError:
                set_pedido_submission_status(
                    "error",
                    "❌ Falla al subir el pedido.",
                    "No se encontró la columna ID_Pedido en la hoja.",
                )
                rerun_with_pedido_loading()

            try:
                with st.spinner("Registrando pedido en Google Sheets..."):
                    append_row_with_confirmation(
                        worksheet=worksheet,
                        values=values,
                        pedido_id=pedido_id,
                        id_col_index=id_col_index,
                    )
            except Exception as e:
                schedule_pending_submission_retry(pending_cache_key)
                set_pedido_submission_status(
                    "error",
                    "❌ Falla al subir el pedido.",
                    f"Error al registrar el pedido: {e}",
                )
                rerun_with_pedido_loading()

            cliente_local_history_notice = ""
            local_route_upload_notice = ""
            if usa_hoja_ruta_local and not is_local_pasa_bodega:
                try:
                    inserted, _history_message = upsert_cliente_local_if_missing(
                        build_clientes_locales_record_from_form()
                    )
                    if inserted:
                        cliente_local_history_notice = " Se agregó el cliente al historial local."
                except Exception as e:
                    cliente_local_history_notice = f" No se pudo actualizar Clientes_Locales: {e}"

                local_route_filename = str(
                    st.session_state.get(LOCAL_ROUTE_GENERATED_FILENAME_KEY, "") or ""
                ).strip()
                if local_route_filename:
                    local_route_upload_notice = (
                        f" 📎 La hoja de ruta local se generó y se adjuntó automáticamente: {local_route_filename}."
                    )
                else:
                    local_route_upload_notice = (
                        " 📎 La hoja de ruta local se generó y se adjuntó automáticamente."
                    )

            reset_tab1_form_state()
            id_vendedor_actual = str(st.session_state.get("id_vendedor", "")).strip()
            st.session_state["last_selected_vendedor"] = VENDEDOR_NOMBRE_POR_ID.get(
                normalize_vendedor_id(id_vendedor_actual),
                TAB1_VENDOR_EMPTY_OPTION,
            )
            id_vendedor_segment = (
                f" (ID vendedor: {id_vendedor_actual})" if id_vendedor_actual else ""
            )
            clear_order_related_caches()
            cliente_registrado = str(registro_cliente or "").strip()
            referencia_pedido = cliente_registrado or pedido_id
            set_pedido_submission_status(
                "success",
                f"✅ El pedido {referencia_pedido}{id_vendedor_segment} fue subido correctamente.",
                detail=(
                    f"{aviso_estado_pago_auto}{local_route_upload_notice}{cliente_local_history_notice}"
                ).strip(),
                attachments=adjuntos_urls,
                missing_attachments_warning=pedido_sin_adjuntos,
                client_name=cliente_registrado,
            )
            clear_pending_submission(pending_cache_key)
            if tab1_is_active and st.session_state.get("current_tab_index") == TAB_INDEX_TAB1:
                st.query_params.update({"tab": "0"})
            rerun_with_pedido_loading("⏳ Pedido registrado. Actualizando vista...")

        except Exception as e:
            schedule_pending_submission_retry(pending_cache_key)
            set_pedido_submission_status(
                "error",
                "❌ Falla al subir el pedido.",
                f"Error inesperado al registrar el pedido: {e}",
            )
            rerun_with_pedido_loading()



@st.cache_data(ttl=300)
def cargar_pedidos_combinados():
    """
    Carga y unifica pedidos de 'data_pedidos' y 'casos_especiales'.
    Devuelve un DataFrame con columna 'Fuente' indicando el origen.
    Garantiza columnas usadas por la UI (modificación de surtido, refacturación, folio error, documentos, etc.)
    y mapea Hoja_Ruta_Mensajero -> Adjuntos_Guia para homogeneizar.
    """
    client = build_gspread_client()
    sh = client.open_by_key(GOOGLE_SHEET_ID)

    # ---------------------------
    # data_pedidos
    # ---------------------------
    try:
        ws_datos = sh.worksheet(SHEET_PEDIDOS_OPERATIVOS)
        df_datos, headers_datos = load_sheet_records_with_row_numbers(ws_datos)
    except Exception:
        headers_datos = []
        df_datos = pd.DataFrame()

    if not df_datos.empty:
        # quita filas totalmente vacías en claves mínimas
        claves = ['ID_Pedido', 'Cliente', 'Folio_Factura']
        df_datos = df_datos.dropna(subset=claves, how='all')
        if 'ID_Pedido' in df_datos.columns:
            df_datos = df_datos[df_datos['ID_Pedido'].astype(str).str.strip().ne("")]

        # columnas que la UI puede usar desde data_pedidos
        needed_datos: list[str] = []
        needed_datos += [
            'ID_Pedido','Cliente','Folio_Factura','Vendedor_Registro','Estado','Hora_Registro','Turno','Fecha_Entrega',
            'Comentario','Estado_Pago','Motivo_NotaVenta',
            # archivos/adjuntos
            'Adjuntos','Adjuntos_Guia','Adjuntos_Surtido','Modificacion_Surtido',
            # refacturación
            'Refacturacion_Tipo','Refacturacion_Subtipo','Folio_Factura_Refacturada',
            # seguimiento de modificaciones
            'id_vendedor_Mod',
            # para homogeneidad con casos (puede venir vacío en datos)
            'Folio_Factura_Error','Estado_Caso','Numero_Cliente_RFC','Tipo_Envio','Tipo_Envio_Original',
            'Resultado_Esperado','Motivo_Detallado','Material_Devuelto','Monto_Devuelto',
            'Nota_Credito_URL','Documento_Adicional_URL','Comentarios_Admin_Devolucion',
            'Hoja_Ruta_Mensajero','Fecha_Recepcion_Devolucion','Hora_Proceso','Area_Responsable','Nombre_Responsable',
            'Direccion_Guia_Retorno','Direccion_Envio',
            # seguimiento
            'Seguimiento'
        ]
        for c in needed_datos:
            if c not in df_datos.columns:
                df_datos[c] = ""

        df_datos['Seguimiento'] = df_datos['Seguimiento'].fillna("")

        # asegura tipos string uniformes importantes
        for c in ['Tipo_Envio','Vendedor_Registro','Estado','Folio_Factura','Folio_Factura_Refacturada','id_vendedor_Mod']:
            if c in df_datos.columns:
                df_datos[c] = df_datos[c].astype(str)

        df_datos["Fuente"] = SHEET_PEDIDOS_OPERATIVOS

    # ---------------------------
    # casos_especiales
    # ---------------------------
    try:
        ws_casos = sh.worksheet("casos_especiales")
        df_casos, headers_casos = load_sheet_records_with_row_numbers(ws_casos)
    except Exception:
        headers_casos = []
        df_casos = pd.DataFrame()

    if not df_casos.empty:
        if 'ID_Pedido' in df_casos.columns:
            df_casos = df_casos[df_casos['ID_Pedido'].astype(str).str.strip().ne("")]
        else:
            df_casos["ID_Pedido"] = ""

        # columnas mínimas + TODAS las que usa la UI (incluye Garantías)
        base_cols = [
            'ID_Pedido','Cliente','Folio_Factura','Folio_Factura_Error','Estado','Tipo_Envio','Tipo_Envio_Original',
            'Turno','Fecha_Entrega','Hora_Registro','Hora_Proceso','Vendedor_Registro','Comentario','Estado_Pago',
            # adjuntos/guía/modificación
            'Adjuntos','Adjuntos_Guia','Hoja_Ruta_Mensajero',
            'Adjuntos_Surtido','Modificacion_Surtido',
            # cliente/estatus caso
            'Numero_Cliente_RFC','Estado_Caso',
            # refacturación
            'Refacturacion_Tipo','Refacturacion_Subtipo','Folio_Factura_Refacturada',
            # detalle del caso (dev/garantía)
            'Resultado_Esperado','Motivo_Detallado','Material_Devuelto','Monto_Devuelto',
            'Area_Responsable','Nombre_Responsable',
            'Direccion_Guia_Retorno','Direccion_Envio',
            # ⚙️ NUEVO: Garantías
            'Numero_Serie','Fecha_Compra',   # si tu hoja usa "FechaCompra", abajo lo normalizamos
            # recepción/cierre
            'Fecha_Recepcion_Devolucion','Estado_Recepcion',
            # documentos de cierre
            'Nota_Credito_URL','Documento_Adicional_URL','Comentarios_Admin_Devolucion',
            # seguimiento
            'Seguimiento'
        ]
        for c in base_cols:
            if c not in df_casos.columns:
                df_casos[c] = ""

        df_casos['Seguimiento'] = df_casos['Seguimiento'].fillna("")

        # Normalizar fecha de compra si el encabezado real es "FechaCompra"
        if 'Fecha_Compra' not in headers_casos and 'FechaCompra' in headers_casos:
            df_casos['Fecha_Compra'] = df_casos['FechaCompra']

        # Inferir Tipo_Envio desde Tipo_Caso si viene vacío
        if 'Tipo_Envio' in df_casos.columns:
            df_casos['Tipo_Envio'] = df_casos['Tipo_Envio'].astype(str)
        if 'Tipo_Envio' in df_casos.columns and df_casos['Tipo_Envio'].eq("").any():
            if 'Tipo_Caso' in df_casos.columns:
                def _infer_tipo_envio(row):
                    t_env = str(row.get("Tipo_Envio","")).strip()
                    if t_env:
                        return t_env
                    t_caso = str(row.get("Tipo_Caso","")).lower()
                    if t_caso.startswith("devol"):
                        return "🔁 Devolución"
                    if t_caso.startswith("garan"):
                        return "🛠 Garantía"
                    return "Caso especial"
                df_casos['Tipo_Envio'] = df_casos.apply(_infer_tipo_envio, axis=1)

        # Mapear Hoja_Ruta_Mensajero -> Adjuntos_Guia si esta última está vacía
        if 'Adjuntos_Guia' in df_casos.columns and 'Hoja_Ruta_Mensajero' in df_casos.columns:
            mask_vacios = df_casos['Adjuntos_Guia'].astype(str).str.strip().eq("")
            df_casos.loc[mask_vacios, 'Adjuntos_Guia'] = df_casos.loc[mask_vacios, 'Hoja_Ruta_Mensajero']

        # asegura tipos string uniformes importantes
        for c in ['Tipo_Envio','Vendedor_Registro','Estado','Folio_Factura','Folio_Factura_Error','Folio_Factura_Refacturada']:
            if c in df_casos.columns:
                df_casos[c] = df_casos[c].astype(str)

        df_casos["Fuente"] = "casos_especiales"

    # ---------------------------
    # Unir respetando columnas
    # ---------------------------
    if df_datos.empty and df_casos.empty:
        return pd.DataFrame()
    if df_datos.empty:
        return df_casos.copy()
    if df_casos.empty:
        return df_datos.copy()

    all_cols = list(set(df_datos.columns).union(set(df_casos.columns)))
    df_datos = df_datos.reindex(columns=all_cols, fill_value="")
    df_casos = df_casos.reindex(columns=all_cols, fill_value="")
    df_all = pd.concat([df_datos, df_casos], ignore_index=True)
    return df_all

# --- TAB VENTAS Y REPORTES (solo RUBEN67/JUAN24/FRANKO95) ---
if tab_ventas_reportes is not None:
    with tab_ventas_reportes:
        if TAB_INDEX_REPORTES is not None and default_tab == TAB_INDEX_REPORTES:
            st.session_state["current_tab_index"] = TAB_INDEX_REPORTES

        st.header("📊 Ventas y Reportes")
        st.caption("Pedidos registrados por RUBEN67, JUAN24 y FRANKO95.")

        try:
            df_ventas = cargar_pedidos_ventas_reportes()
        except Exception as e:
            st.error(f"❌ No se pudieron cargar los pedidos: {e}")
            df_ventas = pd.DataFrame()

        if df_ventas.empty:
            st.info("No hay pedidos para mostrar.")
        else:
            if "id_vendedor" not in df_ventas.columns:
                df_ventas["id_vendedor"] = ""

            df_ventas["id_vendedor_norm"] = df_ventas["id_vendedor"].apply(normalize_vendedor_id)
            df_ventas = df_ventas[df_ventas["id_vendedor_norm"].isin(LOCAL_TURNO_CDMX_IDS)].copy()

            columnas_reporte = [
                "Folio_Factura",
                "Cliente",
                "Monto_Comprobante",
                "Forma_Pago_Comprobante",
                "Vendedor_Registro",
                "Hora_Registro",
            ]
            for col in columnas_reporte:
                if col not in df_ventas.columns:
                    df_ventas[col] = ""

            df_ventas_base = df_ventas.copy()

            if "Hora_Registro" in df_ventas.columns:
                fecha_hora_registro = pd.to_datetime(df_ventas["Hora_Registro"], errors="coerce")
                df_ventas["fecha_hora_registro"] = fecha_hora_registro
                df_ventas["mes_registro"] = fecha_hora_registro.dt.to_period("M").astype("string")

                meses_disponibles = (
                    df_ventas.loc[df_ventas["mes_registro"].notna(), "mes_registro"]
                    .drop_duplicates()
                    .sort_values(ascending=False)
                    .tolist()
                )
                filtro_mes = st.selectbox(
                    "🗓️ Filtrar por mes",
                    ["Todos"] + meses_disponibles,
                    index=0,
                    help="Formato AAAA-MM (ejemplo: 2026-04).",
                )
                if filtro_mes != "Todos":
                    df_ventas = df_ventas[df_ventas["mes_registro"] == filtro_mes].copy()

                df_ventas = df_ventas.sort_values(
                    by="fecha_hora_registro",
                    ascending=False,
                )

            st.dataframe(
                df_ventas[columnas_reporte].reset_index(drop=True),
                use_container_width=True,
                hide_index=True,
            )
            st.caption(f"Total de pedidos encontrados: {len(df_ventas)}")

            ventas_excel_buffer = BytesIO()
            with pd.ExcelWriter(ventas_excel_buffer, engine="openpyxl") as writer:
                df_ventas[columnas_reporte].to_excel(writer, index=False, sheet_name="Ventas_Reportes")
            st.download_button(
                label="📥 Descargar ventas (Excel)",
                data=ventas_excel_buffer.getvalue(),
                file_name=f"ventas_reportes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="tab_reportes_descargar_ventas_excel",
            )

            st.markdown("---")
            st.subheader("📅 Reporte Diario")

            columnas_reporte_diario = [
                "Fecha_Pago_Comprobante",
                "Cliente",
                "Folio_Factura",
                "Monto_Comprobante",
                "Forma_Pago_Comprobante",
                "Comentario",
            ]
            st.caption("Solo se muestran pedidos marcados con forma de pago 'Depósito en Efectivo'.")
            for col in columnas_reporte_diario + ["Forma_Pago_Comprobante"]:
                if col not in df_ventas_base.columns:
                    df_ventas_base[col] = ""

            forma_pago_normalizada = (
                df_ventas_base["Forma_Pago_Comprobante"]
                .astype(str)
                .apply(normalizar)
                .str.strip()
                .str.lower()
            )
            df_reporte_diario = df_ventas_base[
                forma_pago_normalizada.eq(normalizar("Depósito en Efectivo").lower())
            ].copy()

            df_reporte_diario["Fecha_Pago_Comprobante_dt"] = pd.to_datetime(
                df_reporte_diario["Fecha_Pago_Comprobante"], errors="coerce"
            )

            fechas_pago_validas = df_reporte_diario["Fecha_Pago_Comprobante_dt"].dropna()
            if fechas_pago_validas.empty:
                st.info("No hay registros de 'Depósito en Efectivo' con fecha de pago válida.")
            else:
                fecha_diaria_default = fechas_pago_validas.max().date()
                fecha_reporte_diario = st.date_input(
                    "📆 Filtrar por día",
                    value=fecha_diaria_default,
                    key="tab_reportes_filtro_fecha_diaria",
                )

                df_reporte_diario = df_reporte_diario[
                    df_reporte_diario["Fecha_Pago_Comprobante_dt"].dt.date == fecha_reporte_diario
                ].copy()

                st.dataframe(
                    df_reporte_diario[columnas_reporte_diario].reset_index(drop=True),
                    use_container_width=True,
                    hide_index=True,
                )
                st.caption(
                    f"Total de registros (Depósito en Efectivo) para {fecha_reporte_diario}: {len(df_reporte_diario)}"
                )

                reporte_diario_excel_buffer = BytesIO()
                with pd.ExcelWriter(reporte_diario_excel_buffer, engine="openpyxl") as writer:
                    df_reporte_diario[columnas_reporte_diario].to_excel(
                        writer, index=False, sheet_name="Reporte_Diario"
                    )
                st.download_button(
                    label="📥 Descargar reporte diario (Excel)",
                    data=reporte_diario_excel_buffer.getvalue(),
                    file_name=f"reporte_diario_{fecha_reporte_diario}_{datetime.now().strftime('%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="tab_reportes_descargar_reporte_diario_excel",
                )


# --- TAB 2: MODIFY EXISTING ORDER ---
reset_inputs_tab2_flag = st.session_state.pop("reset_inputs_tab2", False)
if reset_inputs_tab2_flag:
    # Limpiar entradas controladas por widgets antes de instanciarlos
    for key in [
        "new_modificacion_surtido_input",
        "uploaded_files_surtido",
        "uploaded_comprobantes_extra",
        "tipo_modificacion_mod",
        "refact_tipo_mod_outside",
        "subtipo_datos_outside",
        "subtipo_material_outside",
        "folio_refact_outside",
    ]:
        st.session_state.pop(key, None)

with tab2:
    tab2_is_active = default_tab == TAB_INDEX_TAB2
    if tab2_is_active:
        st.session_state["current_tab_index"] = TAB_INDEX_TAB2
    st.header("✏️ Modificar Pedido Existente")
    st.caption("ℹ️ En esta sección solo saldrán los pedidos que no han viajado.")
    if st.button("🔄 Actualizar pedidos"):
        cargar_pedidos_combinados.clear()

    message_placeholder_tab2 = st.empty()
    loading_message_tab2 = st.session_state.pop(TAB2_LOADING_MESSAGE_KEY, None)
    if loading_message_tab2:
        message_placeholder_tab2.info(loading_message_tab2)


    # 🔄 Cargar pedidos combinados siempre (Tab 1 y Tab 2 activos de forma permanente)
    try:
        df_pedidos = cargar_pedidos_combinados()
    except Exception as e:
        message_placeholder_tab2.error(f"❌ Error al cargar pedidos para modificación: {e}")
        st.stop()

    # ----------------- Estado local -----------------
    selected_order_id = None
    selected_row_data = None
    selected_source = SHEET_PEDIDOS_OPERATIVOS  # por defecto
    current_modificacion_surtido_value = ""
    current_estado_pago_value = "🔴 No Pagado"
    current_adjuntos_list = []
    current_adjuntos_surtido_list = []

    if df_pedidos.empty:
        message_placeholder_tab2.warning("No hay pedidos registrados para modificar.")
    else:
        # 🔧 Normaliza 'Vendedor_Registro' usando 'Vendedor' como respaldo
        if 'Vendedor_Registro' not in df_pedidos.columns:
            df_pedidos['Vendedor_Registro'] = ""
        if 'Vendedor' in df_pedidos.columns:
            df_pedidos['Vendedor_Registro'] = df_pedidos['Vendedor_Registro'].astype(str).str.strip()
            fallback_v = df_pedidos['Vendedor'].astype(str).str.strip()
            df_pedidos.loc[df_pedidos['Vendedor_Registro'] == "", 'Vendedor_Registro'] = fallback_v

        # 🔽 Filtro combinado por envío (usa Turno si es Local)
        df_pedidos['Filtro_Envio_Combinado'] = df_pedidos.apply(
            lambda row: row['Turno'] if (str(row.get('Tipo_Envio',"")) == "📍 Pedido Local" and pd.notna(row.get('Turno')) and str(row.get('Turno')).strip()) else row.get('Tipo_Envio', ''),
            axis=1
        )

        # ----------------- Controles de filtro -----------------
        col1, col2 = st.columns(2)

        with col1:
            if 'Vendedor_Registro' in df_pedidos.columns:
                unique_vendedores_mod = build_vendor_filter_options(
                    df_pedidos["Vendedor_Registro"].dropna().astype(str).tolist(),
                )
                selected_vendedor_mod = st.selectbox(
                    "Filtrar por Vendedor:",
                    options=unique_vendedores_mod,
                    index=ensure_selectbox_vendor_default("vendedor_filter_mod", unique_vendedores_mod),
                    key="vendedor_filter_mod"
                )
            else:
                selected_vendedor_mod = "Todos"

        with col2:
            (
                fecha_inicio_mod,
                fecha_fin_mod,
                _rango_activo_mod,
                rango_valido_mod,
            ) = render_date_filter_controls(
                "📅 Filtrar por Fecha de Registro:",
                "tab2_modificar_filtro",
                recent_days_option=4,
                recent_days_label="Mostrar últimos 4 días",
            )

        # ----------------- Aplicar filtros -----------------
        filtered_orders = df_pedidos.copy()

        # 🔒 Asegura que se preserve el número real de fila en la hoja
        if "Sheet_Row_Number" not in filtered_orders.columns:
            if "Sheet_Row_Number" in df_pedidos.columns:
                filtered_orders["Sheet_Row_Number"] = df_pedidos.loc[
                    filtered_orders.index, "Sheet_Row_Number"
                ].values
            else:
                filtered_orders["Sheet_Row_Number"] = ""

        if selected_vendedor_mod != "Todos":
            filtered_orders = filtered_orders[filtered_orders['Vendedor_Registro'] == selected_vendedor_mod]

        # Filtrar por fecha usando 'Hora_Registro' si existe
        if 'Hora_Registro' in filtered_orders.columns:
            filtered_orders = filtered_orders.copy()
            filtered_orders['Hora_Registro'] = pd.to_datetime(filtered_orders['Hora_Registro'], errors='coerce')
            if rango_valido_mod:
                start_dt = datetime.combine(fecha_inicio_mod, datetime.min.time())
                end_dt = datetime.combine(fecha_fin_mod, datetime.max.time())
                filtered_orders = filtered_orders[
                    filtered_orders['Hora_Registro'].between(start_dt, end_dt)
                ]
            else:
                filtered_orders = filtered_orders.iloc[0:0]

        if filtered_orders.empty:
            if not rango_valido_mod:
                st.info("Ajusta el rango de fechas para continuar.")
            else:
                st.warning("No hay pedidos que coincidan con los filtros seleccionados.")
        else:
            # 🔧 Limpieza para evitar 'nan' en el select
            for col in ['Folio_Factura', 'ID_Pedido', 'Cliente', 'Estado', 'Tipo_Envio']:
                if col in filtered_orders.columns:
                    filtered_orders[col] = (
                        filtered_orders[col]
                        .astype(str)
                        .replace(['nan', 'None'], '')
                        .fillna('')
                        .str.strip()
                    )

            # 🧹 Orden por Fecha_Entrega (más reciente primero) si existe
            if 'Fecha_Entrega' in filtered_orders.columns:
                filtered_orders['Fecha_Entrega'] = pd.to_datetime(filtered_orders['Fecha_Entrega'], errors='coerce')
                filtered_orders = filtered_orders.sort_values(by='Fecha_Entrega', ascending=False).reset_index(drop=True)

            # 🏷️ Etiqueta de display (marca [CE] si es de casos_especiales)
            def _s(x):
                return "" if pd.isna(x) else str(x)

            filtered_orders['display_label'] = filtered_orders.apply(
                lambda row: (
                    f"📄 {(_s(row['Folio_Factura']) or 'Sin Folio')}"
                    f" - {_s(row['Cliente'])}"
                    f" - {_s(row['Estado'])}"
                    f" - {_s(row['Tipo_Envio'])}"
                    f" {'[CE]' if row.get('Fuente','')=='casos_especiales' else ''}"
                ),
                axis=1
            )

            base_option_values = filtered_orders.apply(
                lambda row: (
                    f"{row.get('Fuente', SHEET_PEDIDOS_OPERATIVOS)}|"
                    f"{_s(row.get('ID_Pedido', '')) or 'sin_id'}|"
                    f"{parse_sheet_row_number(row.get('Sheet_Row_Number')) or 'sin_fila'}"
                ),
                axis=1
            )

            filtered_orders['option_value'] = base_option_values

            duplicate_mask = filtered_orders['option_value'].duplicated(keep=False)
            if duplicate_mask.any():
                filtered_orders.loc[duplicate_mask, 'option_value'] = filtered_orders.loc[duplicate_mask].apply(
                    lambda row: f"{base_option_values[row.name]}|{row.name}",
                    axis=1
                )

            option_label_map = dict(
                zip(filtered_orders['option_value'], filtered_orders['display_label'])
            )

            def _format_option(option_key: str) -> str:
                return option_label_map.get(option_key, option_key)

            # ----------------- Selector de pedido -----------------
            selected_option_key = st.selectbox(
                "📝 Seleccionar Pedido para Modificar",
                list(option_label_map.keys()),
                format_func=_format_option,
                key="select_order_to_modify"
            )

            if selected_option_key:
                option_changed = st.session_state.get("tab2_selected_option_key") != selected_option_key
                if option_changed:
                    st.session_state["tab2_selected_option_key"] = selected_option_key
                    st.session_state["tab2_confirm_order"] = False
                    st.session_state.pop("new_modificacion_surtido_input", None)
                    for tab2_local_key in (
                        "tab2_local_route_enabled",
                        "tab2_local_route_client_search",
                        "tab2_local_route_selected_history_label",
                        "tab2_local_route_selected_history_row",
                        "tab2_local_route_pending_registro_cliente",
                    ):
                        st.session_state.pop(tab2_local_key, None)

                matched = filtered_orders[
                    filtered_orders['option_value'] == selected_option_key
                ].iloc[0]
                selected_order_id = matched['ID_Pedido']
                selected_source = matched.get('Fuente', SHEET_PEDIDOS_OPERATIVOS)  # 'data_pedidos' o 'casos_especiales'

                selected_row_data = matched.copy()
                if 'Seguimiento' not in selected_row_data.index:
                    selected_row_data['Seguimiento'] = ''

                # Guarda la fila real de Google Sheets para evitar desalineaciones
                selected_row_number = parse_sheet_row_number(
                    selected_row_data.get("Sheet_Row_Number")
                )
                st.session_state["tab2_row_to_edit"] = selected_row_number
                st.session_state["tab2_row_source"] = selected_source

                # Si viene de casos_especiales y es Devolución/Garantía -> render especial
                tipo_det = __s(selected_row_data.get('Tipo_Envio', ''))
                if selected_source == "casos_especiales" and tipo_det in ("🔁 Devolución", "🛠 Garantía"):
                    render_caso_especial(selected_row_data)
                else:
                    # ----------------- Detalles básicos (para data_pedidos u otros) -----------------
                    st.subheader(
                        f"Detalles del Pedido: Folio {selected_row_data.get('Folio_Factura', 'N/A')}"
                    )

                    fuente_display = (
                        "📄 data_pedidos"
                        if selected_source == SHEET_PEDIDOS_OPERATIVOS
                        else "🔁 casos_especiales"
                    )
                    vendedor_preferido = selected_row_data.get("Vendedor", "")
                    if not vendedor_preferido or str(vendedor_preferido).strip().lower() in {"nan", "none"}:
                        vendedor_preferido = selected_row_data.get(
                            "Vendedor_Registro", "No especificado"
                        )
                    vendedor_display = normalize_case_text(
                        vendedor_preferido, "No especificado"
                    )
                    tipo_envio_val = selected_row_data.get("Tipo_Envio", "N/A")
                    es_local = tipo_envio_val == "📍 Pedido Local"
                    if es_local:
                        turno_local = selected_row_data.get("Turno", "N/A")
                        estado_entrega_local = format_estado_entrega(
                            selected_row_data.get("Estado_Entrega")
                        )
                    else:
                        turno_local = None
                        estado_entrega_local = None

                    detalles_col_izq, detalles_col_der = st.columns(2)

                    with detalles_col_izq:
                        st.markdown(f"**Fuente:** {fuente_display}")
                        st.markdown(f"**Vendedor:** {vendedor_display}")
                        st.markdown(format_id_vendedor_with_mod(selected_row_data))
                        st.markdown(f"**Cliente:** {selected_row_data.get('Cliente', 'N/A')}")
                        st.markdown(
                            f"**Folio de Factura:** {selected_row_data.get('Folio_Factura', 'N/A')}"
                        )

                    with detalles_col_der:
                        st.markdown(f"**Estado Actual:** {selected_row_data.get('Estado', 'N/A')}")
                        st.markdown(
                            f"**Estado de Pago:** {selected_row_data.get('Estado_Pago', '🔴 No Pagado')}"
                        )
                        st.markdown(f"**Tipo de Envío:** {tipo_envio_val}")
                        if es_local:
                            st.markdown(f"**Turno Local:** {turno_local}")
                            st.markdown(f"**Estado_Entrega:** {estado_entrega_local}")
                        st.markdown(
                            f"**Fecha de Entrega:** {selected_row_data.get('Fecha_Entrega', 'N/A')}"
                        )

                    st.markdown("**Comentario Original:**")
                    st.write(selected_row_data.get("Comentario", "N/A"))

                    direccion_envio = selected_row_data.get("Direccion_Guia_Retorno", "")
                    if not str(direccion_envio).strip() or str(direccion_envio).strip().lower() in {"nan", "none"}:
                        direccion_envio = selected_row_data.get("Direccion_Envio", "")
                    direccion_envio = (
                        str(direccion_envio).strip()
                        if pd.notna(direccion_envio)
                        else ""
                    )
                    st.markdown(
                        f"**📬 Dirección para Envió:** {direccion_envio or 'N/A'}"
                    )

                    current_adjuntos_str_basic = selected_row_data.get('Adjuntos', '')
                    current_adjuntos_list_basic = [f.strip() for f in str(current_adjuntos_str_basic).split(',') if f.strip()]
                    current_adjuntos_surtido_str_basic = selected_row_data.get('Adjuntos_Surtido', '')
                    current_adjuntos_surtido_list_basic = [f.strip() for f in str(current_adjuntos_surtido_str_basic).split(',') if f.strip()]

                    with st.expander("📎 Ver adjuntos del pedido", expanded=False):
                        if current_adjuntos_list_basic:
                            st.write("**Adjuntos Originales:**")
                            for adj in current_adjuntos_list_basic:
                                render_attachment_link(adj)
                        else:
                            st.write("**Adjuntos Originales:** Ninguno")

                        if current_adjuntos_surtido_list_basic:
                            st.write("**Adjuntos de Modificación/Surtido:**")
                            for adj_surtido in current_adjuntos_surtido_list_basic:
                                render_attachment_link(adj_surtido)
                        else:
                            st.write("**Adjuntos de Modificación/Surtido:** Ninguno")

                # ----------------- Valores actuales (para formulario) -----------------
                current_modificacion_surtido_value = selected_row_data.get('Modificacion_Surtido', '')
                if option_changed:
                    st.session_state["new_modificacion_surtido_input"] = str(current_modificacion_surtido_value or "")
                current_estado_pago_value = selected_row_data.get('Estado_Pago', '🔴 No Pagado')
                current_adjuntos_str = selected_row_data.get('Adjuntos', '')
                current_adjuntos_list = [f.strip() for f in str(current_adjuntos_str).split(',') if f.strip()]
                current_has_comprobante = any(
                    "comprobante" in str(adj).lower() for adj in current_adjuntos_list
                )
                current_paid_from_sheet = str(current_estado_pago_value or "").strip() == "✅ Pagado"
                current_has_comprobante_effective = current_has_comprobante or current_paid_from_sheet
                current_adjuntos_surtido_str = selected_row_data.get('Adjuntos_Surtido', '')
                current_adjuntos_surtido_list = [f.strip() for f in str(current_adjuntos_surtido_str).split(',') if f.strip()]
                tab2_route_prefix = "tab2_local_route"
                tab2_local_order = selected_row_data.get("Tipo_Envio", "") == "📍 Pedido Local"
                tab2_turno_selector_key = "tab2_local_shift_selector"
                tab2_fecha_entrega_key = "tab2_local_fecha_entrega_requerida"
                id_vendedor_tab2_shift = extract_id_vendedor(
                    selected_row_data,
                    str(st.session_state.get("id_vendedor", "") or ""),
                )
                local_shift_options_tab2 = get_local_shift_options(id_vendedor_tab2_shift)
                tab2_estado_pago_options = ["🔴 No Pagado", "✅ Pagado", "💳 CREDITO"]
                if option_changed and tab2_local_order:
                    st.session_state["tab2_local_route_enabled"] = False
                    if current_estado_pago_value in tab2_estado_pago_options:
                        estado_pago_default_tab2 = current_estado_pago_value
                    elif current_has_comprobante_effective:
                        estado_pago_default_tab2 = "✅ Pagado"
                    else:
                        estado_pago_default_tab2 = "🔴 No Pagado"
                    st.session_state["tab2_local_estado_pago"] = estado_pago_default_tab2
                    st.session_state["tab2_local_route_client_search"] = str(
                        selected_row_data.get("Cliente", "") or ""
                    ).strip()
                    st.session_state[f"{tab2_route_prefix}_recibe"] = str(selected_row_data.get("Recibe", "") or "").strip()
                    st.session_state[f"{tab2_route_prefix}_calle_no"] = str(selected_row_data.get("CalleyNumero", "") or "").strip()
                    st.session_state[f"{tab2_route_prefix}_tipo_inmueble"] = str(selected_row_data.get("Tipo_Inmueble", "") or "").strip() or "Consultorio"
                    st.session_state[f"{tab2_route_prefix}_acceso_privada"] = str(selected_row_data.get("Acceso_Privada", "") or "").strip() or "No aplica"
                    st.session_state[f"{tab2_route_prefix}_municipio"] = str(selected_row_data.get("Municipio", "") or "").strip() or "MONTERREY"
                    st.session_state[f"{tab2_route_prefix}_telefonos"] = display_phone_from_sheets(selected_row_data.get("Tels", ""))
                    st.session_state[f"{tab2_route_prefix}_interior"] = str(selected_row_data.get("Interior", "") or "").strip()
                    st.session_state[f"{tab2_route_prefix}_colonia"] = str(selected_row_data.get("Col", "") or "").strip()
                    st.session_state[f"{tab2_route_prefix}_cp"] = str(selected_row_data.get("C_P.", "") or "").strip()
                    st.session_state[f"{tab2_route_prefix}_forma_pago"] = str(
                        selected_row_data.get("Forma_Pago_Comprobante", "") or ""
                    ).strip() or "TRANSFERENCIA"
                    try:
                        st.session_state[f"{tab2_route_prefix}_total_factura"] = float(selected_row_data.get("Total_Factura", 0) or 0)
                    except (TypeError, ValueError):
                        st.session_state[f"{tab2_route_prefix}_total_factura"] = 0.0
                    try:
                        st.session_state[f"{tab2_route_prefix}_adeudo_anterior"] = float(selected_row_data.get("Adeudo_Anterior", 0) or 0)
                    except (TypeError, ValueError):
                        st.session_state[f"{tab2_route_prefix}_adeudo_anterior"] = 0.0
                    st.session_state[f"{tab2_route_prefix}_referencias"] = str(selected_row_data.get("Referencias", "") or "").strip()
                    turno_local_default_tab2 = str(selected_row_data.get("Turno", "") or "").strip()
                    if turno_local_default_tab2 not in local_shift_options_tab2:
                        turno_local_default_tab2 = local_shift_options_tab2[0]
                    st.session_state[tab2_turno_selector_key] = turno_local_default_tab2
                    fecha_entrega_base_tab2 = pd.to_datetime(
                        selected_row_data.get("Fecha_Entrega"),
                        errors="coerce",
                    )
                    st.session_state[tab2_fecha_entrega_key] = (
                        fecha_entrega_base_tab2.date()
                        if pd.notna(fecha_entrega_base_tab2)
                        else datetime.now().date()
                    )
                    st.session_state["tab2_local_estado_pago_loaded_value"] = estado_pago_default_tab2
                    st.session_state["tab2_local_estado_pago_loaded_order"] = selected_option_key
                elif tab2_local_order and st.session_state.get(tab2_turno_selector_key) not in local_shift_options_tab2:
                    st.session_state[tab2_turno_selector_key] = local_shift_options_tab2[0]
                if tab2_local_order:
                    estado_pago_from_sheet = str(selected_row_data.get("Estado_Pago", "") or "").strip()
                    if estado_pago_from_sheet not in tab2_estado_pago_options:
                        estado_pago_from_sheet = "✅ Pagado" if current_has_comprobante_effective else "🔴 No Pagado"
                    loaded_order_key = st.session_state.get("tab2_local_estado_pago_loaded_order")
                    loaded_estado_pago = st.session_state.get("tab2_local_estado_pago_loaded_value")
                    if loaded_order_key != selected_option_key or loaded_estado_pago != estado_pago_from_sheet:
                        st.session_state["tab2_local_estado_pago"] = estado_pago_from_sheet
                        st.session_state["tab2_local_estado_pago_loaded_value"] = estado_pago_from_sheet
                        st.session_state["tab2_local_estado_pago_loaded_order"] = selected_option_key

                st.markdown("---")
                st.subheader("Modificar Campos y Adjuntos (Surtido)")
                st.markdown("### 🛠 Tipo de modificación")

                # ----------------- Tipo de modificación -----------------
                tipo_modificacion_seleccionada = st.selectbox(
                    "📌 ¿Qué tipo de modificación estás registrando?",
                    ["Otro", "Nueva Ruta", "Refacturación"],
                    index=0,
                    key="tipo_modificacion_mod"
                )

                refact_tipo = ""
                refact_subtipo_val = ""
                refact_folio_nuevo = ""

                if tipo_modificacion_seleccionada == "Refacturación":
                    st.markdown("### 🧾 Detalles de Refacturación")

                    refact_tipo = st.selectbox(
                        "🔍 Razón Principal",
                        ["Datos Fiscales", "Material"],
                        key="refact_tipo_mod_outside"
                    )

                    if refact_tipo == "Datos Fiscales":
                        refact_subtipo_val = st.selectbox(
                            "📌 Subtipo",
                            ["Cambio de RFC", "Cambio de Régimen Fiscal", "Error en Forma de Pago", "Error de uso de Cfdi", "Otro"],
                            key="subtipo_datos_outside",
                            placeholder="Selecciona una opción..."
                        )
                    else:
                        refact_subtipo_val = st.selectbox(
                            "📌 Subtipo",
                            ["Agrego Material", "Quito Material", "Clave de Producto Errónea", "Otro"],
                            key="subtipo_material_outside",
                            placeholder="Selecciona una opción..."
                        )

                    refact_folio_nuevo = st.text_input("📄 Folio de la Nueva Factura", key="folio_refact_outside")

                apply_local_route_update = False
                tab2_route_payload = None
                if tab2_local_order:
                    st.markdown("### 🗺️ Hoja de Ruta Local")
                    apply_local_route_update = st.checkbox(
                        "Habilitar actualización de Hoja de Ruta Local",
                        key="tab2_local_route_enabled",
                        help="Usa la misma lógica de búsqueda automática para actualizar la hoja de ruta del pedido local.",
                    )
                    if apply_local_route_update:
                        st.caption(
                            "🤝 Cliente con búsqueda automática\n\nEscribe el nombre del cliente y dale ENTER. La app buscará coincidencias en el historial local."
                        )
                        pending_tab2_cliente = st.session_state.pop(
                            "tab2_local_route_pending_registro_cliente",
                            None,
                        )
                        if pending_tab2_cliente is not None:
                            st.session_state["tab2_local_route_client_search"] = pending_tab2_cliente

                        tab2_cliente = st.text_input(
                            "🤝 Cliente",
                            key="tab2_local_route_client_search",
                            placeholder="Escribe o pega el nombre del cliente",
                        )
                        clientes_locales_df = load_clientes_locales_dataset()
                        tab2_matches = find_clientes_locales_matches(tab2_cliente, clientes_locales_df)
                        tab2_options: dict[str, dict] = {}
                        normalized_tab2_cliente = normalize_client_history_text(tab2_cliente)
                        exact_tab2_label = None
                        for match in tab2_matches:
                            display_label = f"{str(match.get('Cliente', '')).strip()} | C.P. {str(match.get('C_P.', '')).strip() or 'N/A'}"
                            suffix = 2
                            base_label = display_label
                            while display_label in tab2_options:
                                display_label = f"{base_label} ({suffix})"
                                suffix += 1
                            tab2_options[display_label] = match
                            if (
                                normalized_tab2_cliente
                                and normalize_client_history_text(match.get("Cliente", "")) == normalized_tab2_cliente
                            ):
                                exact_tab2_label = display_label

                        selected_tab2_row = st.session_state.get("tab2_local_route_selected_history_row")
                        previous_tab2_label = st.session_state.get("tab2_local_route_selected_history_label")
                        if exact_tab2_label and exact_tab2_label in tab2_options:
                            selected_tab2_label = exact_tab2_label
                            selected_tab2_record = tab2_options[exact_tab2_label]
                            st.caption(f"✅ Cliente encontrado en historial: {selected_tab2_label}")
                            selected_tab2_row_number = parse_sheet_row_number(selected_tab2_record.get("Sheet_Row_Number"))
                            if selected_tab2_row != selected_tab2_row_number:
                                st.session_state["tab2_local_route_selected_history_label"] = selected_tab2_label
                                st.session_state["tab2_local_route_selected_history_row"] = selected_tab2_row_number
                                apply_cliente_local_record_to_session(
                                    selected_tab2_record,
                                    route_prefix=tab2_route_prefix,
                                    pending_cliente_key="tab2_local_route_pending_registro_cliente",
                                )
                                st.rerun()
                        elif len(tab2_options) == 1:
                            selected_tab2_label, selected_tab2_record = next(iter(tab2_options.items()))
                            st.caption(f"✅ Coincidencia encontrada: {selected_tab2_label}")
                            selected_tab2_row_number = parse_sheet_row_number(selected_tab2_record.get("Sheet_Row_Number"))
                            if selected_tab2_row != selected_tab2_row_number:
                                st.session_state["tab2_local_route_selected_history_label"] = selected_tab2_label
                                st.session_state["tab2_local_route_selected_history_row"] = selected_tab2_row_number
                                apply_cliente_local_record_to_session(
                                    selected_tab2_record,
                                    route_prefix=tab2_route_prefix,
                                    pending_cliente_key="tab2_local_route_pending_registro_cliente",
                                )
                                st.rerun()
                        elif tab2_options:
                            option_labels = list(tab2_options.keys())
                            selected_history_index = option_labels.index(previous_tab2_label) if previous_tab2_label in tab2_options else None
                            selected_tab2_label = st.radio(
                                "Coincidencias encontradas",
                                options=option_labels,
                                index=selected_history_index,
                                key="tab2_local_route_selected_history_label",
                            )
                            selected_tab2_record = tab2_options.get(selected_tab2_label)
                            if selected_tab2_record:
                                selected_tab2_row_number = parse_sheet_row_number(selected_tab2_record.get("Sheet_Row_Number"))
                                if selected_tab2_row != selected_tab2_row_number:
                                    st.session_state["tab2_local_route_selected_history_row"] = selected_tab2_row_number
                                    apply_cliente_local_record_to_session(
                                        selected_tab2_record,
                                        route_prefix=tab2_route_prefix,
                                        pending_cliente_key="tab2_local_route_pending_registro_cliente",
                                    )
                                    st.rerun()
                        elif tab2_cliente.strip():
                            st.caption("🆕 Cliente nuevo sin historial local.")
                            st.session_state["tab2_local_route_selected_history_label"] = None
                            st.session_state["tab2_local_route_selected_history_row"] = None

                        col_local_1, col_local_2 = st.columns(2)
                        with col_local_1:
                            tab2_local_route_recibe = st.text_input("🙋 Recibe", key=f"{tab2_route_prefix}_recibe")
                            tab2_local_route_calle_no = st.text_input("📍 CALLE Y NO.", key=f"{tab2_route_prefix}_calle_no")
                            tab2_local_route_tipo_inmueble = st.selectbox(
                                "🏢 TIPO INMUEBLE",
                                ["Consultorio", "Clínica", "Hospital", "Casa", "Departamento", "Oficina", "Local comercial", "Otro"],
                                key=f"{tab2_route_prefix}_tipo_inmueble",
                            )
                            tab2_local_route_acceso_privada = st.selectbox(
                                "🚧 ACCESO PRIVADA",
                                ["No aplica", "Aplica", "Acceso controlado", "Requiere autorización previa"],
                                key=f"{tab2_route_prefix}_acceso_privada",
                            )
                            tab2_local_route_municipio = st.text_input("🗺️ MUNICIPIO", key=f"{tab2_route_prefix}_municipio")
                            tab2_local_route_telefonos = st.text_input("☎️ TELS", key=f"{tab2_route_prefix}_telefonos")
                        with col_local_2:
                            tab2_local_route_interior = st.text_input("🚪 INTERIOR", key=f"{tab2_route_prefix}_interior")
                            tab2_local_route_colonia = st.text_input("🏘️ COL.", key=f"{tab2_route_prefix}_colonia")
                            tab2_local_route_cp = st.text_input("📮 C.P.", key=f"{tab2_route_prefix}_cp")
                            tab2_forma_pago_key = f"{tab2_route_prefix}_forma_pago"
                            tab2_legacy_forma_pago_map = {
                                "TRANSFERENCIA": "Transferencia",
                                "EFECTIVO": "Depósito en Efectivo",
                                "TARJETA": "Tarjeta de Crédito",
                                "PENDIENTE": "Transferencia",
                            }
                            tab2_current_forma_pago = str(st.session_state.get(tab2_forma_pago_key, "") or "").strip()
                            if tab2_current_forma_pago in tab2_legacy_forma_pago_map:
                                st.session_state[tab2_forma_pago_key] = tab2_legacy_forma_pago_map[tab2_current_forma_pago]
                            tab2_local_route_forma_pago = st.selectbox(
                                "💳 FORMA DE PAGO",
                                [
                                    "Transferencia",
                                    "Depósito en Efectivo",
                                    "Tarjeta de Débito",
                                    "Tarjeta de Crédito",
                                    "Credito TD",
                                    "Cheque",
                                ],
                                key=tab2_forma_pago_key,
                            )
                            tab2_local_route_total_factura = st.number_input(
                                "💵 TOTAL FACTURA",
                                min_value=0.0,
                                format="%.2f",
                                key=f"{tab2_route_prefix}_total_factura",
                            )
                            tab2_local_route_adeudo_anterior = st.number_input(
                                "💸 ADEUDO ANT.",
                                min_value=0.0,
                                format="%.2f",
                                key=f"{tab2_route_prefix}_adeudo_anterior",
                            )
                        tab2_local_route_referencias = st.text_area(
                            "📝 REFERENCIAS Y/O COMENTARIOS (solo hoja de ruta)",
                            key=f"{tab2_route_prefix}_referencias",
                        )

                        tab2_fecha_entrega_requerida = st.date_input(
                            "🗓 Fecha de Entrega Requerida",
                            value=st.session_state.get(tab2_fecha_entrega_key, datetime.now().date()),
                            key=tab2_fecha_entrega_key,
                        )
                        tab2_turno_local = st.selectbox(
                            "⏰ Turno / Local",
                            local_shift_options_tab2,
                            key=tab2_turno_selector_key,
                            help="Usa las mismas opciones de turno del Tab 1 para pedidos locales.",
                        )
                        if tab2_turno_local not in ["☀️ Local Mañana", "🌙 Local Tarde"]:
                            st.warning("⚠️ La hoja de ruta asigna horario automático solo para ☀️ Local Mañana y 🌙 Local Tarde. Para otros turnos se usará el texto del turno seleccionado.")

                        st.subheader("💰 Estado de Pago")
                        if st.session_state.get("tab2_local_estado_pago") not in tab2_estado_pago_options:
                            st.session_state["tab2_local_estado_pago"] = "✅ Pagado" if current_has_comprobante_effective else "🔴 No Pagado"

                        tab2_estado_pago = st.selectbox(
                            "Estado de Pago",
                            tab2_estado_pago_options,
                            index=tab2_estado_pago_options.index(st.session_state.get("tab2_local_estado_pago", "🔴 No Pagado")),
                            key="tab2_local_estado_pago",
                        )
                        if tab2_estado_pago == "✅ Pagado":
                            if current_has_comprobante_effective:
                                st.info("ℹ️ El pedido ya tiene comprobante previo; puedes guardar sin subir uno adicional.")
                            else:
                                st.warning("⚠️ Estado en PAGADO: debes adjuntar al menos un comprobante para guardar la modificación.")
                        else:
                            st.caption("ℹ️ Los Comprobantes son obligatorios cuando el estado sea '✅ Pagado' y no haya comprobantes registrados.")

                        tab2_route_payload = build_local_route_payload(
                            fecha_entrega=tab2_fecha_entrega_requerida,
                            registro_cliente=tab2_cliente,
                            subtipo_local=str(tab2_turno_local or local_shift_options_tab2[0]),
                            recibe=tab2_local_route_recibe,
                            referencias_hoja_ruta=tab2_local_route_referencias,
                            calle_no=tab2_local_route_calle_no,
                            tipo_inmueble=tab2_local_route_tipo_inmueble,
                            interior=tab2_local_route_interior,
                            acceso_privada=tab2_local_route_acceso_privada,
                            colonia=tab2_local_route_colonia,
                            municipio=tab2_local_route_municipio,
                            cp=tab2_local_route_cp,
                            telefonos=tab2_local_route_telefonos,
                            estado_pago=tab2_estado_pago,
                            forma_pago=tab2_local_route_forma_pago,
                            vendedor=get_session_vendedor_name() or str(selected_row_data.get("Vendedor", "") or ""),
                            total_factura=tab2_local_route_total_factura,
                            adeudo_anterior=tab2_local_route_adeudo_anterior,
                            folio=str(selected_row_data.get("Folio_Factura", "") or ""),
                        )
                        st.caption(
                            f"📅 Día de entrega: {tab2_route_payload.get('dia_entrega', 'N/A')} | "
                            f"🕒 Horario asignado: {tab2_route_payload.get('hora_entrega', 'N/A')}"
                        )

                # ----------------- Formulario de modificación -----------------
                with st.form(key="modify_pedido_form_inner", clear_on_submit=False):
                    if reset_inputs_tab2_flag:
                        st.session_state["new_modificacion_surtido_input"] = ""
                    elif "new_modificacion_surtido_input" not in st.session_state:
                        st.session_state["new_modificacion_surtido_input"] = str(current_modificacion_surtido_value or "")

                    new_modificacion_surtido_input = st.text_area(
                        "✍️ Notas de Modificación/Surtido",
                        height=100,
                        key="new_modificacion_surtido_input"
                    )

                    uploaded_files_surtido = st.file_uploader(
                        "📎 Subir Archivos para Modificación/Surtido",
                        type=["pdf", "jpg", "jpeg", "png", "xlsx", "docx"],
                        accept_multiple_files=True,
                        key="uploaded_files_surtido"
                    )

                    uploaded_comprobantes_extra = st.file_uploader(
                        "💲 Comprobante(s) de Pago",
                        type=["pdf", "jpg", "jpeg", "png"],
                        accept_multiple_files=True,
                        key="uploaded_comprobantes_extra"
                    )
                    render_uploaded_files_preview("Comprobantes de pago seleccionados", uploaded_comprobantes_extra)

                    folio_confirm = selected_row_data.get("Folio_Factura", "N/A")
                    cliente_confirm = selected_row_data.get("Cliente", "N/A")
                    confirm_order = st.checkbox(
                        f"✅ Confirmo que el pedido/cliente mostrado es el correcto (Folio: {folio_confirm} | Cliente: {cliente_confirm})",
                        key="tab2_confirm_order"
                    )

                    # Botón para procesar la modificación del pedido
                    modify_button = st.form_submit_button("✅ Procesar Modificación")
                    feedback_slot = st.empty()

                    if modify_button:
                        feedback_slot.empty()
                        can_process_modification = True
                        if not confirm_order:
                            feedback_slot.error(
                                "⚠️ Confirma que el pedido y cliente son correctos antes de procesar la modificación."
                            )
                            can_process_modification = False
                        if not str(new_modificacion_surtido_input).strip():
                            feedback_slot.empty()
                            feedback_slot.error(
                                "⚠️ El campo 'Notas de Modificación/Surtido' es obligatorio para procesar la modificación."
                            )
                            can_process_modification = False

                        if can_process_modification and apply_local_route_update and tab2_route_payload:
                            missing_route_fields = get_local_route_missing_fields(tab2_route_payload)
                            if missing_route_fields:
                                feedback_slot.empty()
                                feedback_slot.error(
                                    "⚠️ Completa los datos obligatorios de la hoja de ruta: "
                                    + ", ".join(missing_route_fields)
                                )
                                can_process_modification = False
                            if (
                                can_process_modification
                                and
                                st.session_state.get("tab2_local_estado_pago") == "✅ Pagado"
                                and not uploaded_comprobantes_extra
                                and not current_has_comprobante_effective
                            ):
                                feedback_slot.empty()
                                feedback_slot.error("⚠️ Debes adjuntar al menos un comprobante si el estado es '✅ Pagado'.")
                                can_process_modification = False

                        if can_process_modification:
                            try:
                                with st.spinner("⏳ Enviando modificación del pedido..."):
                                    # 1) Enrutar a la hoja correcta según la fuente
                                    client = build_gspread_client()
                                    sh = client.open_by_key(GOOGLE_SHEET_ID)
                                    hoja_objetivo = SHEET_PEDIDOS_OPERATIVOS if selected_source == SHEET_PEDIDOS_OPERATIVOS else "casos_especiales"
                                    worksheet = sh.worksheet(hoja_objetivo)

                                    headers = worksheet.row_values(1)
                                    if "ID_Pedido" not in headers:
                                        feedback_slot.empty()
                                        feedback_slot.error(f"❌ No se encontró la columna 'ID_Pedido' en la hoja {hoja_objetivo}.")
                                        st.stop()

                                    id_col_index = headers.index("ID_Pedido")
                                    selected_order_id_normalized = str(selected_order_id).strip()
                                    selected_folio_normalized = str(selected_row_data.get("Folio_Factura", "")).strip().upper()
                                    selected_cliente_normalized = str(selected_row_data.get("Cliente", "")).strip().upper()
                                    folio_col_index = headers.index("Folio_Factura") if "Folio_Factura" in headers else None
                                    cliente_col_index = headers.index("Cliente") if "Cliente" in headers else None
                                    all_values = worksheet.get_all_values()

                                sheet_row_number = parse_sheet_row_number(
                                    st.session_state.get("tab2_row_to_edit")
                                )
                                if sheet_row_number is None:
                                    sheet_row_number = parse_sheet_row_number(
                                        selected_row_data.get("Sheet_Row_Number")
                                    )

                                def _row_value(row_values, idx):
                                    if idx is None or len(row_values) <= idx:
                                        return ""
                                    return str(row_values[idx]).strip().upper()

                                gsheet_row_index = None
                                if (
                                    sheet_row_number is not None
                                    and sheet_row_number <= len(all_values)
                                ):
                                    candidate_row = all_values[sheet_row_number - 1]
                                    candidate_id = (
                                        str(candidate_row[id_col_index]).strip()
                                        if len(candidate_row) > id_col_index
                                        else ""
                                    )
                                    candidate_folio = _row_value(candidate_row, folio_col_index)
                                    candidate_cliente = _row_value(candidate_row, cliente_col_index)
                                    folio_matches = bool(selected_folio_normalized) and candidate_folio == selected_folio_normalized
                                    cliente_matches = bool(selected_cliente_normalized) and candidate_cliente == selected_cliente_normalized
                                    if candidate_id == selected_order_id_normalized and (folio_matches or cliente_matches):
                                        gsheet_row_index = int(sheet_row_number)

                                if gsheet_row_index is None:
                                    matching_rows = []
                                    for row_number, row_values in enumerate(all_values[1:], start=2):
                                        row_id = (
                                            str(row_values[id_col_index]).strip()
                                            if len(row_values) > id_col_index
                                            else ""
                                        )
                                        if row_id != selected_order_id_normalized:
                                            continue
                                        score = 0
                                        if selected_folio_normalized and _row_value(row_values, folio_col_index) == selected_folio_normalized:
                                            score += 2
                                        if selected_cliente_normalized and _row_value(row_values, cliente_col_index) == selected_cliente_normalized:
                                            score += 1
                                        matching_rows.append((score, row_number))

                                    if matching_rows:
                                        matching_rows.sort(reverse=True)
                                        gsheet_row_index = matching_rows[0][1]

                                if gsheet_row_index is None:
                                    feedback_slot.empty()
                                    feedback_slot.error(
                                        "❌ No se encontró la fila real del pedido en Google Sheets. Intenta refrescar y volver a seleccionar."
                                    )
                                    st.stop()

                                actual_values = (
                                    all_values[gsheet_row_index - 1]
                                    if gsheet_row_index <= len(all_values)
                                    else worksheet.row_values(gsheet_row_index)
                                )
                                actual_row_id = (
                                    str(actual_values[id_col_index]).strip()
                                    if len(actual_values) > id_col_index
                                    else ""
                                )
                                if actual_row_id != selected_order_id_normalized:
                                    feedback_slot.empty()
                                    feedback_slot.error(
                                        "❌ Validación de seguridad: la fila encontrada no corresponde al pedido seleccionado. No se aplicaron cambios."
                                    )
                                    st.stop()
                                if len(actual_values) < len(headers):
                                    actual_values = actual_values + [""] * (len(headers) - len(actual_values))
                                else:
                                    actual_values = actual_values[:len(headers)]

                                changes_made = False

                                cell_updates = []
                                actual_row = dict(zip(headers, actual_values))

                                def col_exists(col):
                                    return col in headers

                                def col_idx(col):
                                    return headers.index(col) + 1

                                # 2) Guardar Modificacion_Surtido (si cambió)
                                if col_exists("Modificacion_Surtido"):
                                    if str(new_modificacion_surtido_input) != str(current_modificacion_surtido_value):
                                        cell_updates.append({
                                            "range": rowcol_to_a1(
                                                gsheet_row_index,
                                                col_idx("Modificacion_Surtido"),
                                            ),
                                            "values": [[str(new_modificacion_surtido_input)]],
                                        })
                                        changes_made = True
                                if tab2_local_order and apply_local_route_update and col_exists("Turno"):
                                    nuevo_turno_tab2 = str(
                                        st.session_state.get(tab2_turno_selector_key, "")
                                        or local_shift_options_tab2[0]
                                    ).strip()
                                    turno_actual_tab2 = str(actual_row.get("Turno", "") or "").strip()
                                    if nuevo_turno_tab2 and nuevo_turno_tab2 != turno_actual_tab2:
                                        cell_updates.append({
                                            "range": rowcol_to_a1(
                                                gsheet_row_index,
                                                col_idx("Turno"),
                                            ),
                                            "values": [[nuevo_turno_tab2]],
                                        })
                                        changes_made = True

                                # 3) Subida de archivos de Surtido -> Adjuntos_Surtido
                                new_adjuntos_surtido_urls = []
                                if uploaded_files_surtido:
                                    for f in uploaded_files_surtido:
                                        ext = os.path.splitext(f.name)[1]
                                        s3_key = f"{selected_order_id}/surtido_{f.name.replace(' ', '_').replace(ext, '')}_{uuid.uuid4().hex[:4]}{ext}"
                                        success, url, error_msg = upload_file_to_s3(s3_client, S3_BUCKET_NAME, f, s3_key)
                                        if success:
                                            new_adjuntos_surtido_urls.append(url)
                                            changes_made = True
                                        else:
                                            feedback_slot.empty()
                                            feedback_slot.warning(
                                                f"⚠️ Falló la subida de {f.name}: {error_msg or 'Error desconocido'}"
                                            )

                                if new_adjuntos_surtido_urls and col_exists("Adjuntos_Surtido"):
                                    current_urls = [x.strip() for x in str(actual_row.get("Adjuntos_Surtido","")).split(",") if x.strip()]
                                    updated_str = ", ".join(current_urls + new_adjuntos_surtido_urls)
                                    cell_updates.append({
                                        "range": rowcol_to_a1(
                                            gsheet_row_index,
                                            col_idx("Adjuntos_Surtido"),
                                        ),
                                        "values": [[updated_str]],
                                    })

                                # 4) Comprobantes extra -> concatenar en 'Adjuntos'
                                comprobante_urls = []
                                if uploaded_comprobantes_extra:
                                    for archivo in uploaded_comprobantes_extra:
                                        ext = os.path.splitext(archivo.name)[1]
                                        s3_key = f"{selected_order_id}/comprobante_{selected_order_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}_{uuid.uuid4().hex[:4]}{ext}"
                                        success, url, error_msg = upload_file_to_s3(s3_client, S3_BUCKET_NAME, archivo, s3_key)
                                        if success:
                                            comprobante_urls.append(url)
                                            changes_made = True
                                        else:
                                            feedback_slot.empty()
                                            feedback_slot.warning(
                                                f"⚠️ Falló la subida del comprobante {archivo.name}: {error_msg or 'Error desconocido'}"
                                            )

                                    if comprobante_urls and col_exists("Adjuntos"):
                                        current_adjuntos = [x.strip() for x in str(actual_row.get("Adjuntos","")).split(",") if x.strip()]
                                        updated_adjuntos = ", ".join(current_adjuntos + comprobante_urls)
                                        cell_updates.append({
                                            "range": rowcol_to_a1(
                                                gsheet_row_index,
                                                col_idx("Adjuntos"),
                                            ),
                                            "values": [[updated_adjuntos]],
                                        })

                                if apply_local_route_update and tab2_route_payload:
                                    route_file_payload, route_filename = build_local_route_file_from_payload(
                                        Path("plantillas") / "FORMATO DE ENTREGA LOCAL limpia.xlsx",
                                        tab2_route_payload,
                                    )
                                    if not route_file_payload:
                                        feedback_slot.empty()
                                        feedback_slot.error("❌ No se pudo generar la hoja de ruta local.")
                                        st.stop()

                                    route_file_bytes = BytesIO(
                                        base64.b64decode(str(route_file_payload.get("content_b64", "") or ""))
                                    )
                                    route_file_bytes.name = str(route_filename or route_file_payload.get("name") or "hoja_ruta_local.xlsx")
                                    s3_key = f"{selected_order_id}/hoja_ruta_mod_{route_file_bytes.name.replace(' ', '_')}"
                                    success, route_url, route_error = upload_file_to_s3(
                                        s3_client,
                                        S3_BUCKET_NAME,
                                        route_file_bytes,
                                        s3_key,
                                    )
                                    if not success:
                                        feedback_slot.empty()
                                        feedback_slot.error(f"❌ No se pudo subir la hoja de ruta: {route_error or 'Error desconocido'}")
                                        st.stop()

                                    current_adjuntos = [x.strip() for x in str(actual_row.get("Adjuntos", "")).split(",") if x.strip()]
                                    updated_adjuntos = ", ".join(current_adjuntos + [route_url])
                                    if col_exists("Adjuntos"):
                                        cell_updates.append({
                                            "range": rowcol_to_a1(
                                                gsheet_row_index,
                                                col_idx("Adjuntos"),
                                            ),
                                            "values": [[updated_adjuntos]],
                                        })
                                        changes_made = True

                                    tab2_route_fields_map = {
                                        "Cliente": tab2_route_payload.get("cliente", ""),
                                        "Recibe": tab2_route_payload.get("recibe", ""),
                                        "CalleyNumero": tab2_route_payload.get("calle_no", ""),
                                        "Tipo_Inmueble": tab2_route_payload.get("tipo_inmueble", ""),
                                        "Acceso_Privada": tab2_route_payload.get("acceso_privada", ""),
                                        "Municipio": tab2_route_payload.get("municipio", ""),
                                        "Tels": normalize_phone_for_sheets(tab2_route_payload.get("telefonos", "")),
                                        "Interior": tab2_route_payload.get("interior", ""),
                                        "Col": tab2_route_payload.get("colonia", ""),
                                        "C_P.": tab2_route_payload.get("cp", ""),
                                        "Referencias": tab2_route_payload.get("referencias", ""),
                                        "Forma_Pago_Comprobante": tab2_route_payload.get("forma_pago", ""),
                                        "Estado_Pago": st.session_state.get("tab2_local_estado_pago", "🔴 No Pagado"),
                                        "Monto_Comprobante": f"{float(st.session_state.get(f'{tab2_route_prefix}_total_factura', 0) or 0) + float(st.session_state.get(f'{tab2_route_prefix}_adeudo_anterior', 0) or 0):.2f}",
                                    }
                                    for col_name, col_value in tab2_route_fields_map.items():
                                        if col_exists(col_name) and str(actual_row.get(col_name, "")) != str(col_value):
                                            cell_updates.append({
                                                "range": rowcol_to_a1(
                                                    gsheet_row_index,
                                                    col_idx(col_name),
                                                ),
                                                "values": [[col_value]],
                                            })
                                            changes_made = True

                                # 5) Refacturación (si las columnas existen en ESA hoja)
                                if tipo_modificacion_seleccionada == "Refacturación":
                                    campos_refact = {
                                        "Refacturacion_Tipo": refact_tipo,
                                        "Refacturacion_Subtipo": refact_subtipo_val,
                                        "Folio_Factura_Refacturada": refact_folio_nuevo
                                    }
                                    for campo, valor in campos_refact.items():
                                        if col_exists(campo):
                                            cell_updates.append({
                                                "range": rowcol_to_a1(
                                                    gsheet_row_index,
                                                    col_idx(campo),
                                                ),
                                                "values": [[valor]],
                                            })
                                    st.toast("🧾 Refacturación registrada.")
                                else:
                                    for campo in ["Refacturacion_Tipo","Refacturacion_Subtipo","Folio_Factura_Refacturada"]:
                                        if col_exists(campo):
                                            cell_updates.append({
                                                "range": rowcol_to_a1(
                                                    gsheet_row_index,
                                                    col_idx(campo),
                                                ),
                                                "values": [[""]],
                                            })

                                # 6) Ajustar estado del pedido para reflejar modificación
                                if col_exists("Estado"):
                                    estado_actual_raw = str(actual_row.get("Estado", "")).strip()
                                    estado_modificacion = "✏️ Modificación"

                                    if estado_actual_raw != estado_modificacion:
                                        cell_updates.append({
                                            "range": rowcol_to_a1(
                                                gsheet_row_index,
                                                col_idx("Estado"),
                                            ),
                                            "values": [[estado_modificacion]],
                                        })
                                        changes_made = True
                                        feedback_slot.empty()
                                        feedback_slot.info(
                                            f"📌 El estado del pedido se actualizó a '{estado_modificacion}'."
                                        )

                                if (
                                    selected_source == SHEET_PEDIDOS_OPERATIVOS
                                    and col_exists("Fecha_Completado")
                                ):
                                    cell_updates.append({
                                        "range": rowcol_to_a1(
                                            gsheet_row_index,
                                            col_idx("Fecha_Completado"),
                                        ),
                                        "values": [[""]],
                                    })

                                if col_exists("Completados_Limpiado"):
                                    cell_updates.append({
                                        "range": rowcol_to_a1(
                                            gsheet_row_index,
                                            col_idx("Completados_Limpiado"),
                                        ),
                                        "values": [[""]],
                                    })

                                # 7) Registrar quién modificó el pedido
                                id_vendedor_actual = str(st.session_state.get("id_vendedor", "")).strip()
                                if id_vendedor_actual and col_exists("id_vendedor_Mod"):
                                    existing_ids_raw = str(actual_row.get("id_vendedor_Mod", "")).strip()
                                    existing_ids = [
                                        entry.strip().upper()
                                        for entry in existing_ids_raw.split(",")
                                        if entry.strip()
                                    ]
                                    normalized_vendedor = id_vendedor_actual.upper()
                                    if normalized_vendedor not in existing_ids:
                                        updated_ids = existing_ids + [normalized_vendedor]
                                        cell_updates.append({
                                            "range": rowcol_to_a1(
                                                gsheet_row_index,
                                                col_idx("id_vendedor_Mod"),
                                            ),
                                            "values": [[", ".join(updated_ids)]],
                                        })
                                        changes_made = True

                                if cell_updates:
                                    if col_exists("Fecha_Modificacion"):
                                        fecha_mod = datetime.now(timezone("America/Mexico_City")).strftime(
                                            "%Y-%m-%d %H:%M:%S"
                                        )
                                        cell_updates.append({
                                            "range": rowcol_to_a1(
                                                gsheet_row_index,
                                                col_idx("Fecha_Modificacion"),
                                            ),
                                            "values": [[fecha_mod]],
                                        })
                                        changes_made = True
                                    safe_batch_update(worksheet, cell_updates)

                                # 8) Mensajes y limpieza de inputs
                                if changes_made:
                                    st.session_state["reset_inputs_tab2"] = True
                                    st.session_state["show_success_message"] = True
                                    st.session_state["last_updated_order_id"] = selected_order_id
                                    st.session_state["last_updated_cliente"] = str(
                                        selected_row_data.get("Cliente", "")
                                    ).strip()
                                    if tab2_is_active and st.session_state.get("current_tab_index") == TAB_INDEX_TAB2:
                                        st.query_params.update({"tab": str(TAB_INDEX_TAB2)})  # mantener UX actual
                                    rerun_with_tab2_loading("⏳ Guardando cambios del pedido...")
                                else:
                                    feedback_slot.empty()
                                    feedback_slot.info("ℹ️ No se detectaron cambios nuevos para guardar.")

                            except Exception as e:
                                feedback_slot.empty()
                                feedback_slot.error(f"❌ Error inesperado al guardar: {e}")

                if (
                    st.session_state.get("show_success_message")
                    and st.session_state.get("last_updated_order_id")
                ):
                    cliente_actualizado = str(st.session_state.get("last_updated_cliente", "")).strip()
                    referencia_actualizada = (
                        cliente_actualizado
                        or st.session_state.get("last_updated_order_id")
                    )
                    st.success(
                        f"🎉 ¡Cambios guardados con éxito para el pedido **{referencia_actualizada}**!"
                    )
                    if st.button("Aceptar", key="ack_mod_success"):
                        for state_key in (
                            "show_success_message",
                            "last_updated_order_id",
                            "last_updated_cliente",
                            "_mod_tab2_success_feedback_sent",
                        ):
                            st.session_state.pop(state_key, None)
                        st.rerun()
                    if not st.session_state.get("_mod_tab2_success_feedback_sent"):
                        st.toast(f"✅ Pedido de {referencia_actualizada} actualizado", icon="📦")
                        st.session_state["_mod_tab2_success_feedback_sent"] = True


# --- TAB 3: PENDING PROOF OF PAYMENT ---
with tab3:
    tab3_is_active = default_tab == TAB_INDEX_TAB3
    if tab3_is_active:
        st.session_state["current_tab_index"] = TAB_INDEX_TAB3
    st.header("🧾 Pedidos Pendientes de Comprobante")

    df_pedidos_comprobante = pd.DataFrame()
    worksheets_by_source: dict[str, object] = {}
    headers_by_source: dict[str, list[str]] = {}

    if tab3_is_active:
        try:
            tab3_refresh_token = st.session_state.get(
                "tab3_pending_comprobante_refresh_token",
                0.0,
            )
            df_pedidos_comprobante, headers_by_source = get_tab3_pending_comprobante_dataset(
                tab3_refresh_token
            )

            for source_name, getter in (
                (SHEET_PEDIDOS_HISTORICOS, get_worksheet_historico),
                (SHEET_PEDIDOS_OPERATIVOS, get_worksheet_operativa),
            ):
                if source_name not in headers_by_source:
                    continue
                worksheet_source = getter()
                if worksheet_source is not None:
                    worksheets_by_source[source_name] = worksheet_source

            if df_pedidos_comprobante.empty:
                st.warning("No se encontraron datos disponibles en las hojas de pedidos para comprobantes.")
        except Exception as e:
            st.error(f"❌ Error al cargar pedidos para comprobante: {e}")
    else:
        render_lazy_tab_placeholder(
            3,
            "tab3_lazy",
            LAZY_TAB_MESSAGE,
        )

    if not tab3_is_active:
        pass
    elif df_pedidos_comprobante.empty:
        st.info("No hay pedidos registrados.")
    else:
        filtered_pedidos_comprobante = df_pedidos_comprobante.copy()

        col3_tab3, col4_tab3 = st.columns(2)
        with col3_tab3:
            if 'Vendedor_Registro' in filtered_pedidos_comprobante.columns:
                vendedores_detectados = sorted(filtered_pedidos_comprobante['Vendedor_Registro'].dropna().astype(str).unique().tolist())
                vendedores_extra = [v for v in vendedores_detectados if v not in VENDEDORES_LIST]
                unique_vendedores_comp = ["Todos"] + VENDEDORES_LIST + vendedores_extra
                selected_vendedor_comp = st.selectbox(
                    "Filtrar por Vendedor:",
                    options=unique_vendedores_comp,
                    index=ensure_selectbox_vendor_default("comprobante_vendedor_filter", unique_vendedores_comp),
                    key="comprobante_vendedor_filter"
                )
                if selected_vendedor_comp != "Todos":
                    filtered_pedidos_comprobante = filtered_pedidos_comprobante[filtered_pedidos_comprobante['Vendedor_Registro'] == selected_vendedor_comp]

        with col4_tab3:
            (
                fecha_inicio_comp,
                fecha_fin_comp,
                _rango_activo_comp,
                rango_valido_comp,
            ) = render_date_filter_controls(
                "📅 Filtrar por Fecha de Registro:",
                "tab3_comprobantes_filtro",
                recent_days_option=7,
                recent_days_label="Mostrar solo últimos 7 días",
            )

        # Filtrar por fecha si existe la columna 'Hora_Registro'
        if 'Hora_Registro' in filtered_pedidos_comprobante.columns:
            filtered_pedidos_comprobante = filtered_pedidos_comprobante.copy()
            filtered_pedidos_comprobante['Hora_Registro'] = pd.to_datetime(
                filtered_pedidos_comprobante['Hora_Registro'],
                errors='coerce'
            )
            if rango_valido_comp:
                start_dt = datetime.combine(fecha_inicio_comp, datetime.min.time())
                end_dt = datetime.combine(fecha_fin_comp, datetime.max.time())
                filtered_pedidos_comprobante = filtered_pedidos_comprobante[
                    filtered_pedidos_comprobante['Hora_Registro'].between(start_dt, end_dt)
                ]
            else:
                filtered_pedidos_comprobante = filtered_pedidos_comprobante.iloc[0:0]

        filtered_pedidos_comprobante = filtered_pedidos_comprobante[
            filtered_pedidos_comprobante['ID_Pedido'].astype(str).str.strip().ne('') &
            filtered_pedidos_comprobante['Cliente'].astype(str).str.strip().ne('') &
            filtered_pedidos_comprobante['Folio_Factura'].astype(str).str.strip().ne('')
        ]

        if 'Estado_Pago' in filtered_pedidos_comprobante.columns and 'Adjuntos' in filtered_pedidos_comprobante.columns:
            pedidos_sin_comprobante = filtered_pedidos_comprobante[
                (filtered_pedidos_comprobante['Estado_Pago'] == '🔴 No Pagado') &
                (~filtered_pedidos_comprobante['Adjuntos'].astype(str).str.contains('comprobante', na=False, case=False))
            ].copy()
        else:
            st.warning("Las columnas 'Estado_Pago' o 'Adjuntos' no se encontraron. No se puede filtrar por comprobantes.")
            pedidos_sin_comprobante = pd.DataFrame()

        if pedidos_sin_comprobante.empty:
            if not rango_valido_comp:
                st.info("Ajusta el rango de fechas para continuar.")
            else:
                st.success("🎉 Todos los pedidos están marcados como pagados o tienen comprobante.")
        else:
            st.warning(f"⚠️ Hay {len(pedidos_sin_comprobante)} pedidos pendientes de comprobante.")

            columnas_mostrar = [
                'ID_Pedido', 'Cliente', 'Folio_Factura', 'Vendedor_Registro', 'Tipo_Envio', 'Turno',
                'Fecha_Entrega', 'Estado', 'Estado_Pago', 'Comentario', 'Modificacion_Surtido', 'Adjuntos', 'Adjuntos_Surtido'
            ]
            columnas_mostrar = [c for c in columnas_mostrar if c in pedidos_sin_comprobante.columns]

            st.dataframe(pedidos_sin_comprobante[columnas_mostrar].sort_values(by='Fecha_Entrega'), use_container_width=True, hide_index=True)

            # ✅ Bloque de subida o marca sin comprobante SOLO si hay pedidos pendientes
            st.markdown("---")
            st.subheader("Subir Comprobante para un Pedido")

            # 🆕 Ordenar por Fecha_Entrega descendente para mostrar los más recientes primero
            if 'Fecha_Entrega' in pedidos_sin_comprobante.columns:
                pedidos_sin_comprobante['Fecha_Entrega'] = pd.to_datetime(pedidos_sin_comprobante['Fecha_Entrega'], errors='coerce')
                pedidos_sin_comprobante = pedidos_sin_comprobante.sort_values(by='Fecha_Entrega', ascending=False).reset_index(drop=True)



            pedidos_sin_comprobante['display_label'] = pedidos_sin_comprobante.apply(lambda row:
                f"📄 {row.get('Folio_Factura', 'N/A') or row.get('ID_Pedido', 'N/A')} - {row.get('Cliente', 'N/A')} - {row.get('Estado', 'N/A')} [{row.get('Fuente', 'N/A')}]", axis=1)
            base_option_values = pedidos_sin_comprobante.apply(
                lambda row: (
                    f"{str(row.get('Fuente', SHEET_PEDIDOS_HISTORICOS)).strip()}|"
                    f"{str(row.get('ID_Pedido', '')).strip() or 'sin_id'}|"
                    f"{parse_sheet_row_number(row.get('Sheet_Row_Number')) or 'sin_fila'}"
                ),
                axis=1,
            )
            pedidos_sin_comprobante['option_value'] = base_option_values
            duplicate_mask = pedidos_sin_comprobante['option_value'].duplicated(keep=False)
            if duplicate_mask.any():
                pedidos_sin_comprobante.loc[duplicate_mask, 'option_value'] = pedidos_sin_comprobante.loc[duplicate_mask].apply(
                    lambda row: f"{base_option_values[row.name]}|{row.name}",
                    axis=1,
                )

            option_label_map = dict(zip(pedidos_sin_comprobante['option_value'], pedidos_sin_comprobante['display_label']))

            selected_pending_option_key = st.selectbox(
                "📝 Seleccionar Pedido para Subir Comprobante",
                list(option_label_map.keys()),
                format_func=lambda option_key: option_label_map.get(option_key, option_key),
                key="select_pending_order_comprobante"
            )

            if selected_pending_option_key:
                selected_pending_row_data = pedidos_sin_comprobante[
                    pedidos_sin_comprobante['option_value'] == selected_pending_option_key
                ].iloc[0]
                selected_pending_order_id = str(selected_pending_row_data.get('ID_Pedido', '')).strip()
                selected_pending_folio = str(selected_pending_row_data.get('Folio_Factura', '')).strip().upper()
                selected_pending_cliente = str(selected_pending_row_data.get('Cliente', '')).strip().upper()
                selected_source_name = str(selected_pending_row_data.get('Fuente', SHEET_PEDIDOS_HISTORICOS)).strip()
                worksheet_obj = worksheets_by_source.get(selected_source_name)
                headers_source = headers_by_source.get(selected_source_name, [])
                sheet_row_number = parse_sheet_row_number(selected_pending_row_data.get('Sheet_Row_Number'))

                st.info(
                    f"Subiendo comprobante para: Folio {selected_pending_row_data.get('Folio_Factura')} "
                    f"(ID {selected_pending_order_id}) en {selected_source_name}"
                )

                with st.form(key=f"upload_comprobante_form_{selected_pending_order_id}"):
                    comprobante_files = st.file_uploader(
                        "💲 Comprobante(s) de Pago",
                        type=["pdf", "jpg", "jpeg", "png"],
                        accept_multiple_files=True,
                        key=f"comprobante_uploader_{selected_pending_order_id}"
                    )

                    submit_comprobante = st.form_submit_button("✅ Subir Comprobante y Actualizar Estado")

                    if submit_comprobante:
                        if comprobante_files:
                            try:
                                if worksheet_obj is None:
                                    st.error("❌ No se encontró la hoja de origen del pedido seleccionado.")
                                    st.stop()
                                if not headers_source:
                                    headers_source = worksheet_obj.row_values(1)
                                if "ID_Pedido" not in headers_source:
                                    st.error("❌ La hoja no contiene la columna 'ID_Pedido'.")
                                    st.stop()

                                id_col_idx = headers_source.index("ID_Pedido")
                                folio_col_idx = headers_source.index("Folio_Factura") if "Folio_Factura" in headers_source else None
                                cliente_col_idx = headers_source.index("Cliente") if "Cliente" in headers_source else None
                                all_values_source = worksheet_obj.get_all_values()

                                def _row_value(row_values, idx):
                                    if idx is None or len(row_values) <= idx:
                                        return ""
                                    return str(row_values[idx]).strip().upper()

                                def _resolve_target_row():
                                    if sheet_row_number and sheet_row_number <= len(all_values_source):
                                        candidate = all_values_source[sheet_row_number - 1]
                                        candidate_id = str(candidate[id_col_idx]).strip() if len(candidate) > id_col_idx else ""
                                        folio_ok = bool(selected_pending_folio) and _row_value(candidate, folio_col_idx) == selected_pending_folio
                                        cliente_ok = bool(selected_pending_cliente) and _row_value(candidate, cliente_col_idx) == selected_pending_cliente
                                        if candidate_id == selected_pending_order_id and (folio_ok or cliente_ok):
                                            return int(sheet_row_number)

                                    matches = []
                                    for row_number, row_values in enumerate(all_values_source[1:], start=2):
                                        row_id = str(row_values[id_col_idx]).strip() if len(row_values) > id_col_idx else ""
                                        if row_id != selected_pending_order_id:
                                            continue
                                        score = 0
                                        if selected_pending_folio and _row_value(row_values, folio_col_idx) == selected_pending_folio:
                                            score += 2
                                        if selected_pending_cliente and _row_value(row_values, cliente_col_idx) == selected_pending_cliente:
                                            score += 1
                                        matches.append((score, row_number))
                                    if not matches:
                                        return None
                                    matches.sort(reverse=True)
                                    return matches[0][1]

                                sheet_row = _resolve_target_row()
                                if not sheet_row:
                                    st.error("❌ No se pudo resolver la fila real del pedido seleccionado en Google Sheets.")
                                    st.stop()

                                current_row_values = all_values_source[sheet_row - 1] if sheet_row <= len(all_values_source) else worksheet_obj.row_values(sheet_row)
                                current_row_id = str(current_row_values[id_col_idx]).strip() if len(current_row_values) > id_col_idx else ""
                                if current_row_id != selected_pending_order_id:
                                    st.error("❌ Validación de seguridad: la fila encontrada no coincide con el pedido seleccionado.")
                                    st.stop()

                                new_urls = []
                                for archivo in comprobante_files:
                                    ext = os.path.splitext(archivo.name)[1]
                                    s3_key = f"{selected_pending_order_id}/comprobante_{selected_pending_order_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}_{uuid.uuid4().hex[:4]}{ext}"
                                    success, url, error_msg = upload_file_to_s3(s3_client, S3_BUCKET_NAME, archivo, s3_key)
                                    if success:
                                        new_urls.append(url)
                                    else:
                                        st.warning(f"⚠️ Falló la subida de {archivo.name}: {error_msg or 'Error desconocido'}")

                                if new_urls:
                                    adj_col_idx = headers_source.index('Adjuntos')
                                    current_adjuntos = str(current_row_values[adj_col_idx]).strip() if len(current_row_values) > adj_col_idx else ''
                                    adjuntos_list = [x.strip() for x in current_adjuntos.split(',') if x.strip()]
                                    adjuntos_list.extend(new_urls)

                                    required_cols = ['Adjuntos', 'Estado_Pago', 'Fecha_Pago_Comprobante']
                                    missing_cols = [col for col in required_cols if col not in headers_source]
                                    if missing_cols:
                                        st.error(f"❌ Faltan columnas requeridas en la hoja: {', '.join(missing_cols)}")
                                        st.stop()

                                    updates = [
                                        {
                                            "range": rowcol_to_a1(
                                                sheet_row,
                                                headers_source.index('Adjuntos') + 1,
                                            ),
                                            "values": [[", ".join(adjuntos_list)]],
                                        },
                                        {
                                            "range": rowcol_to_a1(
                                                sheet_row,
                                                headers_source.index('Estado_Pago') + 1,
                                            ),
                                            "values": [["✅ Pagado"]],
                                        },
                                        {
                                            "range": rowcol_to_a1(
                                                sheet_row,
                                                headers_source.index('Fecha_Pago_Comprobante') + 1,
                                            ),
                                            "values": [[datetime.now(timezone("America/Mexico_City")).strftime('%Y-%m-%d')]],
                                        },
                                    ]
                                    safe_batch_update(worksheet_obj, updates)

                                    st.success("✅ Comprobantes subidos y estado actualizado con éxito.")
                                    st.session_state["tab3_pending_comprobante_refresh_token"] = time.time()
                                    get_tab3_pending_comprobante_dataset.clear()
                                    st.rerun()
                                else:
                                    st.warning("⚠️ No se subió ningún archivo correctamente.")
                            except Exception as e:
                                st.error(f"❌ Error al subir comprobantes: {e}")
                        else:
                            st.warning("⚠️ Por favor, sube al menos un archivo.")

                if st.button("✅ Marcar como Pagado sin Comprobante", key=f"btn_sin_cp_{selected_pending_order_id}"):
                    try:
                        if worksheet_obj is None:
                            st.error("❌ No se encontró la hoja de origen del pedido seleccionado.")
                            st.stop()
                        if not headers_source:
                            headers_source = worksheet_obj.row_values(1)
                        if 'ID_Pedido' not in headers_source:
                            st.error("❌ La hoja no contiene la columna 'ID_Pedido'.")
                            st.stop()
                        if 'Estado_Pago' not in headers_source:
                            st.error("❌ La hoja no contiene la columna 'Estado_Pago'.")
                            st.stop()

                        id_col_idx = headers_source.index('ID_Pedido')
                        folio_col_idx = headers_source.index('Folio_Factura') if 'Folio_Factura' in headers_source else None
                        cliente_col_idx = headers_source.index('Cliente') if 'Cliente' in headers_source else None
                        all_values_source = worksheet_obj.get_all_values()

                        def _row_value(row_values, idx):
                            if idx is None or len(row_values) <= idx:
                                return ""
                            return str(row_values[idx]).strip().upper()

                        sheet_row = None
                        if sheet_row_number and sheet_row_number <= len(all_values_source):
                            candidate = all_values_source[sheet_row_number - 1]
                            candidate_id = str(candidate[id_col_idx]).strip() if len(candidate) > id_col_idx else ''
                            folio_ok = bool(selected_pending_folio) and _row_value(candidate, folio_col_idx) == selected_pending_folio
                            cliente_ok = bool(selected_pending_cliente) and _row_value(candidate, cliente_col_idx) == selected_pending_cliente
                            if candidate_id == selected_pending_order_id and (folio_ok or cliente_ok):
                                sheet_row = int(sheet_row_number)

                        if sheet_row is None:
                            matches = []
                            for row_number, row_values in enumerate(all_values_source[1:], start=2):
                                row_id = str(row_values[id_col_idx]).strip() if len(row_values) > id_col_idx else ''
                                if row_id != selected_pending_order_id:
                                    continue
                                score = 0
                                if selected_pending_folio and _row_value(row_values, folio_col_idx) == selected_pending_folio:
                                    score += 2
                                if selected_pending_cliente and _row_value(row_values, cliente_col_idx) == selected_pending_cliente:
                                    score += 1
                                matches.append((score, row_number))
                            if matches:
                                matches.sort(reverse=True)
                                sheet_row = matches[0][1]

                        if sheet_row is None:
                            st.error("❌ No se pudo resolver la fila real del pedido seleccionado en Google Sheets.")
                            st.stop()

                        resolved_row_values = all_values_source[sheet_row - 1] if sheet_row <= len(all_values_source) else worksheet_obj.row_values(sheet_row)
                        resolved_row_id = str(resolved_row_values[id_col_idx]).strip() if len(resolved_row_values) > id_col_idx else ''
                        if resolved_row_id != selected_pending_order_id:
                            st.error("❌ Validación de seguridad: la fila encontrada no coincide con el pedido seleccionado.")
                            st.stop()

                        updates = [
                            {
                                "range": rowcol_to_a1(
                                    sheet_row,
                                    headers_source.index('Estado_Pago') + 1,
                                ),
                                "values": [["✅ Pagado"]],
                            }
                        ]

                        if 'Fecha_Pago_Comprobante' in headers_source:
                            updates.append({
                                "range": rowcol_to_a1(
                                    sheet_row,
                                    headers_source.index('Fecha_Pago_Comprobante') + 1,
                                ),
                                "values": [[datetime.now(timezone("America/Mexico_City")).strftime('%Y-%m-%d')]],
                            })

                        safe_batch_update(worksheet_obj, updates)

                        st.success("✅ Pedido marcado como pagado sin comprobante.")
                        st.session_state["tab3_pending_comprobante_refresh_token"] = time.time()
                        get_tab3_pending_comprobante_dataset.clear()
                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ Error al marcar como pagado sin comprobante: {e}")

# ----------------- HELPERS FALTANTES -----------------

def partir_urls(value):
    """
    Normaliza un campo de adjuntos que puede venir como JSON (lista o dict),
    o como texto separado por comas/; / saltos de línea. Devuelve lista de URLs únicas.
    """
    if value is None:
        return []
    s = str(value).strip()
    if not s or s.lower() in ("nan", "none", "n/a"):
        return []
    urls = []
    # Intento como JSON
    try:
        obj = json.loads(s)
        if isinstance(obj, list):
            for it in obj:
                if isinstance(it, str) and it.strip():
                    urls.append(it.strip())
                elif isinstance(it, dict):
                    for k in ("url", "URL", "href", "link"):
                        if k in it and str(it[k]).strip():
                            urls.append(str(it[k]).strip())
        elif isinstance(obj, dict):
            for k in ("url", "URL", "href", "link"):
                if k in obj and str(obj[k]).strip():
                    urls.append(str(obj[k]).strip())
    except Exception:
        # Separadores comunes
        for p in re.split(r"[,\n;]+", s):
            p = p.strip()
            if p:
                urls.append(p)
    # De-duplicar preservando orden
    out, seen = [], set()
    for u in urls:
        if u not in seen:
            seen.add(u); out.append(u)
    return out


@st.cache_data(ttl=90)
def get_tab4_casos_especiales_dataset(
    refresh_token: float | None = None,
) -> tuple[pd.DataFrame, list[str]]:
    """Carga casos especiales con número de fila para la pestaña 4."""
    _ = refresh_token

    ws = get_worksheet_casos_especiales()
    df, headers = load_sheet_records_with_row_numbers(ws)
    if df.empty:
        return pd.DataFrame(), headers

    columnas_necesarias = [
        "ID_Pedido","Cliente","Vendedor_Registro","Folio_Factura","Folio_Factura_Error",
        "Hora_Registro","Tipo_Envio","Tipo_Caso","Estado","Estado_Caso","Turno",
        "Refacturacion_Tipo","Refacturacion_Subtipo","Folio_Factura_Refacturada",
        "Resultado_Esperado","Motivo_Detallado","Material_Devuelto","Monto_Devuelto","Motivo_NotaVenta",
        "Area_Responsable","Nombre_Responsable","Numero_Cliente_RFC","Tipo_Envio_Original","Estatus_OrigenF",
        "Direccion_Guia_Retorno","Direccion_Envio","Numero_Serie","Fecha_Compra",
        "Fecha_Entrega","Fecha_Recepcion_Devolucion","Estado_Recepcion",
        "Nota_Credito_URL","Documento_Adicional_URL","Comentarios_Admin_Devolucion",
        "Modificacion_Surtido","Adjuntos_Surtido","Adjuntos","Hoja_Ruta_Mensajero",
        "Hora_Proceso","Seguimiento","id_vendedor",
    ]
    for col in columnas_necesarias:
        if col not in df.columns:
            df[col] = ""

    if "Fecha_Compra" not in df.columns and "FechaCompra" in df.columns:
        df["Fecha_Compra"] = df["FechaCompra"]
    elif "Fecha_Compra" in df.columns and "FechaCompra" in df.columns and df["Fecha_Compra"].eq("").all():
        df["Fecha_Compra"] = df["Fecha_Compra"].where(df["Fecha_Compra"].astype(str).str.strip() != "", df["FechaCompra"])

    return df, headers


@st.cache_data(ttl=300)
def cargar_casos_especiales():
    """
    Lee la hoja 'casos_especiales' usando tu helper get_worksheet_casos_especiales()
    y garantiza todas las columnas que la UI usa.
    """
    ws = get_worksheet_casos_especiales()
    data = ws.get_all_records()
    df = pd.DataFrame(data)

    columnas_necesarias = [
        # Identificación y encabezado
        "ID_Pedido","Cliente","Vendedor_Registro","Folio_Factura","Folio_Factura_Error",
        "Hora_Registro","Tipo_Envio","Estado","Estado_Caso","Turno",
        # Refacturación
        "Refacturacion_Tipo","Refacturacion_Subtipo","Folio_Factura_Refacturada",
        # Detalle del caso
        "Resultado_Esperado","Motivo_Detallado","Material_Devuelto","Monto_Devuelto","Motivo_NotaVenta",
        "Area_Responsable","Nombre_Responsable","Numero_Cliente_RFC","Tipo_Envio_Original","Estatus_OrigenF",
        "Direccion_Guia_Retorno","Direccion_Envio",
        # ⚙️ NUEVO: Garantías
        "Numero_Serie","Fecha_Compra",  # (si tu hoja usa 'FechaCompra', abajo la normalizamos)
        # Fechas/recepción
        "Fecha_Entrega","Fecha_Recepcion_Devolucion","Estado_Recepcion",
        # Documentos de cierre
        "Nota_Credito_URL","Documento_Adicional_URL","Comentarios_Admin_Devolucion",
        # Modificación de surtido
        "Modificacion_Surtido","Adjuntos_Surtido",
        # Adjuntos/guía
        "Adjuntos","Hoja_Ruta_Mensajero",
        # Otros
        "Hora_Proceso",
        # Seguimiento
        "Seguimiento"
    ]
    for c in columnas_necesarias:
        if c not in df.columns:
            df[c] = ""

    # Quitar casos cerrados
    df["Seguimiento"] = df["Seguimiento"].fillna("")
    df = df[~df["Seguimiento"].astype(str).str.lower().eq("cerrado")]

    # Normaliza 'Fecha_Compra' si en la hoja existe como 'FechaCompra'
    if "Fecha_Compra" not in df.columns and "FechaCompra" in df.columns:
        df["Fecha_Compra"] = df["FechaCompra"]
    elif "Fecha_Compra" in df.columns and "FechaCompra" in df.columns and df["Fecha_Compra"].eq("").all():
        # Si ambas existen pero 'Fecha_Compra' viene vacía, usa 'FechaCompra'
        df["Fecha_Compra"] = df["Fecha_Compra"].where(df["Fecha_Compra"].astype(str).str.strip() != "", df["FechaCompra"])

    return df




# --- TAB 4: CASOS ESPECIALES ---
with tab4:
    tab4_is_active = default_tab == TAB_INDEX_TAB4
    if tab4_is_active:
        st.session_state["current_tab_index"] = TAB_INDEX_TAB4
    st.header("📁 Casos Especiales")

    df_casos_ref = pd.DataFrame()
    headers_casos_ref: list[str] = []
    ws_casos_ref = None

    if tab4_is_active:
        try:
            tab4_refresh_token = st.session_state.get(
                "tab4_casos_refresh_token",
                0.0,
            )
            df_casos_ref, headers_casos_ref = get_tab4_casos_especiales_dataset(tab4_refresh_token)
            df_casos = df_casos_ref.copy()
            ws_casos_ref = get_worksheet_casos_especiales()

            if "Seguimiento" in df_casos.columns:
                df_casos["Seguimiento"] = df_casos["Seguimiento"].fillna("")
                df_casos = df_casos[~df_casos["Seguimiento"].astype(str).str.lower().eq("cerrado")]
        except Exception as e:
            st.error(f"❌ Error al cargar casos especiales: {e}")
            df_casos = pd.DataFrame()
            df_casos_ref = pd.DataFrame()
            headers_casos_ref = []
            ws_casos_ref = None
    else:
        render_lazy_tab_placeholder(
            4,
            "tab4_lazy",
            LAZY_TAB_MESSAGE,
        )
        df_casos = pd.DataFrame()

    if not tab4_is_active:
        pass
    elif df_casos.empty:
        st.info("No hay casos especiales.")
    else:
        if "id_vendedor" not in df_casos.columns:
            df_casos["id_vendedor"] = ""

        id_vendedor_sesion = normalize_vendedor_id(st.session_state.get("id_vendedor", ""))
        seguimiento_autorizacion = "Autorización de devolución"

        st.markdown("#### Devoluciones sin refacturar — ✍️ Captura el Folio Nuevo")
        st.caption("Solo se muestran devoluciones del vendedor logeado con Folio Nuevo pendiente. Captura el Folio Nuevo y guarda; se almacena con prefijo * para auditoría post-registro.")

        if ws_casos_ref is None:
            st.error("❌ No fue posible conectar con la hoja de casos especiales para devoluciones.")

        if not df_casos_ref.empty:
            for col in ["id_vendedor", "Tipo_Envio", "Tipo_Caso", "Seguimiento", "Folio_Factura", "Hora_Registro"]:
                if col not in df_casos_ref.columns:
                    df_casos_ref[col] = ""

        if df_casos_ref.empty or ws_casos_ref is None:
            st.info("No hay devoluciones sin refacturar por mostrar.")
        else:
            df_sin_refacturar = df_casos_ref[df_casos_ref.apply(is_devolucion_case_row, axis=1)].copy()
            df_sin_refacturar = df_sin_refacturar[
                df_sin_refacturar["Seguimiento"].astype(str).str.strip().eq(seguimiento_autorizacion)
                & df_sin_refacturar["Folio_Factura"].apply(is_empty_folio)
            ]
            if id_vendedor_sesion:
                df_sin_refacturar = df_sin_refacturar[
                    df_sin_refacturar["id_vendedor"].apply(normalize_vendedor_id) == id_vendedor_sesion
                ]
            else:
                df_sin_refacturar = df_sin_refacturar.iloc[0:0]

            if not df_sin_refacturar.empty:
                df_sin_refacturar["_hora_sort"] = pd.to_datetime(
                    df_sin_refacturar["Hora_Registro"], errors="coerce"
                )
                df_sin_refacturar = df_sin_refacturar.sort_values(
                    by=["_hora_sort"],
                    ascending=[True],
                    na_position="last",
                )

            if df_sin_refacturar.empty:
                st.info("No tienes devoluciones pendientes de Folio Nuevo.")
            else:
                for _, row in df_sin_refacturar.iterrows():
                    sheet_row_number = parse_sheet_row_number(row.get("Sheet_Row_Number"))
                    row_key = f"devol_sin_ref_{sheet_row_number or uuid.uuid4().hex}"
                    with st.container(border=True):
                        st.markdown(
                            f"👤 **Cliente:** {row.get('Cliente', 'N/A') or 'N/A'}  |  "
                            f"🧾 **Folio Error:** {row.get('Folio_Factura_Error', 'N/A') or 'N/A'}"
                        )
                        st.markdown(
                            f"📌 **Seguimiento:** {row.get('Seguimiento', 'N/A') or 'N/A'}  |  "
                            f"🕒 **Hora Registro:** {row.get('Hora_Registro', 'N/A') or 'N/A'}"
                        )
                        with st.form(key=f"{row_key}_form_folio_nuevo", clear_on_submit=False):
                            folio_input = st.text_input(
                                "📄 Folio Nuevo",
                                key=f"{row_key}_folio_input",
                                placeholder="Ej. F197176",
                            )

                            notas_devolucion = st.text_area(
                                "✍️ Notas de Devolucion Pendiente",
                                key=f"{row_key}_notas_devolucion",
                                height=100,
                            )
                            direccion_guia_retorno_pendiente = st.text_area(
                                "📬 Dirección Guia_Retorno (Opcional)",
                                key=f"{row_key}_direccion_guia_retorno",
                                height=80,
                                help="Si lo dejas vacío, se limpiará el valor previo en Excel/Sheets. Si capturas un valor, reemplazará el existente.",
                            )
                            uploaded_files_devolucion = st.file_uploader(
                                "📎 Subir Archivos de Devolucion",
                                type=["pdf", "jpg", "jpeg", "png", "xlsx", "docx"],
                                accept_multiple_files=True,
                                key=f"{row_key}_archivos_devolucion",
                            )
                            uploaded_comprobantes_extra = st.file_uploader(
                                "🧾 Subir Comprobante(s) Adicional(es)",
                                type=["pdf", "jpg", "jpeg", "png"],
                                accept_multiple_files=True,
                                key=f"{row_key}_comprobantes_extra",
                            )
                            submit_folio_nuevo = st.form_submit_button("Guardar Folio Nuevo")

                        if submit_folio_nuevo:
                            folio_sanitizado = str(folio_input or "").strip()
                            if not folio_sanitizado:
                                st.error("❌ El Folio Nuevo no puede estar vacío.")
                            elif not str(notas_devolucion or "").strip():
                                st.error("❌ El campo 'Notas de Devolucion Pendiente' es obligatorio.")
                            elif sheet_row_number is None:
                                st.error("❌ No se pudo identificar la fila real en Google Sheets para actualizar.")
                            else:
                                try:
                                    valor_guardar = f"*{folio_sanitizado}"
                                    row_idx = int(sheet_row_number)
                                    current_row_values = ws_casos_ref.row_values(row_idx)
                                    if len(current_row_values) < len(headers_casos_ref):
                                        current_row_values += [""] * (len(headers_casos_ref) - len(current_row_values))
                                    current_row = dict(zip(headers_casos_ref, current_row_values))

                                    cell_updates = []

                                    def col_exists(col_name: str) -> bool:
                                        return col_name in headers_casos_ref

                                    def col_idx(col_name: str) -> int:
                                        return headers_casos_ref.index(col_name) + 1

                                    if col_exists("Folio_Factura"):
                                        cell_updates.append({
                                            "range": rowcol_to_a1(row_idx, col_idx("Folio_Factura")),
                                            "values": [[valor_guardar]],
                                        })

                                    if col_exists("Modificacion_Surtido"):
                                        cell_updates.append({
                                            "range": rowcol_to_a1(row_idx, col_idx("Modificacion_Surtido")),
                                            "values": [[str(notas_devolucion).strip()]],
                                        })

                                    direccion_guia_retorno_normalizada = str(direccion_guia_retorno_pendiente or "").strip()
                                    if col_exists("Direccion_Guia_Retorno"):
                                        # Siempre sobrescribir: vacío => limpia celda, con valor => reemplaza contenido previo.
                                        cell_updates.append({
                                            "range": rowcol_to_a1(row_idx, col_idx("Direccion_Guia_Retorno")),
                                            "values": [[direccion_guia_retorno_normalizada]],
                                        })

                                    new_adjuntos_surtido_urls = []
                                    archivos_devolucion_subidos = []
                                    if uploaded_files_devolucion:
                                        for f in uploaded_files_devolucion:
                                            ext = os.path.splitext(f.name)[1]
                                            s3_key = f"{row.get('ID_Pedido','sin_id')}/devolucion_{f.name.replace(' ', '_').replace(ext, '')}_{uuid.uuid4().hex[:4]}{ext}"
                                            success, url, error_msg = upload_file_to_s3(s3_client, S3_BUCKET_NAME, f, s3_key)
                                            if success:
                                                new_adjuntos_surtido_urls.append(url)
                                                archivos_devolucion_subidos.append(f.name)
                                            else:
                                                st.warning(f"⚠️ Falló la subida de {f.name}: {error_msg or 'Error desconocido'}")

                                    if new_adjuntos_surtido_urls and col_exists("Adjuntos_Surtido"):
                                        current_urls = [x.strip() for x in str(current_row.get("Adjuntos_Surtido", "")).split(",") if x.strip()]
                                        updated_adjuntos_surtido = ", ".join(current_urls + new_adjuntos_surtido_urls)
                                        cell_updates.append({
                                            "range": rowcol_to_a1(row_idx, col_idx("Adjuntos_Surtido")),
                                            "values": [[updated_adjuntos_surtido]],
                                        })

                                    comprobante_urls = []
                                    comprobantes_subidos = []
                                    if uploaded_comprobantes_extra:
                                        for archivo in uploaded_comprobantes_extra:
                                            ext = os.path.splitext(archivo.name)[1]
                                            s3_key = f"{row.get('ID_Pedido','sin_id')}/comprobante_devolucion_{datetime.now().strftime('%Y%m%d%H%M%S')}_{uuid.uuid4().hex[:4]}{ext}"
                                            success, url, error_msg = upload_file_to_s3(s3_client, S3_BUCKET_NAME, archivo, s3_key)
                                            if success:
                                                comprobante_urls.append(url)
                                                comprobantes_subidos.append(archivo.name)
                                            else:
                                                st.warning(f"⚠️ Falló la subida del comprobante {archivo.name}: {error_msg or 'Error desconocido'}")

                                    if comprobante_urls and col_exists("Adjuntos"):
                                        current_adjuntos = [x.strip() for x in str(current_row.get("Adjuntos", "")).split(",") if x.strip()]
                                        updated_adjuntos = ", ".join(current_adjuntos + comprobante_urls)
                                        cell_updates.append({
                                            "range": rowcol_to_a1(row_idx, col_idx("Adjuntos")),
                                            "values": [[updated_adjuntos]],
                                        })

                                    if col_exists("Estado"):
                                        cell_updates.append({
                                            "range": rowcol_to_a1(row_idx, col_idx("Estado")),
                                            "values": [["✏️ Modificación"]],
                                        })

                                    completado_actual = str(current_row.get("Completados_Limpiado", "")).strip().lower()
                                    if col_exists("Completados_Limpiado") and completado_actual in {"si", "sí"}:
                                        cell_updates.append({
                                            "range": rowcol_to_a1(row_idx, col_idx("Completados_Limpiado")),
                                            "values": [[""]],
                                        })

                                    if cell_updates:
                                        safe_batch_update(ws_casos_ref, cell_updates)

                                    st.success(f"✅ Folio Nuevo guardado correctamente: {folio_sanitizado}")
                                    st.info("📝 Notas de devolución pendientes actualizadas correctamente.")

                                    if archivos_devolucion_subidos:
                                        st.success(
                                            "📎 Archivos de devolución subidos correctamente: "
                                            + ", ".join(archivos_devolucion_subidos)
                                        )

                                    if comprobantes_subidos:
                                        st.success(
                                            "🧾 Comprobante(s) adicional(es) subido(s) correctamente: "
                                            + ", ".join(comprobantes_subidos)
                                        )

                                    if not archivos_devolucion_subidos and not comprobantes_subidos:
                                        st.info("ℹ️ No se adjuntaron archivos en este guardado.")

                                    st.session_state.pop(f"{row_key}_folio_input", None)
                                    st.session_state.pop(f"{row_key}_notas_devolucion", None)
                                    st.session_state.pop(f"{row_key}_direccion_guia_retorno", None)
                                    cargar_casos_especiales.clear()
                                    get_tab4_casos_especiales_dataset.clear()
                                    st.session_state["tab4_casos_refresh_token"] = time.time()
                                    obtener_devoluciones_autorizadas_sin_folio.clear()
                                    st.rerun()
                                except Exception as e:
                                    st.error(f"❌ No se pudo guardar el Folio Nuevo: {e}")

        df_casos = df_casos[
            df_casos["Tipo_Envio"].isin(["🔁 Devolución", "🛠 Garantía"]) &
            (df_casos["Seguimiento"] != "Cerrado")
        ]

        if df_casos.empty:
            st.info("No hay casos especiales abiertos.")
        else:
            with st.expander("🔎 Filtrar y revisar casos especiales", expanded=True):
                if "Hora_Registro" in df_casos.columns:
                    df_casos["Hora_Registro"] = pd.to_datetime(df_casos["Hora_Registro"], errors="coerce")

                if "Vendedor_Registro" in df_casos.columns:
                    df_casos["Vendedor_Registro"] = (
                        df_casos["Vendedor_Registro"].astype(str).str.strip()
                    )

                col_vend_casos, col_fecha_casos = st.columns(2)

                with col_vend_casos:
                    vendedores_casos = ["Todos"]
                    if "Vendedor_Registro" in df_casos.columns:
                        vendedores_casos = build_vendor_filter_options(
                            df_casos["Vendedor_Registro"].dropna().astype(str).tolist(),
                        )
                    selected_vendedor_casos = st.selectbox(
                        "Filtrar por Vendedor:",
                        options=vendedores_casos,
                        index=ensure_selectbox_vendor_default("filtro_vendedor_casos_especiales", vendedores_casos),
                        key="filtro_vendedor_casos_especiales"
                    )

                with col_fecha_casos:
                    (
                        fecha_inicio_casos,
                        fecha_fin_casos,
                        _rango_activo_casos,
                        rango_valido_casos,
                    ) = render_date_filter_controls(
                        "📅 Filtrar por Fecha de Registro:",
                        "tab4_casos_filtro",
                        recent_days_option=7,
                        recent_days_label="Mostrar solo últimos 7 días",
                    )

                filtered_casos = df_casos.copy()

                if (
                    selected_vendedor_casos != "Todos"
                    and "Vendedor_Registro" in filtered_casos.columns
                ):
                    filtered_casos = filtered_casos[
                        filtered_casos["Vendedor_Registro"] == selected_vendedor_casos
                    ]

                if "Hora_Registro" in filtered_casos.columns:
                    filtered_casos = filtered_casos.copy()
                    hora_registro_series = filtered_casos["Hora_Registro"]
                    if not pd.api.types.is_datetime64_any_dtype(hora_registro_series):
                        hora_registro_series = pd.to_datetime(
                            hora_registro_series,
                            errors="coerce",
                        )
                        filtered_casos["Hora_Registro"] = hora_registro_series

                    if rango_valido_casos:
                        start_dt = datetime.combine(fecha_inicio_casos, datetime.min.time())
                        end_dt = datetime.combine(fecha_fin_casos, datetime.max.time())
                        filtered_casos = filtered_casos[
                            hora_registro_series.between(start_dt, end_dt)
                        ]
                    else:
                        filtered_casos = filtered_casos.iloc[0:0]

                if filtered_casos.empty:
                    if not rango_valido_casos:
                        st.info("Ajusta el rango de fechas para continuar.")
                    else:
                        st.warning("No hay casos especiales que coincidan con los filtros seleccionados.")
                else:
                    filtered_casos = filtered_casos.reset_index(drop=True)
                    columnas_mostrar = ["Estado","Cliente","Vendedor_Registro","Tipo_Envio","Seguimiento"]
                    st.dataframe(filtered_casos[columnas_mostrar], use_container_width=True, hide_index=True)

                    filtered_casos = filtered_casos.copy()
                    filtered_casos["display_label"] = filtered_casos.apply(
                        lambda r: f"{r['Estado']} - {r['Cliente']} ({r['Tipo_Envio']})", axis=1
                    )
                    selected_case = st.selectbox(
                        "📂 Selecciona un caso para ver detalles",
                        filtered_casos["display_label"].tolist(),
                        key="select_caso_especial_tab4"
                    )

                    if selected_case:
                        case_row = filtered_casos[
                            filtered_casos["display_label"] == selected_case
                        ].iloc[0]
                        render_caso_especial(case_row)


# --- TAB 5: GUIAS CARGADAS ---
def fijar_tab5_activa():
    """Mantiene referencia de pestaña activa sin tocar query params en render."""
    st.session_state["current_tab_index"] = TAB_INDEX_TAB5

@st.cache_data(ttl=60)
def cargar_datos_guias_unificadas(refresh_token: float | None = None):
    # ---------- A) hojas de pedidos (histórico + operativa) ----------
    _ = refresh_token
    def _normalizar_guias_pedidos(df_ped: pd.DataFrame, fuente: str) -> pd.DataFrame:
        if df_ped.empty:
            return pd.DataFrame()

        for col in ["ID_Pedido","Cliente","Vendedor_Registro","Tipo_Envio","Estado",
                    "Fecha_Entrega","Hora_Registro","Folio_Factura","Adjuntos_Guia","id_vendedor","Completados_Limpiado"]:
            if col not in df_ped.columns:
                df_ped[col] = ""

        df_res = df_ped[df_ped["Adjuntos_Guia"].astype(str).str.strip() != ""].copy()
        if df_res.empty:
            return df_res

        df_res["Fuente"] = fuente
        df_res["URLs_Guia"] = df_res["Adjuntos_Guia"].astype(str)
        df_res["Ultima_Guia"] = df_res["URLs_Guia"].apply(
            lambda s: s.split(",")[-1].strip() if isinstance(s, str) and s.strip() else ""
        )
        return df_res

    # datos_pedidos (histórico)
    try:
        ws_ped_hist = get_worksheet_historico(refresh_token)
        df_ped_hist = pd.DataFrame(ws_ped_hist.get_all_records())
    except Exception:
        df_ped_hist = pd.DataFrame()

    # data_pedidos (operativa)
    try:
        ws_ped_op = get_worksheet_operativa(refresh_token)
        df_ped_op = pd.DataFrame(ws_ped_op.get_all_records())
    except Exception:
        df_ped_op = pd.DataFrame()

    df_a_hist = _normalizar_guias_pedidos(df_ped_hist, SHEET_PEDIDOS_HISTORICOS)
    df_a_op = _normalizar_guias_pedidos(df_ped_op, SHEET_PEDIDOS_OPERATIVOS)

    # ---------- B) casos_especiales ----------
    try:
        ws_casos = get_worksheet_casos_especiales()
        df_casos = pd.DataFrame(ws_casos.get_all_records())
    except Exception:
        df_casos = pd.DataFrame()

    if df_casos.empty:
        df_b = pd.DataFrame(columns=[
            "ID_Pedido","Cliente","Vendedor_Registro","Tipo_Envio","Estado",
            "Fecha_Entrega","Hora_Registro","Folio_Factura","Adjuntos_Guia",
            "URLs_Guia","Ultima_Guia","Fuente","id_vendedor","Completados_Limpiado"
        ])
    else:
        for col in ["ID_Pedido","Cliente","Vendedor_Registro","Tipo_Envio","Estado",
                    "Fecha_Entrega","Hora_Registro","Folio_Factura","Hoja_Ruta_Mensajero","Tipo_Caso","id_vendedor","Completados_Limpiado"]:
            if col not in df_casos.columns:
                df_casos[col] = ""

        df_b = df_casos[df_casos["Hoja_Ruta_Mensajero"].astype(str).str.strip() != ""].copy()
        if df_b.empty:
            df_b = pd.DataFrame(columns=[
                "ID_Pedido","Cliente","Vendedor_Registro","Tipo_Envio","Estado",
                "Fecha_Entrega","Hora_Registro","Folio_Factura","Adjuntos_Guia",
                "URLs_Guia","Ultima_Guia","Fuente","id_vendedor","Completados_Limpiado"
            ])
        else:
            df_b["Adjuntos_Guia"] = df_b["Hoja_Ruta_Mensajero"].astype(str)
            df_b["URLs_Guia"] = df_b["Adjuntos_Guia"]
            df_b["Ultima_Guia"] = df_b["URLs_Guia"].apply(
                lambda s: s.split(",")[-1].strip() if isinstance(s, str) and s.strip() else ""
            )

            def _infer_tipo_envio(row):
                t_env = str(row.get("Tipo_Envio","")).strip()
                if t_env:
                    return t_env
                t_caso = str(row.get("Tipo_Caso","")).lower()
                if t_caso.startswith("devol"):
                    return "🔁 Devolución"
                if t_caso.startswith("garan"):
                    return "🛠 Garantía"
                return "Caso especial"
            df_b["Tipo_Envio"] = df_b.apply(_infer_tipo_envio, axis=1)
            df_b["Fuente"] = "casos_especiales"

        for col in ["Adjuntos_Guia","URLs_Guia","Ultima_Guia","Fuente"]:
            if col not in df_b.columns:
                df_b[col] = ""

    columnas_finales = ["ID_Pedido","Cliente","Vendedor_Registro","Tipo_Envio","Estado",
                        "Fecha_Entrega","Hora_Registro","Folio_Factura",
                        "Adjuntos_Guia","URLs_Guia","Ultima_Guia","Fuente","id_vendedor","Completados_Limpiado"]
    df_a_hist = df_a_hist[columnas_finales] if not df_a_hist.empty else pd.DataFrame(columns=columnas_finales)
    df_a_op = df_a_op[columnas_finales] if not df_a_op.empty else pd.DataFrame(columns=columnas_finales)
    df_b = df_b[columnas_finales] if not df_b.empty else pd.DataFrame(columns=columnas_finales)

    df = pd.concat([df_a_hist, df_a_op, df_b], ignore_index=True)

    if not df.empty:
        for col_fecha in ["Fecha_Entrega","Hora_Registro"]:
            df[col_fecha] = pd.to_datetime(df[col_fecha], errors="coerce")

        df["Folio_O_ID"] = df["Folio_Factura"].astype(str).str.strip()
        df.loc[df["Folio_O_ID"] == "", "Folio_O_ID"] = df["ID_Pedido"].astype(str).str.strip()

        if df["Fecha_Entrega"].notna().any():
            df = df.sort_values(by="Fecha_Entrega", ascending=False)
        elif df["Hora_Registro"].notna().any():
            df = df.sort_values(by="Hora_Registro", ascending=False)

    return df

with tab5:
    tab5_is_active = default_tab == TAB_INDEX_TAB5
    if tab5_is_active:
        st.session_state["current_tab_index"] = TAB_INDEX_TAB5
    st.header("📦 Pedidos con Guías Subidas desde Almacén y Casos Especiales")

    id_vendedor_sesion = normalize_vendedor_id(st.session_state.get("id_vendedor", ""))

    if st.button("🔄 Actualizar guías"):
        if allow_refresh("guias_last_refresh", cooldown=15):
            st.session_state["guias_refresh_token"] = time.time()

    try:
        df_guias = cargar_datos_guias_unificadas(
            st.session_state.get("guias_refresh_token")
        )
    except Exception as e:
        st.error(f"❌ Error al cargar datos de guías: {e}")
        df_guias = pd.DataFrame()

    if df_guias.empty:
        st.info("No hay pedidos o casos especiales con guías subidas.")
    else:
        st.markdown("### 🔍 Filtros")
        col1_tab5, col2_tab5 = st.columns(2)

        with col1_tab5:
            vendedores = build_vendor_filter_options(
                df_guias["Vendedor_Registro"].dropna().astype(str).tolist(),
            )
            vendedor_filtrado = st.selectbox(
                "Filtrar por Vendedor",
                vendedores,
                index=ensure_selectbox_vendor_default("filtro_vendedor_guias", vendedores),
                key="filtro_vendedor_guias",
                on_change=fijar_tab5_activa
            )

        fecha_inicio_rango = None
        fecha_fin_rango = None
        fecha_filtro_tab5 = None

        with col2_tab5:
            usar_rango_fechas = st.checkbox(
                "🔁 Activar búsqueda por rango de fechas",
                key="filtro_guias_rango_activo",
                on_change=fijar_tab5_activa
            )
            if usar_rango_fechas and st.session_state.get("filtro_guias_7_dias"):
                st.session_state["filtro_guias_7_dias"] = False
            filtrar_7_dias = st.checkbox(
                "Mostrar últimos 7 días",
                key="filtro_guias_7_dias",
                disabled=usar_rango_fechas,
                on_change=fijar_tab5_activa
            )

            if usar_rango_fechas:
                fecha_inicio_rango = st.date_input(
                    "📅 Fecha inicial:",
                    value=st.session_state.get(
                        "filtro_fecha_inicio_guias",
                        datetime.now().date() - timedelta(days=7)
                    ),
                    key="filtro_fecha_inicio_guias",
                    on_change=fijar_tab5_activa
                )
                fecha_fin_rango = st.date_input(
                    "📅 Fecha final:",
                    value=st.session_state.get(
                        "filtro_fecha_fin_guias",
                        datetime.now().date()
                    ),
                    key="filtro_fecha_fin_guias",
                    on_change=fijar_tab5_activa
                )
                if fecha_inicio_rango and fecha_fin_rango and fecha_inicio_rango > fecha_fin_rango:
                    st.warning("⚠️ La fecha inicial no puede ser mayor que la fecha final.")
            else:
                fecha_filtro_tab5 = st.date_input(
                    "📅 Filtrar por Fecha de Registro:",
                    value=st.session_state.get("filtro_fecha_guias", datetime.now().date()),
                    key="filtro_fecha_guias",
                    disabled=filtrar_7_dias,
                    on_change=fijar_tab5_activa
                )

        fecha_col_para_filtrar = None
        if "Hora_Registro" in df_guias.columns and df_guias["Hora_Registro"].notna().any():
            fecha_col_para_filtrar = "Hora_Registro"
        elif "Fecha_Entrega" in df_guias.columns and df_guias["Fecha_Entrega"].notna().any():
            fecha_col_para_filtrar = "Fecha_Entrega"

        if fecha_col_para_filtrar:
            if usar_rango_fechas:
                fecha_inicio_rango = fecha_inicio_rango or st.session_state.get("filtro_fecha_inicio_guias")
                fecha_fin_rango = fecha_fin_rango or st.session_state.get("filtro_fecha_fin_guias")
                if fecha_inicio_rango and fecha_fin_rango and fecha_inicio_rango <= fecha_fin_rango:
                    df_guias = df_guias[df_guias[fecha_col_para_filtrar].dt.date.between(fecha_inicio_rango, fecha_fin_rango)]
            elif filtrar_7_dias:
                hoy = datetime.now().date()
                rango_inicio = hoy - timedelta(days=6)
                df_guias = df_guias[df_guias[fecha_col_para_filtrar].dt.date.between(rango_inicio, hoy)]
            else:
                fecha_filtro_tab5 = fecha_filtro_tab5 or st.session_state.get("filtro_fecha_guias", datetime.now().date())
                df_guias = df_guias[df_guias[fecha_col_para_filtrar].dt.date == fecha_filtro_tab5]

        # Aviso de nuevas guías para pedidos de la sesión/vendedor actual
        if "id_vendedor" not in df_guias.columns:
            df_guias["id_vendedor"] = ""

        df_guias = df_guias.copy()
        df_guias["id_vendedor_norm"] = df_guias["id_vendedor"].apply(normalize_vendedor_id)
        if id_vendedor_sesion:
            df_guias_sesion = df_guias[df_guias["id_vendedor_norm"] == id_vendedor_sesion].copy()
        else:
            df_guias_sesion = pd.DataFrame(columns=df_guias.columns)

        if "Completados_Limpiado" not in df_guias_sesion.columns:
            df_guias_sesion["Completados_Limpiado"] = ""
        df_guias_alertas = df_guias_sesion[
            df_guias_sesion["Completados_Limpiado"].fillna("").astype(str).str.strip() == ""
        ].copy()

        current_guias_map: Dict[str, str] = {}
        if not df_guias_alertas.empty:
            for _, row in df_guias_alertas.iterrows():
                row_key = "::".join([
                    str(row.get("Fuente", "")).strip(),
                    str(row.get("ID_Pedido", "")).strip(),
                    str(row.get("Ultima_Guia", "")).strip(),
                ])
                cliente = str(row.get("Cliente", "")).strip() or "Cliente sin nombre"
                current_guias_map[row_key] = cliente

        guias_signature = "|".join(sorted(current_guias_map.keys()))
        prev_keys_raw = st.session_state.get("tab5_guias_keys", [])
        prev_keys = set(prev_keys_raw if isinstance(prev_keys_raw, list) else [])
        current_keys = set(current_guias_map.keys())
        current_count = int(len(current_keys))
        nuevas_keys = sorted(current_keys - prev_keys)

        if id_vendedor_sesion and prev_keys and nuevas_keys:
            nuevas = len(nuevas_keys)
            clientes_nuevos = [current_guias_map.get(k, "Cliente sin nombre") for k in nuevas_keys]
            clientes_unicos = list(dict.fromkeys(clientes_nuevos))
            detalle_clientes = ", ".join(clientes_unicos[:3])
            if len(clientes_unicos) > 3:
                detalle_clientes = f"{detalle_clientes} y {len(clientes_unicos) - 3} más"

            st.success(
                f"🔔 Se cargaron {nuevas} guía(s) nueva(s) para tus pedidos (ID vendedor: {id_vendedor_sesion})."
            )
            st.info(f"👤 Clientes con nuevas guías: {detalle_clientes}.")
            st.toast(
                f"🔔 Nuevas guías detectadas: {nuevas}",
                icon="📦"
            )
        elif id_vendedor_sesion and current_count == 0:
            st.caption(
                f"Sin nuevas guías detectadas aún para el ID vendedor {id_vendedor_sesion}."
            )

        st.session_state["tab5_guias_signature"] = guias_signature
        st.session_state["tab5_guias_count"] = current_count
        st.session_state["tab5_guias_keys"] = sorted(current_keys)

        if vendedor_filtrado != "Todos":
            df_guias = df_guias[df_guias["Vendedor_Registro"] == vendedor_filtrado]

        columnas_mostrar = ["ID_Pedido","Cliente","Vendedor_Registro","Tipo_Envio","Estado","Fecha_Entrega","Fuente"]
        tabla_guias = df_guias[columnas_mostrar].copy()
        tabla_guias["Fecha_Entrega"] = pd.to_datetime(tabla_guias["Fecha_Entrega"], errors="coerce").dt.strftime("%d/%m/%y")
        st.dataframe(tabla_guias, use_container_width=True, hide_index=True)

        st.markdown("### 📥 Selecciona un Pedido para Ver la Última Guía Subida")

        df_guias["display_label"] = df_guias.apply(
            lambda row: f"📄 {row['Folio_O_ID']} – {row['Cliente']} – {row['Vendedor_Registro']} ({row['Tipo_Envio']}) · {row['Fuente']}",
            axis=1
        )

        pedido_seleccionado = st.selectbox(
            "📦 Pedido/Caso con Guía",
            options=df_guias["display_label"].tolist(),
            key="select_pedido_con_guia"
        )

        if pedido_seleccionado:
            pedido_row = df_guias[df_guias["display_label"] == pedido_seleccionado].iloc[0]
            ultima_guia = str(pedido_row["Ultima_Guia"]).strip()
            fuente = ""
            if "Fuente" in pedido_row:
                fuente = str(pedido_row["Fuente"]).strip()
            guias_unicas = partir_urls(pedido_row.get("URLs_Guia", ""))

            if fuente == "casos_especiales":
                st.markdown("### 📎 Guías Subidas")
                if guias_unicas:
                    for guia_url in guias_unicas:
                        url_encoded = quote(guia_url, safe=':/')
                        render_attachment_link(url_encoded, _infer_display_name(guia_url), bullet=False)
                elif ultima_guia:
                    url_encoded = quote(ultima_guia, safe=':/')
                    render_attachment_link(url_encoded, _infer_display_name(ultima_guia), bullet=False)
                else:
                    st.warning("⚠️ No se encontró una URL válida para la guía.")
            else:
                st.markdown("### 📎 Última Guía Subida")
                if ultima_guia:
                    url_encoded = quote(ultima_guia, safe=':/')
                    nombre = ultima_guia.split("/")[-1]
                    render_attachment_link(url_encoded, f"📄 {nombre}")
                else:
                    st.warning("⚠️ No se encontró una URL válida para la guía.")

# --- TAB 6: PEDIDOS NO ENTREGADOS ---
with tab6:
    tab6_is_active = default_tab == TAB_INDEX_TAB6
    if tab6_is_active:
        st.session_state["current_tab_index"] = TAB_INDEX_TAB6
    st.header("⏳ Pedidos No Entregados")

    if st.button("🔄 Actualizar listado", key="refresh_no_entregados"):
        if allow_refresh("no_entregados_last_refresh"):
            cargar_pedidos.clear()
            st.toast("🔄 Datos de pedidos recargados")
            st.rerun()

    if tab6_is_active:
        try:
            df_pedidos_no_entregados = cargar_pedidos()
        except Exception as e:
            st.error(f"❌ Error al cargar los pedidos: {e}")
            df_pedidos_no_entregados = pd.DataFrame()
    else:
        render_lazy_tab_placeholder(
            5,
            "tab6_lazy",
            LAZY_TAB_MESSAGE,
        )
        df_pedidos_no_entregados = pd.DataFrame()

    if df_pedidos_no_entregados.empty:
        st.info("No se encontraron pedidos para mostrar.")
    elif "Estado_Entrega" not in df_pedidos_no_entregados.columns:
        st.warning("La columna 'Estado_Entrega' no se encontró en los datos de pedidos.")
    else:
        df_pedidos_no_entregados = df_pedidos_no_entregados.copy()
        df_pedidos_no_entregados["Estado_Entrega"] = (
            df_pedidos_no_entregados["Estado_Entrega"].astype(str).str.strip()
        )
        filtro_no_entregados = df_pedidos_no_entregados["Estado_Entrega"] == "⏳ No Entregado"
        df_pedidos_no_entregados = df_pedidos_no_entregados[filtro_no_entregados].reset_index(drop=True)

        if df_pedidos_no_entregados.empty:
            st.success("🎉 No hay pedidos marcados como '⏳ No Entregado' en este momento.")
        else:
            columnas_tabla = [
                col
                for col in [
                    "ID_Pedido",
                    "Cliente",
                    "Tipo_Envio",
                    "Estado",
                    "Fecha_Entrega",
                    "Turno",
                    "Comprobante_Confirmado",
                ]
                if col in df_pedidos_no_entregados.columns
            ]

            if columnas_tabla:
                tabla_visual = df_pedidos_no_entregados[columnas_tabla].copy()
                if "Fecha_Entrega" in tabla_visual.columns:
                    tabla_visual["Fecha_Entrega"] = pd.to_datetime(
                        tabla_visual["Fecha_Entrega"], errors="coerce"
                    ).dt.strftime("%Y-%m-%d")
                st.dataframe(tabla_visual, use_container_width=True, hide_index=True)

            df_pedidos_no_entregados["display_label"] = df_pedidos_no_entregados.apply(
                lambda row: " - ".join(
                    [
                        (
                            str(row.get("Folio_Factura", "")).strip()
                            or str(row.get("ID_Pedido", "")).strip()
                            or "Sin folio"
                        ),
                        str(row.get("Cliente", "")).strip() or "Sin Cliente",
                        str(row.get("Tipo_Envio", "")).strip() or "Sin Tipo",
                    ]
                ),
                axis=1,
            )

            pedido_seleccionado_no_entregado = st.selectbox(
                "📋 Selecciona un pedido para actualizar la entrega",
                options=df_pedidos_no_entregados["display_label"].tolist(),
                key="select_pedido_no_entregado",
            )

            if pedido_seleccionado_no_entregado:
                pedido_fila = df_pedidos_no_entregados[
                    df_pedidos_no_entregados["display_label"] == pedido_seleccionado_no_entregado
                ].iloc[0]

                pedido_id = str(pedido_fila.get("ID_Pedido", "")).strip()
                pedido_folio = str(pedido_fila.get("Folio_Factura", "")).strip()
                folio_display = pedido_folio or pedido_id
                tipo_envio = str(pedido_fila.get("Tipo_Envio", "")).strip()
                fecha_actual = pd.to_datetime(pedido_fila.get("Fecha_Entrega"), errors="coerce")
                turno_actual = str(pedido_fila.get("Turno", "")).strip()

                st.markdown(
                    f"**Cliente:** {pedido_fila.get('Cliente', 'N/D')}  |  **Folio:** {folio_display or 'N/D'}"
                )
                st.markdown(
                    f"**Tipo de envío:** {tipo_envio or 'N/D'}  |  **Estado actual de entrega:** {pedido_fila.get('Estado_Entrega', 'N/D')}"
                )
                st.markdown(
                    f"**Fecha de entrega registrada:** {fecha_actual.date() if pd.notna(fecha_actual) else 'Sin fecha'}  |  **Turno registrado:** {turno_actual or 'Sin turno'}"
                )

                if tipo_envio != "📍 Pedido Local":
                    st.info("Solo se pueden actualizar fecha y turno para pedidos con tipo de envío '📍 Pedido Local'.")
                elif not pedido_id:
                    st.warning("El pedido seleccionado no tiene un 'ID_Pedido' válido para actualizar en Google Sheets.")
                else:
                    turno_options = [
                        "",
                        "🌙 Local Tarde",
                        "☀️ Local Mañana",
                        "📦 Pasa a Bodega",
                        "🌵 Saltillo",
                    ]
                    turno_normalization_map = {
                        "🌙 local tarde": "🌙 Local Tarde",
                        "local tarde": "🌙 Local Tarde",
                        "tarde": "🌙 Local Tarde",
                        "☀️ local mañana": "☀️ Local Mañana",
                        "local mañana": "☀️ Local Mañana",
                        "mañana": "☀️ Local Mañana",
                        "📦 pasa a bodega": "📦 Pasa a Bodega",
                        "pasa a bodega": "📦 Pasa a Bodega",
                        "en bodega": "📦 Pasa a Bodega",
                        "bodega": "📦 Pasa a Bodega",
                        "🌵 saltillo": "🌵 Saltillo",
                        "saltillo": "🌵 Saltillo",
                    }
                    turno_index = 0
                    turno_actual_key = turno_actual.lower()
                    if turno_actual_key == "nan":
                        turno_actual_key = ""
                    turno_actual_estandar = turno_normalization_map.get(
                        turno_actual_key, turno_actual
                    )
                    if turno_actual_estandar in turno_options:
                        turno_index = turno_options.index(turno_actual_estandar)

                    fecha_defecto = fecha_actual.date() if pd.notna(fecha_actual) else date.today()

                    with st.form(key=f"form_actualizar_entrega_{pedido_id}"):
                        nueva_fecha_entrega = st.date_input(
                            "Nueva fecha de entrega",
                            value=fecha_defecto,
                            key=f"fecha_no_entregado_{pedido_id}",
                        )
                        nuevo_turno = st.selectbox(
                            "Selecciona el turno",
                            turno_options,
                            index=turno_index,
                            format_func=lambda x: "Selecciona un turno" if x == "" else x,
                            key=f"turno_no_entregado_{pedido_id}",
                        )
                        submitted = st.form_submit_button("💾 Guardar cambios")

                    if submitted:
                        if nuevo_turno == "":
                            st.warning("Selecciona un turno para continuar.")
                        else:
                            worksheet = get_worksheet()
                            if worksheet is None:
                                st.error("❌ No se pudo acceder a la hoja de Google Sheets para actualizar el pedido.")
                            else:
                                headers = worksheet.row_values(1)
                                try:
                                    df_completo = cargar_pedidos()
                                except Exception as e:
                                    st.error(f"❌ No se pudieron recargar los pedidos desde Google Sheets: {e}")
                                    df_completo = pd.DataFrame()

                                if df_completo.empty or "ID_Pedido" not in df_completo.columns:
                                    st.error("❌ No se encontró la columna 'ID_Pedido' en los datos originales.")
                                elif pedido_id not in df_completo["ID_Pedido"].astype(str).str.strip().tolist():
                                    st.error("❌ No se encontró el pedido seleccionado en los datos originales.")
                                else:
                                    fila_filtrada = df_completo[
                                        df_completo["ID_Pedido"].astype(str).str.strip() == pedido_id
                                    ]
                                    if fila_filtrada.empty:
                                        st.error("❌ No se encontró el pedido seleccionado en la hoja de cálculo.")
                                    else:
                                        fila_original = fila_filtrada.iloc[0]
                                        gsheet_row_index = fila_filtrada.index[0] + 2

                                        updates = []
                                        missing_columns = []

                                        def agregar_actualizacion(columna: str, valor: str) -> None:
                                            if columna not in headers:
                                                missing_columns.append(columna)
                                                return
                                            updates.append(
                                                {
                                                    "range": rowcol_to_a1(
                                                        gsheet_row_index,
                                                        headers.index(columna) + 1,
                                                    ),
                                                    "values": [[valor]],
                                                }
                                            )

                                        fecha_formateada = (
                                            nueva_fecha_entrega.strftime("%Y-%m-%d")
                                            if isinstance(nueva_fecha_entrega, date)
                                            else ""
                                        )

                                        if fecha_formateada:
                                            fecha_existente = pd.to_datetime(
                                                fila_original.get("Fecha_Entrega"), errors="coerce"
                                            )
                                            fecha_existente_date = (
                                                fecha_existente.date() if pd.notna(fecha_existente) else None
                                            )
                                            if fecha_existente_date != nueva_fecha_entrega:
                                                agregar_actualizacion("Fecha_Entrega", fecha_formateada)

                                        turno_actual_estandar = turno_normalization_map.get(
                                            turno_actual.lower() if turno_actual else "",
                                            turno_actual,
                                        )
                                        if turno_actual_estandar == "nan":
                                            turno_actual_estandar = ""
                                        if (
                                            nuevo_turno
                                            and turno_actual_estandar != nuevo_turno
                                        ):
                                            agregar_actualizacion("Turno", nuevo_turno)

                                        comprobante_actual = str(
                                            fila_original.get("Comprobante_Confirmado", "")
                                        ).strip()
                                        comprobante_normalizado = unicodedata.normalize(
                                            "NFKD", comprobante_actual
                                        ).encode("ASCII", "ignore").decode("utf-8").lower()
                                        if comprobante_normalizado == "si":
                                            agregar_actualizacion("Comprobante_Confirmado", "")

                                        if missing_columns:
                                            st.warning(
                                                "No se pudieron actualizar algunas columnas porque no existen en la hoja: "
                                                + ", ".join(missing_columns)
                                            )

                                        if not updates:
                                            st.info("No hay cambios para guardar.")
                                        else:
                                            try:
                                                safe_batch_update(worksheet, updates)
                                                cargar_pedidos.clear()
                                                st.success("✅ Pedido actualizado correctamente.")
                                                st.rerun()
                                            except Exception as e:
                                                st.error(f"❌ Error al actualizar el pedido: {e}")

# --- TAB 7: DOWNLOAD DATA ---
with tab7:
    tab7_is_active = default_tab == TAB_INDEX_TAB7
    if tab7_is_active:
        st.session_state["current_tab_index"] = TAB_INDEX_TAB7
    st.header("⬇️ Descargar Datos de Pedidos")

    @st.cache_data(ttl=60)
    def cargar_todos_los_pedidos():
        worksheet = get_worksheet()
        headers = worksheet.row_values(1)
        if headers:
            df = pd.DataFrame(worksheet.get_all_records())
            if "Adjuntos_Guia" not in df.columns:
                df["Adjuntos_Guia"] = ""
            return df, headers
        return pd.DataFrame(), []

    df_all_pedidos = pd.DataFrame()
    headers = []

    if tab7_is_active:
        try:
            df_all_pedidos, headers = cargar_todos_los_pedidos()
    
            if "Adjuntos_Guia" not in df_all_pedidos.columns:
                df_all_pedidos["Adjuntos_Guia"] = ""
    
            # 🧹 AÑADIDO: Filtrar filas donde 'Folio_Factura' y 'ID_Pedido' son ambos vacíos
            df_all_pedidos = df_all_pedidos.dropna(subset=['Folio_Factura', 'ID_Pedido'], how='all')
    
            # 🧹 Eliminar registros vacíos o inválidos con ID_Pedido en blanco, 'nan', 'N/A'
            df_all_pedidos = df_all_pedidos[
                df_all_pedidos['ID_Pedido'].astype(str).str.strip().ne('') &
                df_all_pedidos['ID_Pedido'].astype(str).str.lower().ne('n/a') &
                df_all_pedidos['ID_Pedido'].astype(str).str.lower().ne('nan')
            ]
    
            if 'Fecha_Entrega' in df_all_pedidos.columns:
                df_all_pedidos['Fecha_Entrega'] = pd.to_datetime(df_all_pedidos['Fecha_Entrega'], errors='coerce')
    
            if 'Vendedor_Registro' in df_all_pedidos.columns:
                df_all_pedidos['Vendedor_Registro'] = df_all_pedidos['Vendedor_Registro'].apply(
                    lambda x: x if x in VENDEDORES_LIST else 'Otro/Desconocido' if pd.notna(x) and str(x).strip() != '' else 'N/A'
                ).astype(str)
            else:
                st.warning("La columna 'Vendedor_Registro' no se encontró en el Google Sheet para el filtrado. Asegúrate de que exista y esté correctamente nombrada.")
    
            if 'Folio_Factura' in df_all_pedidos.columns:
                df_all_pedidos['Folio_Factura'] = df_all_pedidos['Folio_Factura'].astype(str).replace('nan', '')
            else:
                st.warning("La columna 'Folio_Factura' no se encontró en el Google Sheet. No se podrá mostrar en la vista previa.")
        except Exception as e:
            st.error(f"❌ Error al cargar datos para descarga: {e}")
            st.info("Asegúrate de que la primera fila de tu Google Sheet contiene los encabezados esperados y que la API de Google Sheets está habilitada.")
            st.stop()
    else:
        render_lazy_tab_placeholder(
            6,
            "tab7_lazy",
            LAZY_TAB_MESSAGE,
        )

    if not tab7_is_active:
        pass
    elif df_all_pedidos.empty:
        st.info("No hay datos de pedidos para descargar.")
    else:
        st.markdown("---")
        st.subheader("Opciones de Filtro")

        time_filter = st.radio(
            "Selecciona un rango de tiempo:",
            ("Todos los datos", "Últimas 24 horas", "Últimos 7 días", "Últimos 30 días"),
            key="download_time_filter"
        )

        filtered_df_download = df_all_pedidos.copy()

        if time_filter != "Todos los datos" and 'Fecha_Entrega' in filtered_df_download.columns:
            current_time = datetime.now()
            # MODIFICATION 3: Convert Fecha_Entrega to date only for comparison
            filtered_df_download['Fecha_Solo_Fecha'] = filtered_df_download['Fecha_Entrega'].dt.date

            if time_filter == "Últimas 24 horas":
                start_datetime = current_time - timedelta(hours=24)
                filtered_df_download = filtered_df_download[filtered_df_download['Fecha_Entrega'] >= start_datetime]
            else:
                if time_filter == "Últimos 7 días":
                    start_date = current_time.date() - timedelta(days=7)
                elif time_filter == "Últimos 30 días":
                    start_date = current_time.date() - timedelta(days=30)

                filtered_df_download = filtered_df_download[filtered_df_download['Fecha_Solo_Fecha'] >= start_date]

            filtered_df_download = filtered_df_download.drop(columns=['Fecha_Solo_Fecha'])


        if 'Vendedor_Registro' in df_all_pedidos.columns:
            unique_vendedores_en_df = set(filtered_df_download['Vendedor_Registro'].unique())

            options_for_selectbox = ["Todos"]
            for vendedor_nombre in VENDEDORES_LIST:
                if vendedor_nombre in unique_vendedores_en_df:
                    options_for_selectbox.append(vendedor_nombre)

            if 'Otro/Desconocido' in unique_vendedores_en_df and 'Otro/Desconocido' not in options_for_selectbox:
                options_for_selectbox.append('Otro/Desconocido')

            if 'N/A' in unique_vendedores_en_df and 'N/A' not in options_for_selectbox:
                options_for_selectbox.append('N/A')

            selected_vendedor = st.selectbox(
                "Filtrar por Vendedor:",
                options=options_for_selectbox,
                index=ensure_selectbox_vendor_default("download_vendedor_filter_tab6_final", options_for_selectbox),
                key="download_vendedor_filter_tab6_final"
            )

            if selected_vendedor != "Todos":
                filtered_df_download = filtered_df_download[filtered_df_download['Vendedor_Registro'] == selected_vendedor]
        else:
            st.warning("La columna 'Vendedor_Registro' no está disponible en los datos cargados para aplicar este filtro. Por favor, asegúrate de que el nombre de la columna en tu Google Sheet sea 'Vendedor_Registro'.")

        if 'Tipo_Envio' in filtered_df_download.columns:
            unique_tipos_envio_download = [
                "Todos",
                "📍 Pedido Local",
                "🚚 Pedido Foráneo",
                "🎓 Cursos y Eventos",
                "🔁 Devolución",
                "🛠 Garantía",
            ]
            selected_tipo_envio_download = st.selectbox(
                "Filtrar por Tipo de Envío:",
                options=unique_tipos_envio_download,
                key="download_tipo_envio_filter"
            )
            if selected_tipo_envio_download != "Todos":
                filtered_df_download = filtered_df_download[filtered_df_download['Tipo_Envio'] == selected_tipo_envio_download]
        else:
            st.warning("La columna 'Tipo_Envio' no se encontró para aplicar el filtro de tipo de envío.")


        if 'Estado' in filtered_df_download.columns:
            unique_estados = ["Todos"] + list(filtered_df_download['Estado'].dropna().unique())
            selected_estado = st.selectbox("Filtrar por Estado:", unique_estados, key="download_estado_filter_tab6")
            if selected_estado != "Todos":
                filtered_df_download = filtered_df_download[filtered_df_download['Estado'] == selected_estado]

        st.markdown("---")
        st.subheader("Vista Previa de Datos a Descargar")

        # MODIFICATION 3: Format 'Fecha_Entrega' for display
        columnas_excluidas_preview = [
            "ID_Pedido", "Adjuntos", "Adjuntos_Surtido", "Adjuntos_Guia",
            "Completados_Limpiado", "Fecha_Pago_Comprobante",
            "Terminal", "Banco_Destino_Pago", "Forma_Pago_Comprobante",
            "Monto_Comprobante", "Referencia_Comprobante"
        ]
        columnas_preview = [col for col in filtered_df_download.columns if col not in columnas_excluidas_preview]
        display_df = filtered_df_download[columnas_preview].copy()
                
        if 'Fecha_Entrega' in display_df.columns:
            display_df['Fecha_Entrega'] = display_df['Fecha_Entrega'].dt.strftime('%Y-%m-%d')

        st.dataframe(display_df, use_container_width=True, hide_index=True)

        if not filtered_df_download.empty:
            output = BytesIO()
            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                # Crear copia para exportar solo columnas seguras
                columnas_excluidas = [
                    "ID_Pedido", "Adjuntos", "Adjuntos_Surtido", "Adjuntos_Guia",
                    "Completados_Limpiado", "Fecha_Pago_Comprobante",
                    "Terminal", "Banco_Destino_Pago", "Forma_Pago_Comprobante",
                    "Monto_Comprobante", "Referencia_Comprobante"
                ]
                columnas_finales = [col for col in filtered_df_download.columns if col not in columnas_excluidas]

                excel_df = filtered_df_download[columnas_finales].copy()

                # Convertir fechas a texto legible
                for col in excel_df.columns:
                    if "fecha" in col.lower() or "Fecha" in col:
                        excel_df[col] = pd.to_datetime(excel_df[col], errors='coerce').dt.strftime('%Y-%m-%d')


                # Asegúrate de que las fechas estén en formato string
                for fecha_col in ['Fecha_Entrega', 'Fecha_Pago_Comprobante']:
                    if fecha_col in excel_df.columns:
                        excel_df[fecha_col] = pd.to_datetime(excel_df[fecha_col], errors='coerce').dt.strftime('%Y-%m-%d')

                excel_df.to_excel(writer, index=False, sheet_name='Pedidos_Filtrados')

            processed_data = output.getvalue()

            st.download_button(
                label="📥 Descargar Excel Filtrado",
                data=processed_data,
                file_name=f"pedidos_filtrados_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                help="Haz clic para descargar los datos de la tabla mostrada arriba en formato Excel."
            )
        else:
            st.info("No hay datos que coincidan con los filtros seleccionados para descargar.")
# --- Helpers exclusivos para Tab 8 (Buscar Pedido) ---
PEDIDOS_SHEETS = ("datos_pedidos", "data_pedidos")
PEDIDOS_COLUMNAS_MINIMAS = [
    "ID_Pedido", "Hora_Registro", "Cliente", "Estado", "Vendedor_Registro", "Folio_Factura",
    "Comentario", "Comentarios", "Modificacion_Surtido", "Adjuntos_Surtido", "Adjuntos_Guia",
    "Adjuntos", "Direccion_Guia_Retorno", "Nota_Venta", "Tiene_Nota_Venta", "Motivo_NotaVenta",
    "Refacturacion_Tipo", "Refacturacion_Subtipo", "Folio_Factura_Refacturada", "fecha_modificacion", "Fecha_Modificacion"
]
INLINE_EXT = (".pdf", ".jpg", ".jpeg", ".png", ".webp")


def extract_s3_key_busqueda(url_or_key: str) -> str:
    if not isinstance(url_or_key, str):
        return url_or_key
    parsed = urlparse(url_or_key)
    if parsed.scheme and parsed.netloc:
        return unquote(parsed.path.lstrip("/"))
    return url_or_key


def get_s3_file_download_url_busqueda(s3_client_param, object_key_or_url, expires_in=604800):
    if not s3_client_param or not S3_BUCKET_NAME:
        st.error("❌ Configuración de S3 incompleta. Verifica el cliente y el nombre del bucket.")
        return "#"
    try:
        clean_key = extract_s3_key_busqueda(object_key_or_url)
        params = {"Bucket": S3_BUCKET_NAME, "Key": clean_key}
        if isinstance(clean_key, str):
            lower_key = clean_key.lower()
            if lower_key.endswith(INLINE_EXT):
                filename = (clean_key.split("/")[-1] or "archivo").replace('"', "")
                params["ResponseContentDisposition"] = f'inline; filename="{filename}"'
                if lower_key.endswith(".pdf"):
                    params["ResponseContentType"] = "application/pdf"
                elif lower_key.endswith((".jpg", ".jpeg")):
                    params["ResponseContentType"] = "image/jpeg"
                elif lower_key.endswith(".png"):
                    params["ResponseContentType"] = "image/png"
                elif lower_key.endswith(".webp"):
                    params["ResponseContentType"] = "image/webp"
        return s3_client_param.generate_presigned_url("get_object", Params=params, ExpiresIn=expires_in)
    except Exception as e:
        st.error(f"❌ Error al generar URL prefirmada: {e}")
        return "#"

@st.cache_data(ttl=1800)
def get_s3_file_download_url_busqueda_cached(object_key_or_url, expires_in=604800):
    return get_s3_file_download_url_busqueda(s3_client, object_key_or_url, expires_in=expires_in)


def resolver_nombre_y_enlace_busqueda(valor, etiqueta_fallback):
    valor = str(valor).strip()
    if not valor:
        return None, None

    parsed = urlparse(valor)
    nombre_crudo = extract_s3_key_busqueda(valor)
    nombre = nombre_crudo.split("/")[-1] if nombre_crudo else ""
    if not nombre:
        nombre = etiqueta_fallback

    if parsed.scheme and parsed.netloc:
        enlace = valor
    else:
        enlace = get_s3_file_download_url_busqueda_cached(valor)
        if not enlace or enlace == "#":
            enlace = valor

    return nombre, enlace


def normalizar_folio(texto):
    if texto is None:
        return ""
    limpio = normalizar(str(texto).strip())
    limpio_sin_espacios = re.sub(r"\s+", "", limpio)
    return limpio_sin_espacios.upper()


def obtener_fecha_modificacion(row):
    return str(row.get("Fecha_Modificacion") or row.get("fecha_modificacion") or "").strip()


def preparar_resultado_caso_busqueda(row):
    return {
        "__source": "casos",
        "ID_Pedido": str(row.get("ID_Pedido", "")).strip(),
        "Cliente": row.get("Cliente", ""),
        "Vendedor": row.get("Vendedor_Registro", ""),
        "Folio": row.get("Folio_Factura", ""),
        "Folio_Factura_Error": row.get("Folio_Factura_Error", ""),
        "Hora_Registro": row.get("Hora_Registro", ""),
        "Tipo_Envio": row.get("Tipo_Envio", ""),
        "Estado": row.get("Estado", ""),
        "Estado_Caso": row.get("Estado_Caso", ""),
        "Resultado_Esperado": row.get("Resultado_Esperado", ""),
        "Material_Devuelto": row.get("Material_Devuelto", ""),
        "Monto_Devuelto": row.get("Monto_Devuelto", ""),
        "Motivo_Detallado": row.get("Motivo_Detallado", ""),
        "Area_Responsable": row.get("Area_Responsable", ""),
        "Nombre_Responsable": row.get("Nombre_Responsable", ""),
        "Numero_Cliente_RFC": row.get("Numero_Cliente_RFC", ""),
        "Tipo_Envio_Original": row.get("Tipo_Envio_Original", ""),
        "Fecha_Entrega": row.get("Fecha_Entrega", ""),
        "Fecha_Recepcion_Devolucion": row.get("Fecha_Recepcion_Devolucion", ""),
        "Estado_Recepcion": row.get("Estado_Recepcion", ""),
        "Nota_Credito_URL": row.get("Nota_Credito_URL", ""),
        "Documento_Adicional_URL": row.get("Documento_Adicional_URL", ""),
        "Seguimiento": row.get("Seguimiento", ""),
        "Comentarios_Admin_Devolucion": row.get("Comentarios_Admin_Devolucion", ""),
        "Turno": row.get("Turno", ""),
        "Hora_Proceso": row.get("Hora_Proceso", ""),
        "Numero_Serie": row.get("Numero_Serie", ""),
        "Fecha_Compra": row.get("Fecha_Compra", ""),
        "Comentario": str(row.get("Comentario", "")).strip(),
        "Comentarios": str(row.get("Comentarios", "")).strip(),
        "Direccion_Guia_Retorno": str(row.get("Direccion_Guia_Retorno", "")).strip(),
        "Nota_Venta": str(row.get("Nota_Venta", "")).strip(),
        "Tiene_Nota_Venta": str(row.get("Tiene_Nota_Venta", "")).strip(),
        "Motivo_NotaVenta": str(row.get("Motivo_NotaVenta", "")).strip(),
        "Modificacion_Surtido": str(row.get("Modificacion_Surtido", "")).strip(),
        "Fecha_Modificacion_Surtido": obtener_fecha_modificacion(row),
        "Adjuntos_Surtido_urls": partir_urls(row.get("Adjuntos_Surtido", "")),
        "Refacturacion_Tipo": str(row.get("Refacturacion_Tipo", "")).strip(),
        "Refacturacion_Subtipo": str(row.get("Refacturacion_Subtipo", "")).strip(),
        "Folio_Factura_Refacturada": str(row.get("Folio_Factura_Refacturada", "")).strip(),
        "Adjuntos_urls": partir_urls(row.get("Adjuntos", "")),
        "Guias_urls": partir_urls(row.get("Hoja_Ruta_Mensajero", "")),
    }


def render_caso_especial_busqueda(res):
    titulo = f"🧾 Caso Especial – {res.get('Tipo_Envio', '') or 'N/A'}"
    st.markdown(f"### {titulo}")

    tipo_envio_val = str(res.get("Tipo_Envio", ""))
    is_devolucion = tipo_envio_val.strip() == "🔁 Devolución"
    is_garantia = "garant" in tipo_envio_val.lower()
    if is_devolucion:
        folio_nuevo = res.get("Folio", "") or "N/A"
        folio_error = res.get("Folio_Factura_Error", "") or "N/A"
        st.markdown(
            f"📄 **Folio Nuevo:** `{folio_nuevo}`  |  📄 **Folio Error:** `{folio_error}`  |  "
            f"🧑‍💼 **Vendedor:** `{res.get('Vendedor', '') or 'N/A'}`  |  🕒 **Hora:** `{res.get('Hora_Registro', '') or 'N/A'}`"
        )
    else:
        st.markdown(
            f"📄 **Folio:** `{res.get('Folio', '') or 'N/A'}`  |  "
            f"🧑‍💼 **Vendedor:** `{res.get('Vendedor', '') or 'N/A'}`  |  🕒 **Hora:** `{res.get('Hora_Registro', '') or 'N/A'}`"
        )

    st.markdown(f"**👤 Cliente:** {res.get('Cliente', 'N/A')}  |  **RFC:** {res.get('Numero_Cliente_RFC', '') or 'N/A'}")
    st.markdown(
        f"**Estado:** {res.get('Estado', '') or 'N/A'}  |  **Estado del Caso:** {res.get('Estado_Caso', '') or 'N/A'}  |  **Turno:** {res.get('Turno', '') or 'N/A'}"
    )
    if is_garantia:
        st.markdown(
            f"**🔢 Número de Serie:** {res.get('Numero_Serie', '') or 'N/A'}  |  **📅 Fecha de Compra:** {res.get('Fecha_Compra', '') or 'N/A'}"
        )

    comentario_txt = str(res.get("Comentario", "") or res.get("Comentarios", "")).strip()
    if comentario_txt:
        st.markdown("#### 📝 Comentarios del pedido")
        st.info(comentario_txt)

    direccion_retorno = str(res.get("Direccion_Guia_Retorno", "")).strip()
    if direccion_retorno:
        st.markdown("#### 📍 Dirección para guía de retorno")
        st.info(direccion_retorno)

    nota_venta_valor = str(res.get("Nota_Venta", "")).strip()
    tiene_nota_venta = str(res.get("Tiene_Nota_Venta", "")).strip()
    motivo_nota_venta = str(res.get("Motivo_NotaVenta", "")).strip()
    if nota_venta_valor or tiene_nota_venta or motivo_nota_venta:
        st.markdown("#### 🧾 Nota de venta")
        estado_texto = tiene_nota_venta or ("Sí" if nota_venta_valor else "No")
        st.markdown(f"- **¿Tiene nota de venta?:** {estado_texto}")
        if nota_venta_valor:
            st.markdown(f"- **Detalle:** {nota_venta_valor}")
        if motivo_nota_venta:
            st.markdown(f"- **Motivo:** {motivo_nota_venta}")

    ref_t = res.get("Refacturacion_Tipo", "")
    ref_st = res.get("Refacturacion_Subtipo", "")
    ref_f = res.get("Folio_Factura_Refacturada", "")
    if any([ref_t, ref_st, ref_f]):
        st.markdown("**♻️ Refacturación:**")
        st.markdown(f"- **Tipo:** {ref_t or 'N/A'}")
        st.markdown(f"- **Subtipo:** {ref_st or 'N/A'}")
        st.markdown(f"- **Folio refacturado:** {ref_f or 'N/A'}")

    if str(res.get("Resultado_Esperado", "")).strip():
        st.markdown(f"**🎯 Resultado Esperado:** {res.get('Resultado_Esperado', '')}")
    if str(res.get("Motivo_Detallado", "")).strip():
        st.markdown("**📝 Motivo / Descripción:**")
        st.info(str(res.get("Motivo_Detallado", "")).strip())
    if str(res.get("Material_Devuelto", "")).strip():
        st.markdown("**📦 Piezas / Material:**")
        st.info(str(res.get("Material_Devuelto", "")).strip())
    if str(res.get("Monto_Devuelto", "")).strip():
        st.markdown(f"**💵 Monto (dev./estimado):** {res.get('Monto_Devuelto', '')}")

    st.markdown(
        f"**🏢 Área Responsable:** {res.get('Area_Responsable', '') or 'N/A'}  |  **👥 Responsable del Error:** {res.get('Nombre_Responsable', '') or 'N/A'}"
    )
    st.markdown(
        f"**📅 Fecha Entrega/Cierre (si aplica):** {res.get('Fecha_Entrega', '') or 'N/A'}  |  "
        f"**📅 Recepción:** {res.get('Fecha_Recepcion_Devolucion', '') or 'N/A'}  |  "
        f"**📦 Recepción:** {res.get('Estado_Recepcion', '') or 'N/A'}"
    )
    st.markdown(
        f"**🧾 Nota de Crédito:** {res.get('Nota_Credito_URL', '') or 'N/A'}  |  "
        f"**📂 Documento Adicional:** {res.get('Documento_Adicional_URL', '') or 'N/A'}"
    )
    if str(res.get("Comentarios_Admin_Devolucion", "")).strip():
        st.markdown("**🗒️ Comentario Administrativo:**")
        st.info(str(res.get("Comentarios_Admin_Devolucion", "")).strip())

    seguimiento_txt = str(res.get("Seguimiento", ""))
    if (is_devolucion or is_garantia) and seguimiento_txt.strip():
        st.markdown("**📌 Seguimiento:**")
        st.info(seguimiento_txt.strip())

    mod_txt = res.get("Modificacion_Surtido", "") or ""
    mod_fecha = res.get("Fecha_Modificacion_Surtido", "") or ""
    mod_urls = res.get("Adjuntos_Surtido_urls", []) or []
    if mod_txt or mod_urls:
        st.markdown("#### 🛠 Modificación de surtido")
        if mod_fecha:
            st.caption(f"📅 Fecha de modificación: {mod_fecha}")
        if mod_txt:
            st.info(mod_txt)
        if mod_urls:
            st.markdown("**Archivos de modificación:**")
            for u in mod_urls:
                nombre = extract_s3_key_busqueda(u).split("/")[-1]
                tmp = get_s3_file_download_url_busqueda_cached(u)
                st.markdown(f'- <a href="{tmp}" target="_blank">{nombre}</a>', unsafe_allow_html=True)

    with st.expander("📎 Archivos (Adjuntos y Guía)", expanded=False):
        adj = res.get("Adjuntos_urls", []) or []
        guias = res.get("Guias_urls", []) or []
        if adj:
            st.markdown("**Adjuntos:**")
            for u in adj:
                nombre = extract_s3_key_busqueda(u).split("/")[-1]
                tmp = get_s3_file_download_url_busqueda_cached(u)
                st.markdown(f'- <a href="{tmp}" target="_blank">{nombre}</a>', unsafe_allow_html=True)
        if guias:
            st.markdown("**Guías:**")
            for idx, u in enumerate(guias, start=1):
                if not str(u).strip():
                    continue
                nombre = extract_s3_key_busqueda(u).split("/")[-1]
                if not nombre:
                    nombre = f"Guía #{idx}"
                tmp = get_s3_file_download_url_busqueda_cached(u)
                st.markdown(f'- <a href="{tmp}" target="_blank">{nombre}</a>', unsafe_allow_html=True)
        if not adj and not guias:
            st.info("Sin archivos registrados en la hoja.")

    st.markdown("---")


def _leer_registros_hoja_busqueda(nombre_hoja: str, retries: int = 5, base_delay: float = 0.8):
    """Lee una hoja con reintentos para errores transitorios/cuota (incluye 409)."""
    last_error = None
    for attempt in range(retries):
        try:
            sheet = g_spread_client.open_by_key(GOOGLE_SHEET_ID).worksheet(nombre_hoja)
            return sheet.get_all_records()
        except APIError as e:
            last_error = e
            status = getattr(getattr(e, "response", None), "status_code", None)
            transient = status in (409, 429, 500, 503) or "RESOURCE_EXHAUSTED" in str(e)
            if transient and attempt < retries - 1:
                time.sleep(base_delay * (2 ** attempt))
                continue
            raise
        except Exception as e:
            last_error = e
            if attempt < retries - 1:
                time.sleep(base_delay * (2 ** attempt))
                continue
            raise
    if last_error:
        raise last_error
    return []


@st.cache_data(ttl=300)
def cargar_hoja_pedidos_busqueda(nombre_hoja):
    data = _leer_registros_hoja_busqueda(nombre_hoja)
    df = pd.DataFrame(data)
    for c in PEDIDOS_COLUMNAS_MINIMAS:
        if c not in df.columns:
            df[c] = ""
    return df


@st.cache_data(ttl=300)
def cargar_pedidos_busqueda():
    pedidos_frames = [cargar_hoja_pedidos_busqueda(nombre_hoja) for nombre_hoja in PEDIDOS_SHEETS]
    if not pedidos_frames:
        return pd.DataFrame(columns=PEDIDOS_COLUMNAS_MINIMAS)
    return pd.concat(pedidos_frames, ignore_index=True, sort=False)


@st.cache_data(ttl=300)
def cargar_casos_especiales_busqueda():
    data = _leer_registros_hoja_busqueda("casos_especiales")
    df = pd.DataFrame(data)

    columnas_ejemplo = [
        "ID_Pedido", "Hora_Registro", "Vendedor_Registro", "Cliente", "Folio_Factura", "Folio_Factura_Error", "Tipo_Envio",
        "Fecha_Entrega", "Comentario", "Adjuntos", "Estado", "Resultado_Esperado", "Material_Devuelto",
        "Monto_Devuelto", "Motivo_Detallado", "Area_Responsable", "Nombre_Responsable", "Fecha_Completado",
        "Completados_Limpiado", "Estado_Caso", "Hoja_Ruta_Mensajero", "Numero_Cliente_RFC", "Tipo_Envio_Original",
        "Tipo_Caso", "Fecha_Recepcion_Devolucion", "Estado_Recepcion", "Nota_Credito_URL", "Documento_Adicional_URL",
        "Seguimiento", "Comentarios_Admin_Devolucion", "Modificacion_Surtido", "Adjuntos_Surtido", "Refacturacion_Tipo",
        "Refacturacion_Subtipo", "Folio_Factura_Refacturada", "Turno", "Hora_Proceso", "fecha_modificacion", "Fecha_Modificacion",
        "Numero_Serie", "Fecha_Compra", "Comentario", "Comentarios", "Direccion_Guia_Retorno", "Nota_Venta",
        "Tiene_Nota_Venta", "Motivo_NotaVenta"
    ]
    for c in columnas_ejemplo:
        if c not in df.columns:
            df[c] = ""
    return df


# --- TAB 8: SEARCH ORDER ---
with tab8:
    tab8_is_active = default_tab == TAB_INDEX_TAB8
    if tab8_is_active:
        st.session_state["current_tab_index"] = TAB_INDEX_TAB8
    st.subheader("🔍 Buscador de Pedidos por Guía o Cliente")
    if not tab8_is_active:
        st.caption("ℹ️ La búsqueda se habilita al abrir esta pestaña. Si falla la conexión, usa '🔄 Recargar Página y Conexión' arriba.")

    modo_busqueda = st.radio(
        "Selecciona el modo de búsqueda:",
        ["🔢 Por número de guía", "🧑 Por cliente/factura"],
        index=1,
        key="tab_buscar_modo",
    )

    orden_seleccionado = "Más recientes primero"
    recientes_primero = True
    filtrar_por_rango = False
    fecha_inicio_dt = None
    fecha_fin_dt = None
    fecha_inicio_date = None
    fecha_fin_date = None

    if modo_busqueda == "🔢 Por número de guía":
        keyword = st.text_input(
            "📦 Ingresa una palabra clave, número de guía, fragmento o código a buscar:",
            key="tab_buscar_keyword_guia",
        )

        orden_seleccionado = st.selectbox(
            "Orden de los resultados",
            ["Más recientes primero", "Más antiguos primero"],
            index=0,
            key="tab_buscar_orden_resultados_guia",
        )
        recientes_primero = orden_seleccionado == "Más recientes primero"

        filtrar_por_rango = st.checkbox("Filtrar por rango de fechas", value=False, key="tab_buscar_filtrar_rango_guia")
        hoy = date.today()
        inicio_default = hoy - timedelta(days=30)
        rango_fechas_input = st.date_input(
            "Rango de fechas (opcional)",
            value=(inicio_default, hoy),
            format="YYYY-MM-DD",
            disabled=not filtrar_por_rango,
            help="Selecciona una fecha inicial y final para limitar los resultados mostrados.",
            key="tab_buscar_rango_fechas_guia",
        )

        if filtrar_por_rango:
            if isinstance(rango_fechas_input, (list, tuple)):
                if len(rango_fechas_input) == 2:
                    fecha_inicio_date, fecha_fin_date = rango_fechas_input
                elif len(rango_fechas_input) == 1:
                    fecha_inicio_date = fecha_fin_date = rango_fechas_input[0]
            else:
                fecha_inicio_date = fecha_fin_date = rango_fechas_input

            if fecha_inicio_date and fecha_fin_date:
                if fecha_inicio_date > fecha_fin_date:
                    fecha_inicio_date, fecha_fin_date = fecha_fin_date, fecha_inicio_date
                fecha_inicio_dt = datetime.combine(fecha_inicio_date, datetime.min.time())
                fecha_fin_dt = datetime.combine(fecha_fin_date, datetime.max.time())

        buscar_btn = st.button("🔎 Buscar", key="tab_buscar_btn_guia")

    else:
        keyword = st.text_input(
            "🧑 Ingresa el nombre del cliente o folio de factura a buscar:",
            help="Puedes escribir el nombre del cliente o el folio de factura; la búsqueda ignora mayúsculas, acentos y espacios en el folio.",
            key="tab_buscar_keyword_cliente",
        )
        buscar_btn = st.button("🔍 Buscar Pedido del Cliente", key="tab_buscar_btn_cliente")

    filtro_fechas_activo = bool(filtrar_por_rango and fecha_inicio_dt and fecha_fin_dt)

    if "tab_buscar_resultados" not in st.session_state:
        st.session_state["tab_buscar_resultados"] = []
        st.session_state["tab_buscar_modo_last"] = None
        st.session_state["tab_buscar_filtro_fechas_activo"] = False
        st.session_state["tab_buscar_orden_last"] = "Más recientes primero"
        st.session_state["tab_buscar_fecha_inicio_last"] = None
        st.session_state["tab_buscar_fecha_fin_last"] = None

    if buscar_btn:
        if modo_busqueda == "🔢 Por número de guía":
            st.info("🔄 Buscando, por favor espera... puede tardar unos segundos...")

        resultados = []

        df_pedidos = cargar_pedidos_busqueda()
        if "Hora_Registro" in df_pedidos.columns:
            df_pedidos["Hora_Registro"] = pd.to_datetime(df_pedidos["Hora_Registro"], errors="coerce")
            df_pedidos = df_pedidos.sort_values(by="Hora_Registro", ascending=not recientes_primero)
            if filtro_fechas_activo:
                mask_validas = df_pedidos["Hora_Registro"].notna()
                df_pedidos = df_pedidos[mask_validas & df_pedidos["Hora_Registro"].between(fecha_inicio_dt, fecha_fin_dt)]
            df_pedidos = df_pedidos.reset_index(drop=True)

        if modo_busqueda == "🧑 Por cliente/factura":
            if not keyword.strip():
                st.warning("⚠️ Ingresa un nombre de cliente.")
                st.stop()

            keyword_cliente_normalizado = normalizar(keyword.strip())
            keyword_folio_normalizado = normalizar_folio(keyword.strip())

            for _, row in df_pedidos.iterrows():
                nombre = str(row.get("Cliente", "")).strip()
                folio = str(row.get("Folio_Factura", "")).strip()

                nombre_normalizado = normalizar(nombre) if nombre else ""
                folio_normalizado = normalizar_folio(folio)

                coincide_cliente = bool(nombre) and keyword_cliente_normalizado in nombre_normalizado
                coincide_folio = bool(folio_normalizado) and keyword_folio_normalizado == folio_normalizado

                if not coincide_cliente and not coincide_folio:
                    continue

                pedido_id = str(row.get("ID_Pedido", "")).strip()
                if not pedido_id:
                    continue

                prefix = obtener_prefijo_s3(pedido_id)
                todos_los_archivos = obtener_todos_los_archivos(prefix) if prefix else []

                comprobantes = [f for f in todos_los_archivos if "comprobante" in f["Key"].lower()]
                facturas = [f for f in todos_los_archivos if "factura" in f["Key"].lower()]
                otros = [f for f in todos_los_archivos if f not in comprobantes and f not in facturas]

                resultados.append({
                    "__source": "pedidos",
                    "ID_Pedido": pedido_id,
                    "Cliente": row.get("Cliente", ""),
                    "Estado": row.get("Estado", ""),
                    "Vendedor": row.get("Vendedor_Registro", ""),
                    "Folio": row.get("Folio_Factura", ""),
                    "Hora_Registro": row.get("Hora_Registro", ""),
                    "Comentario": str(row.get("Comentario", "")).strip(),
                    "Comentarios": str(row.get("Comentarios", "")).strip(),
                    "Direccion_Guia_Retorno": str(row.get("Direccion_Guia_Retorno", "")).strip(),
                    "Nota_Venta": str(row.get("Nota_Venta", "")).strip(),
                    "Tiene_Nota_Venta": str(row.get("Tiene_Nota_Venta", "")).strip(),
                    "Motivo_NotaVenta": str(row.get("Motivo_NotaVenta", "")).strip(),
                    "Modificacion_Surtido": str(row.get("Modificacion_Surtido", "")).strip(),
                    "Fecha_Modificacion_Surtido": obtener_fecha_modificacion(row),
                    "Adjuntos_Surtido_urls": partir_urls(row.get("Adjuntos_Surtido", "")),
                    "Adjuntos_Guia_urls": partir_urls(row.get("Adjuntos_Guia", "")),
                    "Adjuntos_urls": partir_urls(row.get("Adjuntos", "")),
                    "Refacturacion_Tipo": str(row.get("Refacturacion_Tipo", "")).strip(),
                    "Refacturacion_Subtipo": str(row.get("Refacturacion_Subtipo", "")).strip(),
                    "Folio_Factura_Refacturada": str(row.get("Folio_Factura_Refacturada", "")).strip(),
                    "Coincidentes": [],
                    "Comprobantes": [(f["Key"], get_s3_file_download_url_busqueda_cached(f["Key"])) for f in comprobantes],
                    "Facturas": [(f["Key"], get_s3_file_download_url_busqueda_cached(f["Key"])) for f in facturas],
                    "Otros": [(f["Key"], get_s3_file_download_url_busqueda_cached(f["Key"])) for f in otros],
                })

            df_casos = cargar_casos_especiales_busqueda()
            if "Hora_Registro" in df_casos.columns:
                df_casos["Hora_Registro"] = pd.to_datetime(df_casos["Hora_Registro"], errors="coerce")
                df_casos = df_casos.sort_values(by="Hora_Registro", ascending=not recientes_primero)
                if filtro_fechas_activo:
                    mask_validas_casos = df_casos["Hora_Registro"].notna()
                    df_casos = df_casos[mask_validas_casos & df_casos["Hora_Registro"].between(fecha_inicio_dt, fecha_fin_dt)]
                df_casos = df_casos.reset_index(drop=True)

            for _, row in df_casos.iterrows():
                nombre = str(row.get("Cliente", "")).strip()
                folio = str(row.get("Folio_Factura", "")).strip()

                nombre_normalizado = normalizar(nombre) if nombre else ""
                folio_normalizado = normalizar_folio(folio)

                coincide_cliente = bool(nombre) and keyword_cliente_normalizado in nombre_normalizado
                coincide_folio = bool(folio_normalizado) and keyword_folio_normalizado == folio_normalizado

                if not coincide_cliente and not coincide_folio:
                    continue
                resultados.append(preparar_resultado_caso_busqueda(row))

        elif modo_busqueda == "🔢 Por número de guía":
            clave = keyword.strip()
            if not clave:
                st.warning("⚠️ Ingresa una palabra clave o número de guía.")
                st.stop()

            for _, row in df_pedidos.iterrows():
                pedido_id = str(row.get("ID_Pedido", "")).strip()
                if not pedido_id:
                    continue

                prefix = obtener_prefijo_s3(pedido_id)
                if not prefix:
                    continue

                archivos_validos = obtener_archivos_pdf_validos(prefix)
                archivos_coincidentes = []

                for archivo in archivos_validos:
                    key = archivo["Key"]
                    texto = extraer_texto_pdf(key)

                    clave_sin_espacios = clave.replace(" ", "")
                    texto_limpio = texto.replace(" ", "").replace("\n", "")

                    coincide = (
                        clave in texto
                        or clave_sin_espacios in texto_limpio
                        or re.search(re.escape(clave), texto_limpio)
                        or re.search(re.escape(clave_sin_espacios), texto_limpio)
                    )

                    if coincide:
                        waybill_match = re.search(r"WAYBILL[\s:]*([0-9 ]{8,})", texto, re.IGNORECASE)
                        if waybill_match:
                            st.code(f"📦 WAYBILL detectado: {waybill_match.group(1)}")

                        archivos_coincidentes.append((key, get_s3_file_download_url_busqueda_cached(key)))
                        todos_los_archivos = obtener_todos_los_archivos(prefix)
                        comprobantes = [f for f in todos_los_archivos if "comprobante" in f["Key"].lower()]
                        facturas = [f for f in todos_los_archivos if "factura" in f["Key"].lower()]
                        otros = [f for f in todos_los_archivos if f not in comprobantes and f not in facturas and f["Key"] != archivos_coincidentes[0][0]]

                        resultados.append({
                            "__source": "pedidos",
                            "ID_Pedido": pedido_id,
                            "Cliente": row.get("Cliente", ""),
                            "Estado": row.get("Estado", ""),
                            "Vendedor": row.get("Vendedor_Registro", ""),
                            "Folio": row.get("Folio_Factura", ""),
                            "Hora_Registro": row.get("Hora_Registro", ""),
                            "Comentario": str(row.get("Comentario", "")).strip(),
                            "Comentarios": str(row.get("Comentarios", "")).strip(),
                            "Direccion_Guia_Retorno": str(row.get("Direccion_Guia_Retorno", "")).strip(),
                            "Nota_Venta": str(row.get("Nota_Venta", "")).strip(),
                            "Tiene_Nota_Venta": str(row.get("Tiene_Nota_Venta", "")).strip(),
                            "Motivo_NotaVenta": str(row.get("Motivo_NotaVenta", "")).strip(),
                            "Modificacion_Surtido": str(row.get("Modificacion_Surtido", "")).strip(),
                            "Fecha_Modificacion_Surtido": obtener_fecha_modificacion(row),
                            "Adjuntos_Surtido_urls": partir_urls(row.get("Adjuntos_Surtido", "")),
                            "Adjuntos_Guia_urls": partir_urls(row.get("Adjuntos_Guia", "")),
                            "Adjuntos_urls": partir_urls(row.get("Adjuntos", "")),
                            "Refacturacion_Tipo": str(row.get("Refacturacion_Tipo", "")).strip(),
                            "Refacturacion_Subtipo": str(row.get("Refacturacion_Subtipo", "")).strip(),
                            "Folio_Factura_Refacturada": str(row.get("Folio_Factura_Refacturada", "")).strip(),
                            "Coincidentes": archivos_coincidentes,
                            "Comprobantes": [(f["Key"], get_s3_file_download_url_busqueda_cached(f["Key"])) for f in comprobantes],
                            "Facturas": [(f["Key"], get_s3_file_download_url_busqueda_cached(f["Key"])) for f in facturas],
                            "Otros": [(f["Key"], get_s3_file_download_url_busqueda_cached(f["Key"])) for f in otros],
                        })
                        break

                if archivos_coincidentes:
                    break

        st.session_state["tab_buscar_resultados"] = resultados
        st.session_state["tab_buscar_modo_last"] = modo_busqueda
        st.session_state["tab_buscar_filtro_fechas_activo"] = filtro_fechas_activo
        st.session_state["tab_buscar_orden_last"] = orden_seleccionado
        st.session_state["tab_buscar_fecha_inicio_last"] = fecha_inicio_date
        st.session_state["tab_buscar_fecha_fin_last"] = fecha_fin_date

    resultados = st.session_state.get("tab_buscar_resultados", [])
    modo_busqueda_render = st.session_state.get("tab_buscar_modo_last") or modo_busqueda
    filtro_fechas_activo_render = bool(st.session_state.get("tab_buscar_filtro_fechas_activo", False))
    orden_render = st.session_state.get("tab_buscar_orden_last", orden_seleccionado)
    fecha_inicio_render = st.session_state.get("tab_buscar_fecha_inicio_last")
    fecha_fin_render = st.session_state.get("tab_buscar_fecha_fin_last")

    if buscar_btn or resultados:
        st.markdown("---")
        if resultados:
            mensaje_exito = f"✅ Se encontraron coincidencias en {len(resultados)} registro(s)."
            if filtro_fechas_activo_render:
                mensaje_exito += " (Filtro temporal aplicado)"
            st.success(mensaje_exito)

            detalles_filtros = [f"Orden: {orden_render}"]
            if filtro_fechas_activo_render and fecha_inicio_render and fecha_fin_render:
                detalles_filtros.append(f"Rango: {fecha_inicio_render.strftime('%Y-%m-%d')} → {fecha_fin_render.strftime('%Y-%m-%d')}")
            if detalles_filtros:
                st.caption(" | ".join(detalles_filtros))

            def _parse_dt(v):
                try:
                    return pd.to_datetime(v)
                except Exception:
                    return pd.NaT

            resultados = sorted(resultados, key=lambda r: _parse_dt(r.get("Hora_Registro")), reverse=recientes_primero)

            for res in resultados:
                if res.get("__source") == "casos":
                    render_caso_especial_busqueda(res)
                else:
                    st.markdown(f"### 🤝 {res['Cliente'] or 'Cliente N/D'}")
                    st.markdown(
                        f"📄 **Folio:** `{res['Folio'] or 'N/D'}`  |  🔍 **Estado:** `{res['Estado'] or 'N/D'}`  |  🧑‍💼 **Vendedor:** `{res['Vendedor'] or 'N/D'}`  |  🕒 **Hora:** `{res['Hora_Registro'] or 'N/D'}`"
                    )

                    comentario_txt = str(res.get("Comentario", "") or res.get("Comentarios", "")).strip()
                    if comentario_txt:
                        st.markdown("#### 📝 Comentarios del pedido")
                        st.info(comentario_txt)

                    direccion_retorno = str(res.get("Direccion_Guia_Retorno", "")).strip()
                    if direccion_retorno:
                        st.markdown("#### 📍 Dirección para guía de retorno")
                        st.info(direccion_retorno)

                    nota_venta_valor = str(res.get("Nota_Venta", "")).strip()
                    tiene_nota_venta = str(res.get("Tiene_Nota_Venta", "")).strip()
                    motivo_nota_venta = str(res.get("Motivo_NotaVenta", "")).strip()
                    if nota_venta_valor or tiene_nota_venta or motivo_nota_venta:
                        st.markdown("#### 🧾 Nota de venta")
                        estado_texto = tiene_nota_venta or ("Sí" if nota_venta_valor else "No")
                        st.markdown(f"- **¿Tiene nota de venta?:** {estado_texto}")
                        if nota_venta_valor:
                            st.markdown(f"- **Detalle:** {nota_venta_valor}")
                        if motivo_nota_venta:
                            st.markdown(f"- **Motivo:** {motivo_nota_venta}")

                    mod_txt = res.get("Modificacion_Surtido", "") or ""
                    mod_fecha = res.get("Fecha_Modificacion_Surtido", "") or ""
                    mod_urls = res.get("Adjuntos_Surtido_urls", []) or []
                    if mod_txt or mod_urls:
                        st.markdown("#### 🛠 Modificación de surtido")
                        if mod_fecha:
                            st.caption(f"📅 Fecha de modificación: {mod_fecha}")
                        if mod_txt:
                            st.info(mod_txt)
                        if mod_urls:
                            st.markdown("**Archivos de modificación:**")
                            for u in mod_urls:
                                nombre = extract_s3_key_busqueda(u).split("/")[-1]
                                tmp = get_s3_file_download_url_busqueda_cached(u)
                                st.markdown(f'- <a href="{tmp}" target="_blank">{nombre}</a>', unsafe_allow_html=True)

                    ref_t = res.get("Refacturacion_Tipo", "")
                    ref_st = res.get("Refacturacion_Subtipo", "")
                    ref_f = res.get("Folio_Factura_Refacturada", "")
                    if any([ref_t, ref_st, ref_f]):
                        with st.expander("♻️ Refacturación", expanded=False):
                            st.markdown(f"- **Tipo:** {ref_t or 'N/A'}")
                            st.markdown(f"- **Subtipo:** {ref_st or 'N/A'}")
                            st.markdown(f"- **Folio refacturado:** {ref_f or 'N/A'}")

                    with st.expander("📁 Archivos del Pedido", expanded=True):
                        guia_hoja = res.get("Adjuntos_Guia_urls") or []
                        if guia_hoja:
                            st.markdown("#### 🧾 Guías registradas en la hoja:")
                            for idx, raw_url in enumerate(guia_hoja, start=1):
                                nombre, enlace = resolver_nombre_y_enlace_busqueda(raw_url, f"Guía hoja #{idx}")
                                if not enlace:
                                    continue
                                st.markdown(f'- <a href="{enlace}" target="_blank">🧾 {nombre} (hoja)</a>', unsafe_allow_html=True)

                        if res.get("Coincidentes"):
                            st.markdown("#### 🔍 Guías detectadas en S3:")
                            for key, url in res["Coincidentes"]:
                                nombre = key.split("/")[-1]
                                st.markdown(f'- <a href="{url}" target="_blank">🔍 {nombre}</a>', unsafe_allow_html=True)
                        if res.get("Comprobantes"):
                            st.markdown("#### 🧾 Comprobantes:")
                            for key, url in res["Comprobantes"]:
                                nombre = key.split("/")[-1]
                                st.markdown(f'- <a href="{url}" target="_blank">📄 {nombre}</a>', unsafe_allow_html=True)
                        if res.get("Facturas"):
                            st.markdown("#### 📁 Facturas:")
                            for key, url in res["Facturas"]:
                                nombre = key.split("/")[-1]
                                st.markdown(f'- <a href="{url}" target="_blank">📄 {nombre}</a>', unsafe_allow_html=True)
                        adjuntos_hoja = res.get("Adjuntos_urls") or []
                        otros_s3 = res.get("Otros") or []
                        otros_items = []
                        claves_vistas = set()

                        def _normalizar_clave(valor):
                            if not valor:
                                return None
                            valor_str = str(valor).strip()
                            if not valor_str:
                                return None
                            return valor_str.lower()

                        def _registrar_clave(valor):
                            clave_norm = _normalizar_clave(valor)
                            if clave_norm:
                                claves_vistas.add(clave_norm)

                        def _esta_registrada(valor):
                            clave_norm = _normalizar_clave(valor)
                            if not clave_norm:
                                return False
                            return clave_norm in claves_vistas

                        for raw_url in guia_hoja:
                            clave = extract_s3_key_busqueda(raw_url) or raw_url
                            _registrar_clave(clave)
                            _registrar_clave(raw_url)

                        for key, url in res.get("Coincidentes") or []:
                            clave = extract_s3_key_busqueda(key) or key
                            _registrar_clave(clave)
                            if url:
                                _registrar_clave(extract_s3_key_busqueda(url) or url)

                        for key, url in res.get("Comprobantes") or []:
                            clave = extract_s3_key_busqueda(key) or key
                            _registrar_clave(clave)
                            if url:
                                _registrar_clave(extract_s3_key_busqueda(url) or url)

                        for key, url in res.get("Facturas") or []:
                            clave = extract_s3_key_busqueda(key) or key
                            _registrar_clave(clave)
                            if url:
                                _registrar_clave(extract_s3_key_busqueda(url) or url)

                        for key, url in otros_s3:
                            clave = extract_s3_key_busqueda(key) or key or url
                            if _esta_registrada(clave) or _esta_registrada(url):
                                continue
                            nombre = key.split("/")[-1] if key else "Archivo"
                            otros_items.append((nombre, url))
                            _registrar_clave(clave)
                            if url:
                                _registrar_clave(url)

                        for idx, raw_url in enumerate(adjuntos_hoja, start=1):
                            nombre, enlace = resolver_nombre_y_enlace_busqueda(raw_url, f"Adjunto hoja #{idx}")
                            if not enlace:
                                continue
                            clave = extract_s3_key_busqueda(raw_url) or enlace
                            if _esta_registrada(clave):
                                continue
                            otros_items.append((nombre or f"Adjunto hoja #{idx}", enlace))
                            _registrar_clave(clave)

                        if otros_items:
                            st.markdown("#### 📂 Otros Archivos:")
                            for nombre, enlace in otros_items:
                                st.markdown(f'- <a href="{enlace}" target="_blank">📌 {nombre}</a>', unsafe_allow_html=True)

        else:
            mensaje = (
                "⚠️ No se encontraron coincidencias en ningún archivo PDF."
                if modo_busqueda_render == "🔢 Por número de guía"
                else "⚠️ No se encontraron pedidos o casos para el cliente ingresado."
            )
            if filtro_fechas_activo_render:
                mensaje += " Revisa el rango de fechas seleccionado."
            st.warning(mensaje)
